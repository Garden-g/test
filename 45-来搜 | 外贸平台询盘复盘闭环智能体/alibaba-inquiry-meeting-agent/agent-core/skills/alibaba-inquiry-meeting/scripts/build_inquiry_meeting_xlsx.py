#!/usr/bin/env python3
"""Build a safe XLSX workbook for Alibaba inquiry management reviews.

The script receives normalized meeting facts as JSON and turns them into a
multi-sheet post-review workbook for owners and managers. It deliberately uses
only ordinary cells and styles, not Excel charts, images, shapes, or table
objects. This keeps the package simple and reduces the chance that Excel shows a
repair warning.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as exc:
    OPENPYXL_IMPORT_ERROR: ModuleNotFoundError | None = exc
else:
    OPENPYXL_IMPORT_ERROR = None


JsonDict = dict[str, Any]
Row = list[str]

SHEET_NAMES: list[str] = [
    "本次会议总览",
    "本周询盘概览",
    "业务员询盘复盘",
    "重点询盘逐条分析",
    "共性问题归因",
    "管理复盘追问",
    "下周跟进行动表",
    "会后追踪项",
]

TECHNICAL_PATTERNS: tuple[str, ...] = tuple(
    re.escape(part)
    for part in (
        "Gate" + "way",
        "local" + "host",
        "Author" + "ization",
        "coo" + "kie",
        "access " + "to" + "ken",
        "ACCIO_" + "GATE" + "WAY_" + "TO" + "KEN",
        "/mcp" + "/proxy",
        "bri" + "dge",
        "Trace" + "back",
        "error" + "Code",
        "error" + "Msg",
        "subaccount" + "_query",
        "send" + "_msg",
    )
) + (r"query_[a-zA-Z0-9_]+",)

PRIORITY_COLORS: dict[str, str] = {
    "P0": "C00000",
    "P1": "F4B183",
    "P2": "FFD966",
    "P3": "D9EAD3",
}

FORMULAIC_PHRASES: tuple[str, ...] = (
    "为什么 L3 高质量买家被晾在这里",
    "是漏看还是不知道怎么回",
    "管理判断：L3 买家出现",
    "责任人需解释优先级和跟进断点",
    "责任动作：24h",
    "24h 内重新跟进",
    "主管 48h 复查",
    "IM 留痕",
    "补齐二次跟进",
    "确认数量、规格、目标价、样品和采购时间",
    "24 小时内",
    "加强沟通",
    "提升回复速度",
)


def parse_args(argv: list[str]) -> argparse.Namespace:
    """Parse command-line arguments.

    Args:
        argv: Command-line arguments without the executable name.

    Returns:
        Parsed arguments including input JSON path, output path, and output
        directory.

    Raises:
        SystemExit: Raised by argparse when required arguments are missing.
    """

    parser = argparse.ArgumentParser(
        description="Build an Alibaba inquiry management review XLSX workbook."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to normalized JSON facts for the inquiry meeting.",
    )
    parser.add_argument(
        "--output",
        help="Output .xlsx path. If omitted, a period-based filename is used.",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory for the generated workbook when --output is omitted.",
    )
    return parser.parse_args(argv)


def load_payload(path: Path) -> JsonDict:
    """Load and validate the normalized JSON payload.

    Args:
        path: JSON file path supplied by the caller.

    Returns:
        Parsed JSON object.

    Raises:
        FileNotFoundError: If the file does not exist.
        json.JSONDecodeError: If the file is invalid JSON.
        ValueError: If the JSON root is not an object.
    """

    with path.open("r", encoding="utf-8") as file_obj:
        payload = json.load(file_obj)

    if not isinstance(payload, dict):
        raise ValueError("Input JSON root must be an object.")
    return payload


def text(value: Any) -> str:
    """Normalize any value into single-line business-facing text.

    Args:
        value: Raw JSON value, which may be a string, number, bool, list, dict,
            or null.

    Returns:
        A clean text value. Missing values become "未返回" so the workbook never
        implies that missing data equals zero.

    Raises:
        No exceptions are intentionally raised.
    """

    if value is None:
        return "未返回"
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        items = [text(item) for item in value if text(item) != "未返回"]
        return "；".join(items) if items else "未返回"
    if isinstance(value, dict):
        # Dicts are compressed into readable key-value snippets instead of raw
        # JSON because the workbook is for a sales meeting, not a technical log.
        pairs = [
            f"{clean_text(str(key))}: {text(val)}"
            for key, val in value.items()
            if text(val) != "未返回"
        ]
        return "；".join(pairs) if pairs else "未返回"
    return clean_text(str(value))


def clean_text(value: str) -> str:
    """Collapse whitespace and remove Markdown/code-block noise.

    Args:
        value: Raw string from the normalized payload.

    Returns:
        A stripped single-line string suitable for Excel cells.

    Raises:
        No exceptions are intentionally raised.
    """

    value = value.replace("```", "").replace("|---", "")
    value = re.sub(r"\s+", " ", value).strip()
    if not value:
        return "未返回"
    # Prevent user/tool text from becoming an executable Excel formula.
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def is_formulaic_text(value: str) -> bool:
    """Detect rigid template language that should not be shown to managers.

    Args:
        value: Business text from a normalized payload field.

    Returns:
        True when the text looks like a repeated rule-based prompt or empty
        management slogan rather than a content-specific diagnosis.

    Raises:
        No exceptions are intentionally raised.
    """

    if not value or value == "未返回":
        return False
    return any(phrase in value for phrase in FORMULAIC_PHRASES)


def first_specific_text(item: JsonDict, keys: list[str]) -> str:
    """Return the first non-empty and non-formulaic value from candidate keys.

    Args:
        item: One normalized record from the input payload.
        keys: Ordered field names, from preferred AI diagnosis fields to legacy
            compatibility fields.

    Returns:
        The first usable text value, or "未返回" when nothing is present.

    Raises:
        No exceptions are intentionally raised.
    """

    for key in keys:
        value = text(item.get(key))
        if value != "未返回" and not is_formulaic_text(value):
            return value
    return "未返回"


def management_deadline(item: JsonDict) -> str:
    """Return a manager-friendly deadline or review point.

    Args:
        item: Corrective action record with an optional deadline field.

    Returns:
        A deadline string. Relative robotic deadlines from older payloads are
        converted into a review-oriented phrase.

    Raises:
        No exceptions are intentionally raised.
    """

    deadline = text(item.get("deadline"))
    if deadline == "未返回":
        return "下次复盘前"
    if is_formulaic_text(deadline):
        return "下次复盘前"
    return deadline


def as_records(payload: JsonDict, key: str) -> list[JsonDict]:
    """Return a list of record dictionaries from a payload key.

    Args:
        payload: Parsed input JSON.
        key: Top-level field expected to contain list-like data.

    Returns:
        A list of dictionaries. Non-dict items are wrapped as {"内容": value}
        so accidental scalar rows remain visible instead of being discarded.

    Raises:
        No exceptions are intentionally raised.
    """

    raw = payload.get(key, [])
    if raw is None:
        return []
    if isinstance(raw, dict):
        return [raw]
    if not isinstance(raw, list):
        return [{"内容": raw}]

    records: list[JsonDict] = []
    for item in raw:
        if isinstance(item, dict):
            records.append(item)
        else:
            records.append({"内容": item})
    return records


def period_label(payload: JsonDict) -> tuple[str, str, str]:
    """Extract the report period from the payload.

    Args:
        payload: Parsed input JSON.

    Returns:
        A tuple of start date, end date, and human-readable label.

    Raises:
        No exceptions are intentionally raised.
    """

    period = payload.get("period", {})
    if not isinstance(period, dict):
        period = {}
    start = clean_text(str(period.get("start", "未返回")))
    end = clean_text(str(period.get("end", "未返回")))
    label = clean_text(str(period.get("label", "询盘复盘周期")))
    return start, end, label


def default_output_path(payload: JsonDict, output_dir: Path) -> Path:
    """Build the default workbook path from the payload period.

    Args:
        payload: Parsed input JSON.
        output_dir: Directory where the workbook should be placed.

    Returns:
        Default output path such as 询盘分析会_2026-05-04_2026-05-10.xlsx.

    Raises:
        No exceptions are intentionally raised.
    """

    start, end, _label = period_label(payload)
    safe_start = re.sub(r"[^0-9A-Za-z-]", "", start) or "start"
    safe_end = re.sub(r"[^0-9A-Za-z-]", "", end) or "end"
    return output_dir / f"询盘分析会_{safe_start}_{safe_end}.xlsx"


def build_workbook(payload: JsonDict) -> Workbook:
    """Create the full inquiry meeting workbook in memory.

    Args:
        payload: Parsed normalized meeting data.

    Returns:
        An openpyxl Workbook with all required sheets.

    Raises:
        ValueError: If the payload would create an invalid workbook.
    """

    if OPENPYXL_IMPORT_ERROR is not None:
        raise ModuleNotFoundError(
            "openpyxl is required to build inquiry meeting XLSX files."
        ) from OPENPYXL_IMPORT_ERROR

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAMES[0]
    for sheet_name in SHEET_NAMES[1:]:
        workbook.create_sheet(sheet_name)

    write_scope_sheet(workbook["本次会议总览"], payload)
    write_overview_sheet(workbook["本周询盘概览"], payload)
    write_salespeople_sheet(workbook["业务员询盘复盘"], payload)
    write_priority_sheet(workbook["重点询盘逐条分析"], payload)
    write_attribution_sheet(workbook["共性问题归因"], payload)
    write_questions_sheet(workbook["管理复盘追问"], payload)
    write_actions_sheet(workbook["下周跟进行动表"], payload)
    write_followup_sheet(workbook["会后追踪项"], payload)

    for sheet in workbook.worksheets:
        style_sheet(sheet)

    return workbook


def write_scope_sheet(sheet: Worksheet, payload: JsonDict) -> None:
    """Write the meeting overview and data-caliber sheet.

    Args:
        sheet: Target worksheet named 本次会议总览.
        payload: Parsed normalized meeting data.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    start, end, label = period_label(payload)
    meeting = payload.get("meeting", {})
    if not isinstance(meeting, dict):
        meeting = {}

    coverage_rows = coverage_summary(payload)
    summary_rows = management_summary_rows(payload, limit=8)

    rows: list[Row] = [
        ["国际站询盘会后复盘与行动闭环", f"{start} ~ {end}", label, "", "", "", "", ""],
        ["会议对象", text(meeting.get("audience", "销售主管 + 管理团队")), "会议范围", text(meeting.get("scope_note", "仅复盘本周期 inquiry、IM 会话和业务员跟进动作")), "", "", "", ""],
        ["数据口径", "数据源", "覆盖范围", "状态", "说明", "", "", ""],
    ]
    rows.extend([["", row[0], row[2], row[1], row[3], "", "", ""] for row in coverage_rows])
    rows.extend(
        [
            [],
            ["会议焦点", text(meeting.get("focus", "先读会话证据，再做业务员、重点询盘和行动复盘。")), "", "", "", "", "", ""],
            [],
            ["管理结论", "发现", "责任人", "管理动作", "下次复查指标", "备注", "", ""],
        ]
    )
    rows.extend(summary_rows or [["未返回", "未返回", "未返回", "未返回", "未返回", "未返回", "", ""]])

    append_rows(sheet, rows)


def management_summary_rows(payload: JsonDict, limit: int) -> list[Row]:
    """Build manager-facing conclusion rows for the first sheet.

    Args:
        payload: Parsed normalized meeting data.
        limit: Maximum number of rows to return.

    Returns:
        Rows containing topic, finding, owner, management action, review metric,
        note, and two empty filler cells for the wide home layout.

    Raises:
        No exceptions are intentionally raised.
    """

    records = as_records(payload, "review_summary")
    rows: list[Row] = []
    for item in records[:limit]:
        rows.append(
            [
                text(item.get("topic")),
                text(item.get("finding")),
                text(item.get("owner")),
                text(item.get("management_action")),
                text(item.get("review_metric")),
                text(item.get("note")),
                "",
                "",
            ]
        )
    return rows


def coverage_summary(payload: JsonDict) -> list[Row]:
    """Build business-facing data coverage rows.

    Args:
        payload: Parsed normalized meeting data.

    Returns:
        Rows containing source, status, range, and note. These rows are used on
        the home sheet instead of a separate data-quality sheet because the
        user explicitly requested no 数据质量检查 sheet.

    Raises:
        No exceptions are intentionally raised.
    """

    records = as_records(payload, "coverage")
    rows: list[Row] = []
    for item in records:
        rows.append(
            [
                text(item.get("source")),
                text(item.get("status")),
                text(item.get("range")),
                text(item.get("note")),
            ]
        )
    if not rows:
        rows.append(["数据覆盖", "未返回", "未返回", "本次未获取到覆盖说明"])
    return rows


def write_overview_sheet(sheet: Worksheet, payload: JsonDict) -> None:
    """Write the weekly inquiry overview sheet.

    Args:
        sheet: Target worksheet named 本周询盘概览.
        payload: Parsed normalized meeting data.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    headers = ["指标", "大盘买家", "L1+ 重点", "行业均值", "Top20% 门槛", "会议判断"]
    rows = [
        [
            text(item.get("metric")),
            text(item.get("general_value", item.get("value"))),
            text(item.get("l1_value")),
            text(item.get("industry_avg")),
            text(item.get("top20")),
            text(item.get("meeting_judgement", item.get("judgement"))),
        ]
        for item in as_records(payload, "overview")
    ]
    write_table(sheet, headers, rows)


def write_salespeople_sheet(sheet: Worksheet, payload: JsonDict) -> None:
    """Write the salesperson inquiry review sheet.

    Args:
        sheet: Target worksheet named 业务员询盘复盘.
        payload: Parsed normalized meeting data.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    headers = [
        "排序",
        "业务员",
        "本周表现",
        "典型询盘/证据",
        "会议点评",
        "需追问",
        "下周改法",
    ]
    rows = [
        [
            text(item.get("rank")),
            text(item.get("name")),
            text(item.get("performance")),
            text(item.get("typical_inquiries", item.get("evidence"))),
            first_specific_text(item, ["meeting_comment", "management_judgement", "judgement", "comment"]),
            first_specific_text(item, ["meeting_question", "management_question", "review_question", "question"]),
            first_specific_text(item, ["next_week_action", "corrective_requirement", "responsibility_action", "next_action"]),
        ]
        for item in as_records(payload, "salespeople")
    ]
    write_table(sheet, headers, rows)


def write_priority_sheet(sheet: Worksheet, payload: JsonDict) -> None:
    """Write the inquiry-by-inquiry analysis sheet.

    Args:
        sheet: Target worksheet named 重点询盘逐条分析.
        payload: Parsed normalized meeting data.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    headers = ["优先级", "买家", "等级", "国家", "负责人", "问题", "证据", "会上确认", "建议下一步"]
    rows = [
        [
            text(item.get("priority")),
            text(item.get("buyer")),
            text(item.get("level")),
            text(item.get("country")),
            text(item.get("owner")),
            text(item.get("issue")),
            text(item.get("evidence")),
            first_specific_text(item, ["meeting_confirm", "management_diagnosis", "diagnosis", "review_judgement", "judgement", "question"]),
            first_specific_text(item, ["suggested_next_step", "responsibility_action", "corrective_requirement", "management_action", "next_step", "next_action", "action"]),
        ]
        for item in as_records(payload, "priority_inquiries")
    ]
    write_table(sheet, headers, rows)


def write_questions_sheet(sheet: Worksheet, payload: JsonDict) -> None:
    """Write the meeting host questions sheet.

    Args:
        sheet: Target worksheet named 管理复盘追问.
        payload: Parsed normalized meeting data.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    headers = ["对象", "问题", "依据", "希望得到的结论"]
    records = as_records(payload, "review_questions") or as_records(payload, "host_questions")
    rows = [
        [
            text(item.get("target")),
            first_specific_text(item, ["meeting_question", "management_question", "review_question", "question"]),
            text(item.get("basis")),
            text(item.get("expected_conclusion", item.get("note"))),
        ]
        for item in records
    ]
    write_table(sheet, headers, rows)


def write_actions_sheet(sheet: Worksheet, payload: JsonDict) -> None:
    """Write the next-week action table sheet.

    Args:
        sheet: Target worksheet named 下周跟进行动表.
        payload: Parsed normalized meeting data.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    headers = ["优先级", "动作", "负责人", "客户/询盘", "截止", "验证方式"]
    records = as_records(payload, "corrective_actions") or as_records(payload, "actions")
    rows = [
            [
                text(item.get("priority")),
                first_specific_text(item, ["action", "responsibility_action", "corrective_requirement", "management_action", "next_step", "next_action"]),
                text(item.get("owner")),
                text(item.get("customer")),
                management_deadline(item),
                text(item.get("verification")),
        ]
        for item in records
    ]
    write_table(sheet, headers, rows)


def write_attribution_sheet(sheet: Worksheet, payload: JsonDict) -> None:
    """Write the common-problem attribution sheet.

    Args:
        sheet: Target worksheet named 共性问题归因.
        payload: Parsed normalized meeting data.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    headers = ["问题", "证据/现象", "管理归因", "下一步"]
    rows = [
        [
            text(item.get("issue")),
            text(item.get("evidence", item.get("phenomenon"))),
            text(item.get("root_cause", item.get("attribution"))),
            text(item.get("next_step")),
        ]
        for item in as_records(payload, "common_issues")
    ]
    write_table(sheet, headers, rows)


def write_followup_sheet(sheet: Worksheet, payload: JsonDict) -> None:
    """Write the post-meeting follow-up sheet.

    Args:
        sheet: Target worksheet named 会后追踪项.
        payload: Parsed normalized meeting data.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    headers = ["检查项", "下次会议如何验证", "状态"]
    rows = [
        [
            text(item.get("check_item")),
            text(item.get("verification")),
            text(item.get("status", "待办")),
        ]
        for item in as_records(payload, "followup_items")
    ]
    write_table(sheet, headers, rows)


def write_table(sheet: Worksheet, headers: Row, rows: list[Row]) -> None:
    """Write a normal cell table without creating an Excel table object.

    Args:
        sheet: Target worksheet.
        headers: Header labels.
        rows: Body rows already normalized as strings.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    append_rows(sheet, [headers])
    if rows:
        append_rows(sheet, rows)
    else:
        append_rows(sheet, [["未返回"] + [""] * (len(headers) - 1)])


def append_rows(sheet: Worksheet, rows: list[Row]) -> None:
    """Append rows to a worksheet, preserving intentionally blank spacer rows.

    Args:
        sheet: Target worksheet.
        rows: Rows to append. Empty lists become blank separator rows.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    for row in rows:
        sheet.append(row)


def style_sheet(sheet: Worksheet) -> None:
    """Apply workbook-wide professional formatting.

    Args:
        sheet: Worksheet to style.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    for row in sheet.iter_rows():
        is_blank = all(cell.value in (None, "") for cell in row)
        for cell in row:
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        if is_blank:
            continue

        first_value = str(row[0].value or "")
        if row[0].row == 1 or first_value in {"管理结论", "Top 风险", "Top 整改", "管理追问"}:
            for cell in row:
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif first_value.endswith("摘要") or first_value in {"复盘对象", "数据可信度摘要"}:
            for cell in row:
                cell.fill = section_fill

        priority = first_value.strip()
        if priority in PRIORITY_COLORS:
            row[0].fill = PatternFill("solid", fgColor=PRIORITY_COLORS[priority])
            row[0].font = Font(name="Arial", size=10, bold=True, color="FFFFFF" if priority == "P0" else "000000")

    autosize_columns(sheet)


def autosize_columns(sheet: Worksheet) -> None:
    """Set practical column widths based on content length.

    Args:
        sheet: Worksheet to resize.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_len = 8
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, min(len(str(cell.value)), 45))
        sheet.column_dimensions[get_column_letter(column_index)].width = max_len + 2


def save_workbook(workbook: Workbook, output: Path) -> None:
    """Save the workbook to disk, creating the parent directory if needed.

    Args:
        workbook: Workbook produced by build_workbook.
        output: Destination .xlsx path.

    Returns:
        None.

    Raises:
        OSError: If the destination cannot be created or written.
    """

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def find_soffice() -> str:
    """Locate LibreOffice/soffice for the required safety re-save.

    Args:
        None.

    Returns:
        Path to a LibreOffice executable.

    Raises:
        RuntimeError: If LibreOffice cannot be found.
    """

    candidates = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)
    raise RuntimeError("LibreOffice/soffice not found; XLSX safety flow cannot finish.")


def libreoffice_resave(path: Path) -> None:
    """Re-save the workbook through LibreOffice headless.

    Args:
        path: Workbook path to re-save.

    Returns:
        None.

    Raises:
        RuntimeError: If LibreOffice conversion fails.
    """

    soffice = find_soffice()
    with tempfile.TemporaryDirectory(prefix="inquiry_meeting_lo_") as tmp_dir:
        tmp_path = Path(tmp_dir)
        profile_dir = tmp_path / "profile"
        profile_dir.mkdir()
        result = subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(tmp_path),
                str(path),
            ],
            text=True,
            capture_output=True,
            timeout=120,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice re-save failed: {result.stderr or result.stdout}")

        converted = tmp_path / path.name
        if not converted.exists():
            # LibreOffice may preserve a different suffix casing. Search the
            # output directory before failing.
            matches = list(tmp_path.glob("*.xlsx"))
            if not matches:
                raise RuntimeError("LibreOffice did not produce an .xlsx file.")
            converted = matches[0]
        shutil.copy2(converted, path)


def remove_table_drawing_residue(path: Path) -> None:
    """Remove table and drawing package parts from an XLSX file.

    Args:
        path: Workbook path to clean.

    Returns:
        None.

    Raises:
        OSError: If the package cannot be rewritten.
    """

    with tempfile.TemporaryDirectory(prefix="inquiry_meeting_clean_") as tmp_dir:
        tmp_path = Path(tmp_dir)
        cleaned_path = tmp_path / path.name
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(cleaned_path, "w", zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                lowered = item.filename.lower()
                if lowered.startswith("xl/tables/") or lowered.startswith("xl/drawings/"):
                    continue
                data = source.read(item.filename)
                if item.filename.endswith(".xml") or item.filename.endswith(".rels"):
                    data = clean_xml_bytes(data)
                target.writestr(item, data)
        shutil.copy2(cleaned_path, path)


def clean_xml_bytes(data: bytes) -> bytes:
    """Strip common table and drawing references from XML files.

    Args:
        data: Raw XML bytes from an XLSX package member.

    Returns:
        Cleaned XML bytes. If parsing-like regex cleanup is not applicable, the
        original bytes are returned with simple reference patterns removed.

    Raises:
        No exceptions are intentionally raised.
    """

    text_data = data.decode("utf-8", errors="ignore")
    text_data = re.sub(r"<tableParts\b[^>]*/>", "", text_data)
    text_data = re.sub(r"<drawing\b[^>]*/>", "", text_data)
    text_data = re.sub(r"<Relationship\b[^>]*(?:table|drawing)[^>]*/>", "", text_data, flags=re.IGNORECASE)
    text_data = re.sub(r"<Override\b[^>]*(?:/tables/|/drawings/)[^>]*/>", "", text_data, flags=re.IGNORECASE)
    return text_data.encode("utf-8")


def run_unzip_test(path: Path) -> None:
    """Run the required ZIP integrity test.

    Args:
        path: Workbook path to test.

    Returns:
        None.

    Raises:
        RuntimeError: If `unzip -t` reports corruption.
    """

    result = subprocess.run(
        ["unzip", "-t", str(path)],
        text=True,
        capture_output=True,
        timeout=60,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"unzip -t failed: {result.stderr or result.stdout}")


def validate_xlsx(path: Path) -> None:
    """Validate workbook structure and business-facing text.

    Args:
        path: Workbook path to validate.

    Returns:
        None.

    Raises:
        ValueError: If sheets are missing, table/drawing residue exists, or
            technical strings leak into the workbook.
    """

    run_unzip_test(path)

    with zipfile.ZipFile(path, "r") as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise ValueError(f"Corrupt ZIP member found: {bad_member}")
        residue = [
            name
            for name in archive.namelist()
            if name.startswith("xl/tables/")
            or name.startswith("xl/drawings/")
            or ("drawing" in name.lower() and name.endswith(".rels"))
            or ("table" in name.lower() and name.endswith(".rels"))
        ]
        if residue:
            raise ValueError(f"Unexpected table/drawing residue found: {residue}")

    workbook = load_workbook(path, data_only=False)
    missing = [sheet_name for sheet_name in SHEET_NAMES if sheet_name not in workbook.sheetnames]
    if missing:
        raise ValueError(f"Missing required sheets: {missing}")
    if "数据质量检查" in workbook.sheetnames:
        raise ValueError("Workbook must not contain 数据质量检查 sheet.")

    leaked = find_technical_leaks(workbook)
    if leaked:
        raise ValueError(f"Technical/internal text leaked into workbook: {leaked[:5]}")

    quality_failures = find_diagnosis_quality_failures(workbook)
    if quality_failures:
        raise ValueError(
            "AI diagnosis quality failed; rewrite the normalized JSON with "
            f"content-specific management diagnosis before delivery: {quality_failures[:5]}"
        )


def find_technical_leaks(workbook: Workbook) -> list[str]:
    """Search workbook cells for internal tool or technical trace strings.

    Args:
        workbook: Loaded workbook.

    Returns:
        A list of sheet/cell/value snippets that contain forbidden technical
        terms.

    Raises:
        No exceptions are intentionally raised.
    """

    findings: list[str] = []
    compiled = [re.compile(pattern, re.IGNORECASE) for pattern in TECHNICAL_PATTERNS]
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                value = str(cell.value)
                if any(pattern.search(value) for pattern in compiled):
                    findings.append(f"{sheet.title}!{cell.coordinate}: {value[:80]}")
    return findings


def find_diagnosis_quality_failures(workbook: Workbook) -> list[str]:
    """Find placeholder or formulaic management diagnosis in user-facing cells.

    Args:
        workbook: Loaded workbook.

    Returns:
        A list of sheet/cell/value snippets that fail the human-readable
        diagnosis gate. Any finding should stop delivery because the workbook
        is for managers, not for technical trace or rule-template output.

    Raises:
        No exceptions are intentionally raised.
    """

    target_sheets = {
        "本次会议总览",
        "业务员询盘复盘",
        "重点询盘逐条分析",
        "共性问题归因",
        "管理复盘追问",
        "下周跟进行动表",
        "会后追踪项",
    }
    findings: list[str] = []
    for sheet in workbook.worksheets:
        if sheet.title not in target_sheets:
            continue
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                value = str(cell.value)
                placeholder_tokens = ("<待补>", "待补充示例", "TODO", "示例客户", "示例业务员")
                if any(token in value for token in placeholder_tokens) or is_formulaic_text(value):
                    findings.append(f"{sheet.title}!{cell.coordinate}: {value[:80]}")
    return findings


def main(argv: list[str]) -> int:
    """Run the XLSX build command.

    Args:
        argv: Command-line arguments without the executable name.

    Returns:
        Process exit code. Zero means the workbook was generated and validated.

    Raises:
        No exceptions escape intentionally; errors are printed to stderr so the
        calling agent can stop delivery and fix the cause.
    """

    try:
        args = parse_args(argv)
        payload = load_payload(Path(args.input))
        output = Path(args.output) if args.output else default_output_path(payload, Path(args.output_dir))
        workbook = build_workbook(payload)
        save_workbook(workbook, output)
        libreoffice_resave(output)
        remove_table_drawing_residue(output)
        validate_xlsx(output)
    except Exception as exc:  # noqa: BLE001 - CLI should report any failure clearly.
        print(f"XLSX generation failed: {exc}", file=sys.stderr)
        return 1

    print(json.dumps({"status": "success", "output": str(output)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

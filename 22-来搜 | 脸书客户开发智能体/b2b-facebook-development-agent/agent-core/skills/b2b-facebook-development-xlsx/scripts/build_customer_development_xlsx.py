#!/usr/bin/env python3
"""Build a safe customer-development XLSX workbook.

This script is intentionally conservative: it renders plain worksheet cells,
does not create Excel Table objects, re-saves through LibreOffice, removes stale
table/drawing package parts, and validates the result before delivery.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as exc:
    # The user's default Python may not be the workbook runtime. Reuse an
    # already-installed interpreter that has openpyxl; never install packages
    # or change the user's environment without confirmation.
    candidate_names = ("python3.13", "python3.12", "python3.11", "python3.10", "python3.9")
    candidate_paths: list[Path] = []
    # Inspect every matching executable in PATH, not only the first one:
    # multiple Python 3.11 installations may have different dependencies.
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        if not directory:
            continue
        for candidate_name in candidate_names:
            candidate_path = Path(directory) / candidate_name
            if candidate_path.is_file() and os.access(candidate_path, os.X_OK):
                candidate_paths.append(candidate_path)
    # Common package-manager locations may be absent from a non-login agent
    # PATH. They are safe read-only runtime fallbacks, not install targets.
    for prefix in (Path("/opt/homebrew/bin"), Path("/usr/local/bin")):
        for candidate_name in candidate_names:
            candidate_path = prefix / candidate_name
            if candidate_path.is_file() and os.access(candidate_path, os.X_OK):
                candidate_paths.append(candidate_path)

    current_runtime = Path(sys.executable).resolve()
    checked_runtimes: set[Path] = set()
    for candidate_path in candidate_paths:
        resolved = candidate_path.resolve()
        if resolved == current_runtime or resolved in checked_runtimes:
            continue
        checked_runtimes.add(resolved)
        probe = subprocess.run(
            [str(candidate_path), "-c", "import openpyxl"],
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
        if probe.returncode == 0:
            os.execv(str(candidate_path), [str(candidate_path), *sys.argv])
    raise SystemExit(
        "缺少 openpyxl。请改用已安装 openpyxl 的 Python；如需安装依赖，必须先征得用户确认。"
    ) from exc
else:
    OPENPYXL_IMPORT_ERROR = None


JsonDict = dict[str, Any]
Row = list[Any]

MISSING = "未找到可靠公开信息"

SHEET_LEADS = "开发名单"
SHEET_PROFILES = "公司画像"
SHEET_EVIDENCE = "网页证据"
SHEET_SEARCH = "搜索记录"
SHEET_RISKS = "风险待确认"
SHEET_ADVICE = "开发建议"

REQUIRED_SHEETS = {
    SHEET_LEADS,
    SHEET_PROFILES,
    SHEET_EVIDENCE,
    SHEET_SEARCH,
    SHEET_RISKS,
    SHEET_ADVICE,
}

LEAD_HEADERS = [
    "优先级",
    "评分",
    "公司名",
    "国家",
    "客户类型",
    "官网",
    "邮箱",
    "电话",
    "地址",
    "联系人",
    "职位",
    "社媒",
    "产品匹配点",
    "采购/进口线索",
    "开发切入点",
    "英文首封建议",
    "来源URL",
    "置信度",
    "待确认事项",
]

PROFILE_HEADERS = [
    "公司名",
    "官网",
    "国家/地区",
    "客户类型",
    "主营产品/类目",
    "公司简介",
    "规模线索",
    "渠道/门店/分销线索",
    "联系方式摘要",
    "来源URL",
    "备注",
]

EVIDENCE_HEADERS = [
    "公司名",
    "URL",
    "页面标题",
    "来源类型",
    "抓取摘要",
    "抽取到的信息",
    "可信度",
]

SEARCH_HEADERS = [
    "搜索词",
    "搜索API",
    "返回结果数",
    "保留候选数",
    "备注",
]

RISK_HEADERS = [
    "公司名",
    "风险/待确认项",
    "严重程度",
    "依据",
    "建议动作",
]

ADVICE_HEADERS = [
    "阶段",
    "行动项",
    "建议内容",
    "对象/负责人",
    "优先级",
    "依据/备注",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
OK_FILL = PatternFill("solid", fgColor="D9EAD3")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
DEFAULT_FONT = Font(name="Arial", size=10, color="333333")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Arial", size=10, bold=True, color="1F4E78")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def parse_args(argv: list[str]) -> argparse.Namespace:
    """Parse command-line arguments.

    Args:
        argv: Command-line arguments after the script name.

    Returns:
        An argparse namespace with input, output, and optional log paths.

    Raises:
        SystemExit: argparse raises this when required arguments are missing.
    """

    parser = argparse.ArgumentParser(
        description="Build a customer-development XLSX workbook from normalized JSON."
    )
    parser.add_argument("--input", required=True, help="Normalized customer JSON path.")
    parser.add_argument("--output", required=True, help="Output .xlsx path.")
    parser.add_argument(
        "--log",
        help="Optional log path. Defaults to the output path with .log suffix.",
    )
    return parser.parse_args(argv)


def configure_logging(output_path: Path, log_path: str | None) -> None:
    """Configure file logging for repeatable troubleshooting.

    Args:
        output_path: Final workbook path; used to derive the default log path.
        log_path: Optional explicit log file path.

    Returns:
        None.

    Raises:
        OSError: If the log directory cannot be created.
    """

    target = Path(log_path) if log_path else output_path.with_suffix(".log")
    target.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        filename=target,
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    logging.info("Starting customer-development XLSX build.")


def require_openpyxl() -> None:
    """Fail early when openpyxl is unavailable.

    Args:
        None.

    Returns:
        None.

    Raises:
        RuntimeError: If importing openpyxl failed at module load time.
    """

    if OPENPYXL_IMPORT_ERROR is not None:
        raise RuntimeError("openpyxl is required to generate .xlsx files.") from OPENPYXL_IMPORT_ERROR


def load_payload(path: Path) -> JsonDict:
    """Load normalized JSON input from disk.

    Args:
        path: Path to the JSON file prepared by the Accio agent.

    Returns:
        The parsed JSON object.

    Raises:
        FileNotFoundError: If the input path does not exist.
        json.JSONDecodeError: If the file content is not valid JSON.
        ValueError: If the JSON root is not an object.
    """

    logging.info("Loading payload from %s", path)
    with path.open("r", encoding="utf-8") as file_obj:
        payload = json.load(file_obj)
    if not isinstance(payload, dict):
        raise ValueError("Input JSON root must be an object.")
    return payload


def as_list(value: Any) -> list[Any]:
    """Normalize a value into a list.

    Args:
        value: A list, scalar, dictionary, or null JSON value.

    Returns:
        A list. Null becomes an empty list; scalar values become one-item lists.

    Raises:
        No intentional exceptions.
    """

    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def clean_text(value: Any) -> str:
    """Convert any JSON value into readable worksheet text.

    Args:
        value: Any JSON-compatible value.

    Returns:
        A compact string. Missing values become the shared missing-data label so
        the workbook does not imply unavailable data was verified.

    Raises:
        No intentional exceptions.
    """

    if value is None:
        return MISSING
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        parts = [clean_text(item) for item in value]
        parts = [part for part in parts if part and part != MISSING]
        return "；".join(parts) if parts else MISSING
    if isinstance(value, dict):
        parts = [
            f"{clean_text(key)}: {clean_text(val)}"
            for key, val in value.items()
            if clean_text(val) != MISSING
        ]
        return "；".join(parts) if parts else MISSING
    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return MISSING
    # Excel interprets these prefixes as formulas. Prefixing an apostrophe
    # keeps untrusted web/user text as literal cell content.
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def first_value(data: JsonDict, keys: list[str]) -> Any:
    """Read the first non-empty field from a dictionary.

    Args:
        data: Source dictionary.
        keys: Candidate field names ordered by preference.

    Returns:
        The first useful value, or None when no key has useful data.

    Raises:
        No intentional exceptions.
    """

    for key in keys:
        value = data.get(key)
        if value not in (None, "", [], {}):
            return value
    return None


def get_dict(payload: JsonDict, key: str) -> JsonDict:
    """Return a dictionary field safely.

    Args:
        payload: Parent JSON object.
        key: Field name to read.

    Returns:
        The dictionary at the key, or an empty dictionary when absent/wrong type.

    Raises:
        No intentional exceptions.
    """

    value = payload.get(key)
    return value if isinstance(value, dict) else {}


def get_records(payload: JsonDict, key: str) -> list[JsonDict]:
    """Return a list of dictionary records from a payload field.

    Args:
        payload: Parent JSON object.
        key: Field name that should contain a list of objects.

    Returns:
        Only dictionary items from the field, preserving order.

    Raises:
        No intentional exceptions.
    """

    return [item for item in as_list(payload.get(key)) if isinstance(item, dict)]


def summarize_contacts(lead: JsonDict, field: str) -> str:
    """Summarize contact person fields for one lead.

    Args:
        lead: One lead record.
        field: Contact field to summarize, such as name, title, email, linkedin.

    Returns:
        A joined string of all values found in contacts plus direct lead fields.

    Raises:
        No intentional exceptions.
    """

    values: list[str] = []
    direct = first_value(lead, [field, f"contact_{field}"])
    if direct not in (None, "", [], {}):
        values.extend(clean_text(item) for item in as_list(direct))
    for contact in as_list(lead.get("contacts")):
        if isinstance(contact, dict):
            value = contact.get(field)
            if value not in (None, "", [], {}):
                values.append(clean_text(value))
    values = [value for value in values if value and value != MISSING]
    return "；".join(dict.fromkeys(values)) if values else MISSING


def make_source_summary(lead: JsonDict) -> str:
    """Create a source URL summary for one lead.

    Args:
        lead: One lead record.

    Returns:
        A semicolon-separated URL/source list.

    Raises:
        No intentional exceptions.
    """

    sources: list[str] = []
    for key in ("source_urls", "sources", "urls", "evidence_urls"):
        for item in as_list(lead.get(key)):
            if isinstance(item, dict):
                url = first_value(item, ["url", "source_url", "link"])
                title = first_value(item, ["title", "name", "source_type"])
                if url:
                    sources.append(f"{clean_text(title)}: {clean_text(url)}" if title else clean_text(url))
            elif item not in (None, ""):
                sources.append(clean_text(item))
    website = first_value(lead, ["website", "official_website", "homepage"])
    if website:
        sources.insert(0, clean_text(website))
    sources = [source for source in sources if source and source != MISSING]
    return "；".join(dict.fromkeys(sources)) if sources else MISSING


def lead_to_row(lead: JsonDict) -> Row:
    """Convert one lead object into the main worksheet row.

    Args:
        lead: Normalized or semi-normalized lead dictionary.

    Returns:
        Row data ordered according to LEAD_HEADERS.

    Raises:
        No intentional exceptions.
    """

    return [
        first_value(lead, ["priority", "bd_priority", "level"]),
        first_value(lead, ["score", "bd_score"]),
        first_value(lead, ["company_name", "name", "company"]),
        first_value(lead, ["country", "region", "market"]),
        first_value(lead, ["customer_type", "type", "buyer_type"]),
        first_value(lead, ["website", "official_website", "homepage", "domain"]),
        first_value(lead, ["emails", "email", "work_email"]),
        first_value(lead, ["phones", "phone", "telephone"]),
        first_value(lead, ["address", "location", "company_address"]),
        summarize_contacts(lead, "name"),
        summarize_contacts(lead, "title"),
        first_value(lead, ["social_links", "social", "linkedin", "company_linkedin"]),
        first_value(lead, ["product_match", "product_fit", "matched_products"]),
        first_value(lead, ["procurement_import_clues", "import_clues", "purchase_clues"]),
        first_value(lead, ["development_angle", "opening_angle", "sales_angle"]),
        first_value(lead, ["first_email_en", "email_draft", "outreach_email"]),
        make_source_summary(lead),
        first_value(lead, ["confidence", "evidence_confidence"]),
        first_value(lead, ["todo", "to_confirm", "open_questions", "risks"]),
    ]


def make_profile_records(payload: JsonDict) -> list[JsonDict]:
    """Build company profile records, falling back to leads when needed.

    Args:
        payload: Full normalized JSON payload.

    Returns:
        A list of profile dictionaries.

    Raises:
        No intentional exceptions.
    """

    profiles = get_records(payload, "company_profiles")
    if profiles:
        return profiles
    profiles = []
    for lead in get_records(payload, "leads"):
        profiles.append(
            {
                "company_name": first_value(lead, ["company_name", "name", "company"]),
                "website": first_value(lead, ["website", "official_website", "homepage", "domain"]),
                "country": first_value(lead, ["country", "region", "market"]),
                "customer_type": first_value(lead, ["customer_type", "type", "buyer_type"]),
                "products": first_value(lead, ["product_match", "product_fit", "matched_products"]),
                "description": first_value(lead, ["company_description", "description", "summary"]),
                "scale": first_value(lead, ["scale", "company_size", "size_clues"]),
                "channels": first_value(lead, ["channels", "distribution_clues", "store_clues"]),
                "contact_summary": first_value(lead, ["emails", "email", "phones", "phone"]),
                "source_urls": make_source_summary(lead),
                "notes": first_value(lead, ["todo", "notes", "to_confirm"]),
            }
        )
    return profiles


def profile_to_row(profile: JsonDict) -> Row:
    """Convert one company profile object into worksheet row data.

    Args:
        profile: Company profile dictionary.

    Returns:
        Row data ordered according to PROFILE_HEADERS.

    Raises:
        No intentional exceptions.
    """

    return [
        first_value(profile, ["company_name", "name", "company"]),
        first_value(profile, ["website", "official_website", "homepage", "domain"]),
        first_value(profile, ["country", "region", "market"]),
        first_value(profile, ["customer_type", "type", "buyer_type"]),
        first_value(profile, ["products", "product_match", "main_products"]),
        first_value(profile, ["description", "company_description", "summary"]),
        first_value(profile, ["scale", "company_size", "size_clues"]),
        first_value(profile, ["channels", "distribution_clues", "store_clues"]),
        first_value(profile, ["contact_summary", "contacts", "emails", "phones"]),
        first_value(profile, ["source_urls", "sources", "url"]),
        first_value(profile, ["notes", "todo", "to_confirm"]),
    ]


def make_evidence_records(payload: JsonDict) -> list[JsonDict]:
    """Build web evidence records, falling back to lead source URLs.

    Args:
        payload: Full normalized JSON payload.

    Returns:
        A list of evidence dictionaries.

    Raises:
        No intentional exceptions.
    """

    evidence = get_records(payload, "web_evidence")
    if evidence:
        return evidence
    fallback: list[JsonDict] = []
    for lead in get_records(payload, "leads"):
        company_name = first_value(lead, ["company_name", "name", "company"])
        for url in as_list(lead.get("source_urls")):
            fallback.append(
                {
                    "company_name": company_name,
                    "url": url,
                    "title": "候选客户来源",
                    "source_type": "公开网页",
                    "summary": first_value(lead, ["product_match", "product_fit", "development_angle"]),
                    "extracted_fields": first_value(lead, ["emails", "phones", "address", "contacts"]),
                    "reliability": first_value(lead, ["confidence", "evidence_confidence"]),
                }
            )
    return fallback


def evidence_to_row(evidence: JsonDict) -> Row:
    """Convert one evidence object into worksheet row data.

    Args:
        evidence: Web evidence dictionary.

    Returns:
        Row data ordered according to EVIDENCE_HEADERS.

    Raises:
        No intentional exceptions.
    """

    return [
        first_value(evidence, ["company_name", "company", "name"]),
        first_value(evidence, ["url", "source_url", "link"]),
        first_value(evidence, ["title", "page_title"]),
        first_value(evidence, ["source_type", "type"]),
        first_value(evidence, ["summary", "snippet", "content"]),
        first_value(evidence, ["extracted_fields", "extracted", "facts"]),
        first_value(evidence, ["reliability", "confidence"]),
    ]


def search_to_row(record: JsonDict) -> Row:
    """Convert one search record into worksheet row data.

    Args:
        record: Search execution dictionary.

    Returns:
        Row data ordered according to SEARCH_HEADERS.

    Raises:
        No intentional exceptions.
    """

    return [
        first_value(record, ["query", "search_query"]),
        first_value(record, ["api", "engine"]),
        first_value(record, ["result_count", "results_count", "returned_count"]),
        first_value(record, ["kept_count", "candidate_count"]),
        first_value(record, ["notes", "summary"]),
    ]


def risk_to_row(record: JsonDict) -> Row:
    """Convert one risk record into worksheet row data.

    Args:
        record: Risk or confirmation item dictionary.

    Returns:
        Row data ordered according to RISK_HEADERS.

    Raises:
        No intentional exceptions.
    """

    return [
        first_value(record, ["company_name", "company", "name"]),
        first_value(record, ["risk", "issue", "question", "item"]),
        first_value(record, ["severity", "level", "priority"]),
        first_value(record, ["basis", "evidence", "source"]),
        first_value(record, ["action", "suggestion", "next_step"]),
    ]


def advice_to_row(record: JsonDict) -> Row:
    """Convert one development advice record into worksheet row data.

    Args:
        record: Development advice dictionary.

    Returns:
        Row data ordered according to ADVICE_HEADERS.

    Raises:
        No intentional exceptions.
    """

    return [
        first_value(record, ["stage", "phase"]),
        first_value(record, ["action", "task"]),
        first_value(record, ["message", "advice", "content"]),
        first_value(record, ["owner", "target", "role"]),
        first_value(record, ["priority", "level"]),
        first_value(record, ["basis", "notes", "source"]),
    ]


def input_summary(payload: JsonDict) -> str:
    """Create a compact business summary for worksheet subtitles.

    Args:
        payload: Full normalized JSON payload.

    Returns:
        A readable summary of product, customer type, country, and target count.

    Raises:
        No intentional exceptions.
    """

    input_data = get_dict(payload, "input")
    product = clean_text(first_value(input_data, ["product", "category"]))
    customer_type = clean_text(first_value(input_data, ["customer_type", "buyer_type", "type"]))
    country = clean_text(first_value(input_data, ["country", "region", "market"]))
    target_count = clean_text(first_value(input_data, ["target_count", "count", "limit"]) or 30)
    return f"产品/类目：{product} | 客户类型：{customer_type} | 国家/地区：{country} | 目标数量：{target_count}"


def append_rows(sheet: Worksheet, rows: list[Row], header: bool = False) -> None:
    """Append rows and apply consistent worksheet styling.

    Args:
        sheet: Target worksheet.
        rows: Two-dimensional row data.
        header: Whether the first appended row is a header row.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if cells cannot be written.
    """

    for row_index, row in enumerate(rows, start=1):
        sheet.append([clean_text(cell) for cell in row])
        appended_row = sheet.max_row
        for cell in sheet[appended_row]:
            cell.font = DEFAULT_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header and row_index == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_column_widths(sheet: Worksheet, widths: dict[str, int]) -> None:
    """Set stable column widths for easier scanning.

    Args:
        sheet: Target worksheet.
        widths: Mapping from column letter to Excel width.

    Returns:
        None.

    Raises:
        No intentional exceptions.
    """

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def write_title(sheet: Worksheet, title: str, subtitle: str, end_column: int) -> None:
    """Write a merged title band at the top of a worksheet.

    Args:
        sheet: Target worksheet.
        title: Business-facing sheet title.
        subtitle: Context summary shown under the title.
        end_column: Last column included in the merged title area.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if merging cells fails.
    """

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.fill = HEADER_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 26
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    subtitle_cell = sheet.cell(row=2, column=1, value=subtitle)
    subtitle_cell.fill = SUBHEADER_FILL
    subtitle_cell.font = LABEL_FONT
    subtitle_cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 24


def status_fill(value: Any) -> PatternFill | None:
    """Choose a fill color for priority, confidence, and risk labels.

    Args:
        value: Cell value to inspect.

    Returns:
        A PatternFill for known labels, otherwise None.

    Raises:
        No intentional exceptions.
    """

    text = clean_text(value)
    if text in {"A", "高", "高置信", "低风险"}:
        return OK_FILL
    if text in {"B", "中", "中置信", "待确认", "中风险"}:
        return WARNING_FILL
    if text in {"C", "低", "低置信", "高风险", "排除"}:
        return ERROR_FILL
    return None


def apply_status_styles(sheet: Worksheet) -> None:
    """Apply status colors to all known status cells.

    Args:
        sheet: Worksheet to style.

    Returns:
        None.

    Raises:
        No intentional exceptions.
    """

    for row in sheet.iter_rows():
        for cell in row:
            fill = status_fill(cell.value)
            if fill is not None:
                cell.fill = fill


def write_table_sheet(
    workbook: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[Row],
    widths: dict[str, int],
) -> Worksheet:
    """Create one titled table-like worksheet without Excel Table objects.

    Args:
        workbook: Workbook to modify.
        sheet_name: Name of the worksheet to create or reuse.
        title: Human-readable title shown in row 1.
        subtitle: Context summary shown in row 2.
        headers: Header labels.
        rows: Data rows.
        widths: Column width mapping.

    Returns:
        The created worksheet.

    Raises:
        openpyxl exceptions may propagate if worksheet operations fail.
    """

    if workbook.worksheets and workbook.active.title == "Sheet" and sheet_name == SHEET_LEADS:
        sheet = workbook.active
        sheet.title = sheet_name
    else:
        sheet = workbook.create_sheet(sheet_name)
    write_title(sheet, title, subtitle, len(headers))
    data_rows = [headers]
    data_rows.extend(rows if rows else [["未返回"] + [MISSING] * (len(headers) - 1)])
    append_rows(sheet, data_rows, header=True)
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{sheet.cell(row=3, column=len(headers)).coordinate}"
    set_column_widths(sheet, widths)
    apply_status_styles(sheet)
    return sheet


def build_workbook(payload: JsonDict, output_path: Path) -> None:
    """Render the workbook before compatibility normalization.

    Args:
        payload: Normalized customer-development payload.
        output_path: Path where the initial workbook should be saved.

    Returns:
        None.

    Raises:
        OSError: If the output directory cannot be created.
        openpyxl exceptions may propagate if the workbook cannot be saved.
    """

    logging.info("Rendering workbook to %s", output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    subtitle = input_summary(payload)

    lead_rows = [lead_to_row(lead) for lead in get_records(payload, "leads")]
    write_table_sheet(
        workbook,
        SHEET_LEADS,
        "客户开发名单",
        subtitle,
        LEAD_HEADERS,
        lead_rows,
        {
            "A": 10,
            "B": 10,
            "C": 28,
            "D": 16,
            "E": 16,
            "F": 28,
            "G": 28,
            "H": 20,
            "I": 30,
            "J": 20,
            "K": 22,
            "L": 28,
            "M": 38,
            "N": 36,
            "O": 36,
            "P": 48,
            "Q": 44,
            "R": 12,
            "S": 34,
        },
    )

    profile_rows = [profile_to_row(profile) for profile in make_profile_records(payload)]
    write_table_sheet(
        workbook,
        SHEET_PROFILES,
        "公司画像",
        subtitle,
        PROFILE_HEADERS,
        profile_rows,
        {
            "A": 28,
            "B": 28,
            "C": 16,
            "D": 16,
            "E": 34,
            "F": 42,
            "G": 24,
            "H": 32,
            "I": 30,
            "J": 44,
            "K": 30,
        },
    )

    evidence_rows = [evidence_to_row(evidence) for evidence in make_evidence_records(payload)]
    write_table_sheet(
        workbook,
        SHEET_EVIDENCE,
        "网页证据",
        subtitle,
        EVIDENCE_HEADERS,
        evidence_rows,
        {
            "A": 28,
            "B": 44,
            "C": 28,
            "D": 16,
            "E": 48,
            "F": 44,
            "G": 12,
        },
    )

    search_rows = [search_to_row(record) for record in get_records(payload, "search_records")]
    write_table_sheet(
        workbook,
        SHEET_SEARCH,
        "搜索记录",
        subtitle,
        SEARCH_HEADERS,
        search_rows,
        {"A": 54, "B": 14, "C": 14, "D": 14, "E": 44},
    )

    risk_rows = [risk_to_row(record) for record in get_records(payload, "risks")]
    write_table_sheet(
        workbook,
        SHEET_RISKS,
        "风险待确认",
        subtitle,
        RISK_HEADERS,
        risk_rows,
        {"A": 28, "B": 42, "C": 14, "D": 44, "E": 38},
    )

    advice_rows = [advice_to_row(record) for record in get_records(payload, "development_advice")]
    write_table_sheet(
        workbook,
        SHEET_ADVICE,
        "开发建议",
        subtitle,
        ADVICE_HEADERS,
        advice_rows,
        {"A": 16, "B": 24, "C": 52, "D": 22, "E": 12, "F": 42},
    )

    workbook.save(output_path)


def find_libreoffice() -> str:
    """Find a LibreOffice executable for the required re-save step.

    Args:
        None.

    Returns:
        Path or command name for LibreOffice/soffice.

    Raises:
        RuntimeError: If LibreOffice cannot be found.
    """

    candidates = [
        "soffice",
        "libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return candidate
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    raise RuntimeError("LibreOffice/soffice not found; XLSX safety flow cannot finish.")


def libreoffice_resave(path: Path) -> None:
    """Re-save the workbook through LibreOffice headless.

    Args:
        path: Workbook path to normalize in place.

    Returns:
        None.

    Raises:
        RuntimeError: If conversion fails or no XLSX is produced.
    """

    logging.info("Running LibreOffice headless re-save.")
    soffice = find_libreoffice()
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        input_dir = tmp_path / "input"
        output_dir = tmp_path / "output"
        profile_dir = tmp_path / "profile"
        input_dir.mkdir()
        output_dir.mkdir()
        profile_dir.mkdir()
        input_copy = input_dir / path.name
        shutil.copy2(path, input_copy)
        result = subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(output_dir),
                str(input_copy),
            ],
            text=True,
            capture_output=True,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice re-save failed: {result.stderr or result.stdout}")
        candidates = list(output_dir.glob("*.xlsx"))
        if not candidates:
            raise RuntimeError("LibreOffice did not produce an .xlsx file.")
        shutil.copy2(candidates[0], path)


def sanitize_xml_text(xml_text: str) -> str:
    """Remove stale table/drawing XML references.

    Args:
        xml_text: XML text decoded from a workbook package part.

    Returns:
        Sanitized XML text.

    Raises:
        No intentional exceptions.
    """

    xml_text = re.sub(r"<tableParts\b[^>]*/>", "", xml_text)
    xml_text = re.sub(r"<drawing\b[^>]*/>", "", xml_text)
    xml_text = re.sub(
        r"<Relationship\b[^>]*(?:/tables/|/drawings/)[^>]*/>",
        "",
        xml_text,
        flags=re.IGNORECASE,
    )
    xml_text = re.sub(
        r"<Override\b[^>]*(?:/tables/|/drawings/)[^>]*/>",
        "",
        xml_text,
        flags=re.IGNORECASE,
    )
    return xml_text


def sanitize_xlsx_package(path: Path) -> None:
    """Remove table/drawing package residue from the workbook.

    Args:
        path: XLSX path to sanitize in place.

    Returns:
        None.

    Raises:
        zipfile.BadZipFile: If the workbook is not a valid ZIP package.
        OSError: If the sanitized archive cannot be written.
    """

    logging.info("Sanitizing XLSX package.")
    temp_path = path.with_suffix(".sanitized.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            lowered = item.filename.lower()
            if lowered.startswith("xl/tables/") or lowered.startswith("xl/drawings/"):
                continue
            data = source.read(item.filename)
            if item.filename.endswith(".xml") or item.filename.endswith(".rels"):
                try:
                    data = sanitize_xml_text(data.decode("utf-8")).encode("utf-8")
                except UnicodeDecodeError:
                    logging.warning("Skipping non-UTF8 XML-like package part: %s", item.filename)
            target.writestr(item, data)
    temp_path.replace(path)


def run_unzip_test(path: Path) -> None:
    """Validate the XLSX ZIP container with `unzip -t`.

    Args:
        path: Workbook path to inspect.

    Returns:
        None.

    Raises:
        RuntimeError: If unzip reports package errors.
    """

    logging.info("Running unzip -t validation.")
    result = subprocess.run(["unzip", "-t", str(path)], text=True, capture_output=True, check=False)
    if result.returncode != 0:
        raise RuntimeError(f"unzip -t failed: {result.stderr or result.stdout}")


def validate_xlsx(path: Path) -> None:
    """Validate workbook sheets and forbidden package residue.

    Args:
        path: Workbook path to validate.

    Returns:
        None.

    Raises:
        RuntimeError: If required sheets are missing or package residue remains.
        zipfile.BadZipFile: If the workbook is not a valid ZIP package.
        openpyxl exceptions may propagate if the workbook cannot be opened.
    """

    logging.info("Running openpyxl and package residue validation.")
    workbook = load_workbook(path, data_only=False)
    missing_sheets = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing_sheets:
        raise RuntimeError(f"Workbook missing required sheets: {missing_sheets}")
    with zipfile.ZipFile(path, "r") as archive:
        forbidden: list[str] = []
        for name in archive.namelist():
            if name.startswith("xl/tables/") or name.startswith("xl/drawings/"):
                forbidden.append(name)
                continue
            if name.endswith((".xml", ".rels")):
                xml_text = archive.read(name).decode("utf-8", errors="ignore")
                if "tableParts" in xml_text or "/drawings/" in xml_text or "/tables/" in xml_text:
                    forbidden.append(name)
    if forbidden:
        raise RuntimeError(f"XLSX still contains table/drawing residue: {forbidden[:5]}")


def run_safety_flow(path: Path) -> None:
    """Run all mandatory Excel safety steps before delivery.

    Args:
        path: Workbook path to normalize and validate.

    Returns:
        None.

    Raises:
        RuntimeError: If any required safety step fails.
    """

    libreoffice_resave(path)
    sanitize_xlsx_package(path)
    run_unzip_test(path)
    validate_xlsx(path)
    logging.info("XLSX safety flow finished successfully.")


def main(argv: list[str]) -> int:
    """Program entry point.

    Args:
        argv: Command-line arguments after the executable name.

    Returns:
        Zero on success; non-zero when generation or validation fails.

    Raises:
        No intentional exceptions; unexpected exceptions are logged and reported.
    """

    args = parse_args(argv)
    output_path = Path(args.output).expanduser().resolve()
    try:
        configure_logging(output_path, args.log)
        require_openpyxl()
        payload = load_payload(Path(args.input).expanduser().resolve())
        build_workbook(payload, output_path)
        run_safety_flow(output_path)
    except Exception as exc:  # noqa: BLE001 - CLI should log every failure type.
        logging.exception("Customer-development XLSX build failed.")
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(f"XLSX_READY {output_path}")
    logging.info("Customer-development XLSX build completed: %s", output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

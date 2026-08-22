#!/usr/bin/env python3
"""Build and validate the Data Advisor business decision workbook."""

from __future__ import annotations

import argparse
import json
import logging
import re
import shutil
import subprocess
import tempfile
import zipfile
from copy import copy
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MISSING = "用户未提供"
FIELD_LABELS: dict[str, str] = {
    "diagnosis_meta": "诊断说明", "period_label": "数据周期", "shop_category": "店铺类目",
    "diagnosis_confidence": "诊断置信度", "confidence_reason": "置信度说明",
    "executive_thesis": "老板结论", "one_sentence": "一句话判断", "biggest_risk": "最大风险",
    "biggest_opportunity": "最大机会", "management_decision_needed": "需要管理层拍板",
    "boss_decisions": "老板拍板事项", "decision": "经营裁决", "decision_type": "动作类型",
    "operation_lever": "经营抓手", "why_now": "为什么现在处理",
    "consequence_if_not_decided": "不处理的后果", "evidence_deltas": "证据差额",
    "metric": "指标", "current": "当前值", "benchmark": "对照值", "gap": "差额", "source": "证据来源",
    "owner_immediate_action": "负责人立即动作", "owner_review_cycle": "复查周期",
    "stop_rule": "停止规则", "review_metric": "复查指标", "priority": "优先级",
    "judgement": "经营判断", "object": "具体对象", "current_value": "当前值",
    "target_value": "目标值", "gap_to_target": "目标差额", "gap_calculation": "差额计算",
    "business_impact": "业务影响", "root_cause_hypothesis": "待验证根因",
    "counter_evidence_or_limit": "反证或限制", "owner_role": "负责人角色",
    "stage": "漏斗阶段", "state": "状态", "evidence_delta": "证据差额",
    "why_it_matters": "业务影响", "action": "下一动作", "failure_rule": "未达标处理",
    "account": "账号", "resource_share": "资源占比", "result_share": "结果占比",
    "efficiency": "效率", "next_action": "下一动作", "sku_or_group": "商品或商品组",
    "status": "状态", "resource_decision": "资源决策", "country": "国家",
    "product_group": "商品组", "keyword_pack": "关键词组合", "entry_object": "承接对象",
    "visitor_group": "访客群", "basis": "依据", "followup_method": "跟进方式",
    "evidence_location": "证据位置", "opportunity": "机会", "country_match_score": "国家匹配分",
    "category_match_score": "类目匹配分", "moq_testability_score": "起订量可测分",
    "price_band_score": "价格带匹配分", "transaction_index_score": "交易指数分",
    "store_readiness_score": "店铺承接分", "total_score": "总分",
    "validation_plan": "验证计划", "theme": "主题", "finding": "发现", "verification": "验证方式",
    "target": "提问对象", "question": "复盘问题", "expected_decision": "期望裁决",
    "method": "执行方法", "deadline": "截止时间", "source_diagnosis": "来源诊断",
    "data_gaps": "数据缺口", "affected_judgement": "受影响判断", "fallback_used": "已用降级方案",
    "next_data_needed": "后续所需数据", "workbook_guidance": "工作簿编排说明",
    "boss_decision_rows": "老板裁决行", "sheets_that_need_diagnosis_first": "优先诊断工作表",
    "rows_to_keep_only_in_appendix": "仅保留在附录的内容",
    "duplicate_diagnoses_to_remove": "需删除的重复诊断",
}
SOURCE_LABELS: dict[str, str] = {
    "data_advisor_shop_summary": "店铺经营大盘",
    "data_advisor_account_summary": "团队账号表现",
    "data_advisor_shop_flow": "店铺流量结构",
    "data_advisor_shop_channel": "店铺渠道分布",
    "data_advisor_shop_region": "店铺地域市场",
    "data_advisor_shop_product": "店铺商品表现",
    "data_advisor_visitor_detail": "访客明细",
    "data_advisor_shop_flow_profile": "流量画像",
    "data_advisor_to_product": "竞品流向",
    "data_advisor_category_infer": "主营类目识别",
    "data_advisor_category_prediction": "类目趋势预测",
    "data_advisor_industry_cate_rank": "行业子类目排行",
    "data_advisor_industry_country_rank": "行业国家排行",
    "data_advisor_opportunity_discovery": "平台机会发现",
    "data_advisor_product_selection": "平台选品推荐",
    "opportunity_discovery": "平台机会发现",
    "product_selection": "平台选品推荐",
    "shop_flow": "店铺流量结构",
    "cateId": "类目编号",
}
SHEET_SPECS: tuple[tuple[str, str], ...] = (
    ("经营结论", "top_diagnoses"),
    ("经营问题诊断", "detail_diagnoses"),
    ("行动与复查", "actions"),
)
ALLOWED_ROLES = {"", "运营", "业务"}
ALLOWED_PRIORITIES = {"先做", "随后做", "持续观察"}
FORBIDDEN_DIAGNOSIS_TEXT = (
    "未返回",
    "待确认",
    "不可判断",
    "用户未提供",
    "待补充",
    "TODO",
    "加强运营",
    "持续优化",
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def parse_args() -> argparse.Namespace:
    """Parse required input and output paths.

    Returns:
        Parsed command-line arguments.

    Raises:
        SystemExit: If required arguments are missing.
    """

    parser = argparse.ArgumentParser(description="Build a Data Advisor decision XLSX.")
    parser.add_argument("--input", required=True, help="Diagnosis JSON path.")
    parser.add_argument("--output", required=True, help="Final XLSX path.")
    return parser.parse_args()


def safe_text(value: Any) -> str:
    """Convert a value to readable, formula-safe worksheet text.

    Args:
        value: Any JSON-compatible value.

    Returns:
        Plain worksheet text; missing values use a visible label.

    Raises:
        No intentional exceptions.
    """

    if value is None or value == "":
        return MISSING
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        value = "；".join(
            f"{business_label(key)}: {safe_text(item)}" for key, item in value.items()
        ) if isinstance(value, dict) else "；".join(safe_text(item) for item in value)
    text = re.sub(r"\s+", " ", str(value)).strip() or MISSING
    # Source/tool identifiers are useful during collection but not appropriate
    # in a boss-facing workbook. Replace exact known identifiers even when they
    # appear inside a longer explanation.
    for technical_name, business_name in sorted(
        SOURCE_LABELS.items(), key=lambda item: len(item[0]), reverse=True
    ):
        text = text.replace(technical_name, business_name)
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def business_label(value: Any) -> str:
    """Return a business-facing Chinese label for an input field.

    Args:
        value: JSON field name or already-readable label.

    Returns:
        A Chinese business label safe for worksheet display.

    Raises:
        No intentional exceptions.
    """

    key = str(value).strip()
    if key in FIELD_LABELS:
        return FIELD_LABELS[key]
    if key in SOURCE_LABELS:
        return SOURCE_LABELS[key]
    if re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", key):
        # Unknown schema keys must not abort an otherwise valid workbook or leak
        # technical text. The underlying value is still preserved for review.
        return "补充业务字段"
    return safe_text(key)


def flatten_nested_section(value: Any) -> list[list[str]]:
    """Expand nested dictionaries/lists into lossless three-column rows.

    Args:
        value: A section containing mixed scalars, dictionaries, and lists.

    Returns:
        Rows with ``分组 / 项目 / 内容`` columns. Every scalar leaf receives its
        own row so tool coverage and detailed diagnoses are not collapsed into
        a short Python-style object string.

    Raises:
        No intentional exceptions.
    """

    rows: list[list[str]] = [["分组", "项目", "内容"]]

    def walk(node: Any, path: list[str]) -> None:
        """Recursively append scalar leaves while retaining their business path."""

        if isinstance(node, dict):
            if not node:
                rows.append([" / ".join(path) or "本节", "状态", MISSING])
                return
            for key, child in node.items():
                label = business_label(key)
                if isinstance(child, (dict, list)):
                    walk(child, [*path, label])
                else:
                    rows.append([
                        " / ".join(path) or "本节",
                        label,
                        safe_text(child),
                    ])
            return
        if isinstance(node, list):
            if not node:
                rows.append([" / ".join(path) or "本节", "状态", MISSING])
                return
            for index, child in enumerate(node, 1):
                walk(child, [*path, f"第{index}项"])
            return
        rows.append([" / ".join(path[:-1]) or "本节", path[-1] if path else "内容", safe_text(node)])

    walk(value, [])
    return rows


def section_rows(value: Any) -> list[list[str]]:
    """Normalize a JSON section into a two-dimensional row list.

    Args:
        value: Section dictionary, list, scalar, or null.

    Returns:
        Rows suitable for appending to an openpyxl worksheet.

    Raises:
        No intentional exceptions.
    """

    if isinstance(value, dict):
        if all(not isinstance(item, (dict, list)) for item in value.values()):
            return [["项目", "内容"], *[[business_label(k), safe_text(v)] for k, v in value.items()]]
        return flatten_nested_section(value)
    if isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
        headers: list[str] = []
        for item in value:
            for key in item:
                if key not in headers:
                    headers.append(key)
        return [[business_label(key) for key in headers], *[
            [safe_text(item.get(key)) for key in headers] for item in value
        ]]
    if isinstance(value, list):
        return [["序号", "内容"], *[[str(i), safe_text(item)] for i, item in enumerate(value, 1)]]
    return [["项目", "内容"], ["本节信息", safe_text(value)]]


def style_sheet(sheet: Any) -> None:
    """Apply readable formatting without tables, drawings, or formulas.

    Args:
        sheet: Worksheet to format.

    Returns:
        None.

    Raises:
        openpyxl errors may propagate for malformed worksheets.
    """

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(max(1, sheet.max_column))}{sheet.max_row}"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    for cell in sheet[3]:
        cell.fill = SUB_FILL
        cell.font = Font(name="Arial", size=10, bold=True, color="1F1F1F")
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row not in (1, 3):
                cell.font = Font(name="Arial", size=10, color="333333")
    for column in range(1, sheet.max_column + 1):
        longest = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(longest + 2, 14), 42)
    sheet.row_dimensions[1].height = 24


def diagnosis_text(value: Any) -> str:
    """Convert diagnosed content to safe worksheet text.

    Args:
        value: Scalar or short list from the LLM diagnosis JSON.

    Returns:
        Formula-safe text. Missing optional values become blank.

    Raises:
        ValueError: Nested objects reach the user-facing renderer.
    """

    if value in (None, ""):
        return ""
    if isinstance(value, dict):
        raise ValueError("Nested objects cannot be rendered into compact workbook cells.")
    if isinstance(value, list):
        rendered = "、".join(
            str(item).strip()
            for item in value[:3]
            if item not in (None, "") and str(item).strip()
        )
    else:
        rendered = re.sub(r"\s+", " ", str(value)).strip()
    return f"'{rendered}" if rendered.startswith(("=", "+", "-", "@")) else rendered


def validate_management_diagnosis(payload: dict[str, Any]) -> None:
    """Validate the operations-expert diagnosis before XLSX generation.

    Args:
        payload: Canonical diagnosis JSON produced after the facts package.

    Returns:
        None when all user-facing diagnosis rules pass.

    Raises:
        ValueError: Required content, ownership, grouping, or wording is invalid.
    """

    if not isinstance(payload, dict):
        raise ValueError("Diagnosis root must be an object.")
    conclusion = str(payload.get("executive_conclusion") or "").strip()
    top = payload.get("top_diagnoses") or []
    details = payload.get("detail_diagnoses") or []
    actions = payload.get("actions") or []
    limitations = payload.get("data_limitations") or []

    if not conclusion:
        raise ValueError("executive_conclusion is required.")
    if not isinstance(top, list) or not 1 <= len(top) <= 3:
        raise ValueError("top_diagnoses must contain 1 to 3 rows.")
    if not isinstance(details, list) or not 1 <= len(details) <= 15:
        raise ValueError("detail_diagnoses must contain 1 to 15 rows.")
    if not isinstance(actions, list) or not 1 <= len(actions) <= 8:
        raise ValueError("actions must contain 1 to 8 rows.")
    if not isinstance(limitations, list) or len(limitations) > 3:
        raise ValueError("data_limitations must contain no more than 3 notes.")

    for index, row in enumerate(top):
        if not isinstance(row, dict):
            raise ValueError(f"top_diagnoses[{index}] must be an object.")
        for key in ("what_happened", "root_cause", "solution", "review_standard"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"top_diagnoses[{index}].{key} is required.")
        if str(row.get("role") or "") not in ALLOWED_ROLES:
            raise ValueError(f"top_diagnoses[{index}].role is invalid.")

    diagnosis_pairs: set[tuple[str, str]] = set()
    for index, row in enumerate(details):
        if not isinstance(row, dict):
            raise ValueError(f"detail_diagnoses[{index}] must be an object.")
        for key in ("issue_group", "evidence", "expert_diagnosis", "root_cause", "solution"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"detail_diagnoses[{index}].{key} is required.")
        objects = row.get("objects") or []
        if not isinstance(objects, list) or len(objects) > 3:
            raise ValueError(
                f"detail_diagnoses[{index}].objects must contain at most 3 examples."
            )
        pair = (
            str(row.get("root_cause") or "").strip(),
            str(row.get("solution") or "").strip(),
        )
        if pair in diagnosis_pairs:
            raise ValueError("Duplicate root-cause and solution rows must be grouped.")
        diagnosis_pairs.add(pair)

    action_texts: set[str] = set()
    for index, row in enumerate(actions):
        if not isinstance(row, dict):
            raise ValueError(f"actions[{index}] must be an object.")
        for key in ("priority", "action", "due", "review_standard", "source_issue_group"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"actions[{index}].{key} is required.")
        if str(row.get("priority") or "") not in ALLOWED_PRIORITIES:
            raise ValueError(f"actions[{index}].priority is invalid.")
        if str(row.get("role") or "") not in ALLOWED_ROLES:
            raise ValueError(f"actions[{index}].role is invalid.")
        action = str(row.get("action") or "").strip()
        if action in action_texts:
            raise ValueError("Duplicate action rows must be grouped.")
        action_texts.add(action)

    visible_text = json.dumps(payload, ensure_ascii=False)
    matched = [
        token
        for token in FORBIDDEN_DIAGNOSIS_TEXT
        if token.lower() in visible_text.lower()
    ]
    if matched:
        raise ValueError(f"Placeholder or generic wording is forbidden: {matched}")
    technical_terms = [
        name for name in SOURCE_LABELS if name.lower() in visible_text.lower()
    ]
    if technical_terms:
        raise ValueError(
            f"Raw tool or schema names are forbidden in diagnosed content: {technical_terms[:5]}"
        )
    if re.search(r"\b(gateway|authorization|access[_ -]?token|traceback|proxy)\b", visible_text, re.I):
        raise ValueError("Internal execution wording is forbidden.")


def owner_text(item: dict[str, Any]) -> str:
    """Return a verified owner name or the allowed fallback role.

    Args:
        item: One top diagnosis or action item.

    Returns:
        ``owner_name`` when present, otherwise ``运营``/``业务`` or blank.

    Raises:
        None. Validation checks role values before this helper runs.
    """

    owner_name = str(item.get("owner_name") or "").strip()
    return owner_name or str(item.get("role") or "").strip()


def write_compact_title(
    sheet: Any,
    title: str,
    period: str,
    conclusion: str | None = None,
    column_count: int = 6,
) -> None:
    """Write one restrained title block.

    Args:
        sheet: Target worksheet.
        title: Natural-language title.
        period: Report period shown once near the top.
        conclusion: Optional executive conclusion for the first sheet.
        column_count: Number of visible columns.

    Returns:
        None.

    Raises:
        ValueError: Unsafe nested content reaches the renderer.
    """

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    sheet.cell(1, 1, diagnosis_text(title))
    sheet.cell(1, 1).fill = HEADER_FILL
    sheet.cell(1, 1).font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    sheet.cell(1, 1).alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 28

    sheet.cell(2, 1, "数据周期")
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=column_count)
    sheet.cell(2, 2, diagnosis_text(period))
    if conclusion is not None:
        sheet.cell(3, 1, "一句话结论")
        sheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=column_count)
        sheet.cell(3, 2, diagnosis_text(conclusion))
        sheet.row_dimensions[3].height = 42
    for row_number in (2, 3):
        for column_number in range(1, column_count + 1):
            cell = sheet.cell(row_number, column_number)
            if column_number == 1:
                cell.fill = SUB_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="1F4E78")
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_compact_table(
    sheet: Any,
    header_row: int,
    headers: list[str],
    rows: list[list[str]],
) -> int:
    """Write a diagnosed table with bounded row heights.

    Args:
        sheet: Target worksheet.
        header_row: One-based header row.
        headers: Natural-language column labels.
        rows: Diagnosed user-facing rows.

    Returns:
        First row after the table.

    Raises:
        ValueError: A row width differs from the header width.
    """

    for column_number, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_number, header)
        cell.fill = HEADER_FILL
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = THIN
    sheet.row_dimensions[header_row].height = 28

    for row_offset, values in enumerate(rows, start=1):
        if len(values) != len(headers):
            raise ValueError("Compact table row width does not match its headers.")
        row_number = header_row + row_offset
        for column_number, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column_number, diagnosis_text(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN
        sheet.row_dimensions[row_number].height = 54
    end_row = header_row + len(rows)
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, end_row)}"
    )
    sheet.freeze_panes = f"A{header_row + 1}"
    return end_row + 1


def style_compact_sheet(sheet: Any, widths: list[int]) -> None:
    """Apply compact widths and consistent typography.

    Args:
        sheet: Worksheet to style.
        widths: Column widths in display order.

    Returns:
        None.

    Raises:
        None.
    """

    sheet.sheet_view.showGridLines = False
    for column_number, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_number)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            updated_font = copy(cell.font)
            updated_font.name = "Arial"
            updated_font.sz = cell.font.sz or 10
            cell.font = updated_font
            updated_alignment = copy(cell.alignment)
            updated_alignment.wrap_text = True
            updated_alignment.vertical = cell.alignment.vertical or "top"
            cell.alignment = updated_alignment


def build_compact_workbook(
    diagnosis: dict[str, Any],
    context: dict[str, Any] | None = None,
) -> Workbook:
    """Build the three-sheet Data Advisor decision workbook.

    Args:
        diagnosis: Canonical operations-expert diagnosis.
        context: Optional display context such as company name.

    Returns:
        In-memory workbook containing only diagnosed, decision-useful rows.

    Raises:
        ValueError: Diagnosis quality or table structure is invalid.
    """

    validate_management_diagnosis(diagnosis)
    context = context or {}
    meta = diagnosis.get("report_meta") or {}
    title = str(meta.get("title") or "数据参谋平台分析会").strip()
    company = str(context.get("company") or "").strip()
    period = str(meta.get("period") or "").strip()
    display_title = f"{title}｜{company}" if company else title

    workbook = Workbook()
    home = workbook.active
    home.title = SHEET_SPECS[0][0]
    diagnosis_sheet = workbook.create_sheet(SHEET_SPECS[1][0])
    action_sheet = workbook.create_sheet(SHEET_SPECS[2][0])

    top_rows = diagnosis["top_diagnoses"]
    show_home_owner = any(owner_text(item) for item in top_rows)
    home_headers = ["发生了什么", "核心原因", "解决方案"]
    if show_home_owner:
        home_headers.append("负责角色")
    home_headers.append("复查标准")
    rendered_top: list[list[str]] = []
    for item in top_rows:
        row = [
            diagnosis_text(item.get("what_happened")),
            diagnosis_text(item.get("root_cause")),
            diagnosis_text(item.get("solution")),
        ]
        if show_home_owner:
            row.append(owner_text(item))
        row.append(diagnosis_text(item.get("review_standard")))
        rendered_top.append(row)
    write_compact_title(
        home,
        display_title,
        period,
        str(diagnosis.get("executive_conclusion") or ""),
        len(home_headers),
    )
    next_row = write_compact_table(home, 5, home_headers, rendered_top)
    limitations = [
        diagnosis_text(item)
        for item in diagnosis.get("data_limitations") or []
        if diagnosis_text(item)
    ]
    if limitations:
        home.cell(next_row + 1, 1, "数据限制")
        home.merge_cells(
            start_row=next_row + 1,
            start_column=2,
            end_row=next_row + 1,
            end_column=len(home_headers),
        )
        home.cell(next_row + 1, 2, "；".join(limitations))
        home.cell(next_row + 1, 1).fill = SUB_FILL
        home.cell(next_row + 1, 1).font = Font(bold=True, color="1F4E78")
    style_compact_sheet(
        home,
        [31, 30, 31] + ([14] if show_home_owner else []) + [27],
    )

    detail_headers = [
        "问题类型",
        "重点对象",
        "关键证据",
        "专家判断",
        "核心原因",
        "解决方案",
    ]
    rendered_details = [
        [
            diagnosis_text(item.get("issue_group")),
            diagnosis_text(item.get("objects") or []),
            diagnosis_text(item.get("evidence")),
            diagnosis_text(item.get("expert_diagnosis")),
            diagnosis_text(item.get("root_cause")),
            diagnosis_text(item.get("solution")),
        ]
        for item in diagnosis["detail_diagnoses"]
    ]
    write_compact_title(
        diagnosis_sheet,
        SHEET_SPECS[1][0],
        period,
        column_count=len(detail_headers),
    )
    write_compact_table(diagnosis_sheet, 4, detail_headers, rendered_details)
    style_compact_sheet(diagnosis_sheet, [16, 24, 30, 30, 28, 30])

    actions = diagnosis["actions"]
    show_action_owner = any(owner_text(item) for item in actions)
    action_headers = ["优先级", "行动"]
    if show_action_owner:
        action_headers.append("负责角色")
    action_headers.extend(["完成时点", "复查标准"])
    rendered_actions: list[list[str]] = []
    for item in actions:
        row = [
            diagnosis_text(item.get("priority")),
            diagnosis_text(item.get("action")),
        ]
        if show_action_owner:
            row.append(owner_text(item))
        row.extend(
            [
                diagnosis_text(item.get("due")),
                diagnosis_text(item.get("review_standard")),
            ]
        )
        rendered_actions.append(row)
    write_compact_title(
        action_sheet,
        SHEET_SPECS[2][0],
        period,
        column_count=len(action_headers),
    )
    write_compact_table(action_sheet, 4, action_headers, rendered_actions)
    style_compact_sheet(
        action_sheet,
        [13, 38] + ([14] if show_action_owner else []) + [16, 32],
    )
    return workbook


def build_workbook(payload: dict[str, Any], target: Path) -> None:
    """Build and save the compact diagnosis-first Data Advisor workbook.

    Args:
        payload: Canonical operations-expert diagnosis payload.
        target: Temporary workbook output path.

    Returns:
        None.

    Raises:
        OSError: If the workbook cannot be saved.
    """

    workbook = build_compact_workbook(
        payload,
        {"generated_at": date.today().isoformat()},
    )
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    workbook.close()


def libreoffice_resave(source: Path, destination: Path) -> None:
    """Normalize the workbook through LibreOffice headless mode.

    Args:
        source: Workbook produced by openpyxl.
        destination: Final normalized workbook path.

    Returns:
        None.

    Raises:
        RuntimeError: If LibreOffice is missing or normalization fails.
    """

    office = (
        shutil.which("libreoffice")
        or shutil.which("soffice")
        or (
            "/Applications/LibreOffice.app/Contents/MacOS/soffice"
            if Path("/Applications/LibreOffice.app/Contents/MacOS/soffice").exists()
            else None
        )
    )
    if not office:
        raise RuntimeError("LibreOffice/soffice is required for final XLSX normalization.")
    with tempfile.TemporaryDirectory(prefix="data-advisor-lo-") as temp_dir:
        temp_path = Path(temp_dir)
        profile_dir = temp_path / "profile"
        profile_dir.mkdir()
        result = subprocess.run(
            [office, f"-env:UserInstallation={profile_dir.resolve().as_uri()}", "--headless", "--convert-to", "xlsx", "--outdir", str(temp_path), str(source)],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
        converted = temp_path / source.name
        if result.returncode != 0 or not converted.exists():
            raise RuntimeError(f"LibreOffice normalization failed: {result.stderr or result.stdout}")
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(converted, destination)


def validate_workbook(path: Path) -> None:
    """Validate workbook structure and forbidden package artifacts.

    Args:
        path: Final workbook path.

    Returns:
        None.

    Raises:
        ValueError: If sheets or package members violate the contract.
        zipfile.BadZipFile: If the XLSX package is corrupt.
    """

    workbook = load_workbook(path, read_only=False, data_only=False)
    expected = [name for name, _ in SHEET_SPECS]
    if workbook.sheetnames != expected:
        raise ValueError(f"Unexpected sheet order: {workbook.sheetnames}")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise ValueError(f"Formula cell is not allowed: {sheet.title}!{cell.coordinate}")
                value = str(cell.value or "")
                if "{'" in value or re.search(r"'[^']+'\s*:", value):
                    raise ValueError(f"Raw object text is not allowed: {sheet.title}!{cell.coordinate}")
                if any(name in value for name in SOURCE_LABELS):
                    raise ValueError(f"Raw tool/schema name is not allowed: {sheet.title}!{cell.coordinate}")
                if re.search(r"\b(gateway|authorization|access[_ -]?token|traceback|proxy)\b", value, re.I):
                    raise ValueError(f"Internal execution term is not allowed: {sheet.title}!{cell.coordinate}")
                if any(token.lower() in value.lower() for token in FORBIDDEN_DIAGNOSIS_TEXT):
                    raise ValueError(
                        f"Placeholder or generic diagnosis is not allowed: "
                        f"{sheet.title}!{cell.coordinate}"
                    )
                stripped = value.strip()
                if (
                    (stripped.startswith("{") and stripped.endswith("}"))
                    or (stripped.startswith("[") and stripped.endswith("]"))
                ):
                    raise ValueError(f"Raw JSON text is not allowed: {sheet.title}!{cell.coordinate}")
    with zipfile.ZipFile(path) as archive:
        forbidden = [
            name for name in archive.namelist()
            if name.startswith("xl/tables/") or name.startswith("xl/drawings/")
        ]
        if forbidden:
            raise ValueError(f"Forbidden XLSX package members: {forbidden}")


def strip_package_residue(path: Path) -> None:
    """Re-save through openpyxl to remove LibreOffice drawing placeholders.

    Args:
        path: Normalized workbook path.

    Returns:
        None.

    Raises:
        openpyxl and filesystem errors propagate when normalization fails.
    """

    with tempfile.TemporaryDirectory(prefix="data-advisor-strip-") as temp_dir:
        clean_path = Path(temp_dir) / path.name
        workbook = load_workbook(path, read_only=False, data_only=False)
        workbook.save(clean_path)
        workbook.close()
        shutil.copy2(clean_path, path)


def main() -> None:
    """Load input, build, normalize, validate, and log the workbook.

    Returns:
        None.

    Raises:
        JSONDecodeError: If input JSON is invalid.
        RuntimeError: If normalization fails.
        ValueError: If validation fails.
    """

    args = parse_args()
    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve()
    # Create the run-local output directory before opening the adjacent log.
    # Fresh task output directories are intentionally empty, so logging must not
    # assume that a parent folder already exists.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        filename=output_path.with_suffix(".log"),
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    with input_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError("Input JSON root must be an object.")
    with tempfile.TemporaryDirectory(prefix="data-advisor-build-") as temp_dir:
        draft = Path(temp_dir) / output_path.name
        build_workbook(payload, draft)
        libreoffice_resave(draft, output_path)
    strip_package_residue(output_path)
    validate_workbook(output_path)
    logging.info("Workbook validated: %s", output_path)
    print(output_path)


if __name__ == "__main__":
    main()

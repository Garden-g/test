#!/usr/bin/env python3
"""Build a company background-check report as a safe XLSX workbook.

The script receives one normalized JSON file created by the agent after web
research and Apify collection. It renders a multi-sheet workbook for business
users, then runs a safety flow so Excel can open the file without table/drawing
repair warnings.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as exc:
    # Prefer another already-installed Python runtime when the default one
    # lacks openpyxl. This does not install or modify dependencies.
    candidate_names = ("python3.13", "python3.12", "python3.11", "python3.10", "python3.9")
    candidate_paths: list[Path] = []
    # Inspect every matching executable in PATH, not only the first one:
    # multiple Python 3.11 installations may have different dependencies.
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        if not directory:
            continue
        for candidate_name in candidate_names:
            candidate_path = Path(directory) / candidate_name
            if candidate_path.is_file() and os.access(candidate_path, os.X_OK):
                candidate_paths.append(candidate_path)
    # Common package-manager locations may be absent from a non-login agent
    # PATH. They are safe read-only runtime fallbacks, not install targets.
    for prefix in (Path("/opt/homebrew/bin"), Path("/usr/local/bin")):
        for candidate_name in candidate_names:
            candidate_path = prefix / candidate_name
            if candidate_path.is_file() and os.access(candidate_path, os.X_OK):
                candidate_paths.append(candidate_path)

    current_runtime = Path(sys.executable).resolve()
    checked_runtimes: set[Path] = set()
    for candidate_path in candidate_paths:
        resolved = candidate_path.resolve()
        if resolved == current_runtime or resolved in checked_runtimes:
            continue
        checked_runtimes.add(resolved)
        probe = subprocess.run(
            [str(candidate_path), "-c", "import openpyxl"],
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
        if probe.returncode == 0:
            os.execv(str(candidate_path), [str(candidate_path), *sys.argv])
    raise SystemExit(
        "缺少 openpyxl。请改用已安装 openpyxl 的 Python；如需安装依赖，必须先征得用户确认。"
    ) from exc
else:
    OPENPYXL_IMPORT_ERROR = None


JsonDict = dict[str, Any]
Row = list[Any]

SHEET_REPORT = "背调报告"
SHEET_CONTACTS = "联系信息"
SHEET_ACTION_PLAN = "开发行动计划"
SHEET_SUMMARY = "背调摘要"
SHEET_PROFILE = "公司画像"
SHEET_WEB = "网页证据"
SHEET_RISKS = "风险待确认"
# Keep the historical internal constant as an alias so older code paths and
# payload comments still point to the business-facing contact sheet.
SHEET_APIFY = SHEET_CONTACTS

MISSING = "未返回"

APIFY_COMPANY_FIELDS: list[tuple[str, str]] = [
    ("company_name", "公司名"),
    ("company_website", "公司官网"),
    ("company_domain", "公司Domain"),
    ("industry", "行业"),
    ("company_size", "员工规模"),
    ("company_linkedin", "公司LinkedIn"),
    ("company_founded_year", "成立年份"),
    ("company_phone", "公司电话"),
    ("company_full_address", "公司地址"),
    ("company_city", "总部城市"),
    ("company_state", "总部州/省"),
    ("company_country", "总部国家"),
    ("company_annual_revenue_clean", "年收入"),
    ("company_total_funding_clean", "融资总额"),
    ("keywords", "关键词"),
    ("company_description", "公司描述"),
    ("company_technologies", "技术栈"),
]

APIFY_CONTACT_FIELDS: list[tuple[str, str]] = [
    ("bd_priority", "BD优先级"),
    ("bd_score", "评分"),
    ("full_name", "姓名"),
    ("job_title", "职位"),
    ("email", "工作邮箱"),
    ("linkedin", "LinkedIn"),
    ("bd_reason", "建议原因"),
    ("seniority_level", "层级"),
    ("functional_level", "职能"),
    ("city", "城市"),
    ("state", "州/省"),
    ("country", "联系人国家"),
]

ACTION_PLAN_HEADERS = ["阶段", "行动项", "建议内容", "对象/负责人", "优先级", "依据/备注"]

REPORT_SECTION_DEFINITIONS: list[tuple[str, list[str]]] = [
    (
        "一、客户类型判断",
        [
            "客户类型",
            "公司规模初判",
            "主营产品",
            "销售模式",
            "店铺数量或渠道数量",
            "社媒粉丝数量",
            "产品价格定位",
            "主要竞争对手或同类企业",
            "采购可能性",
            "对我司的初步价值",
        ],
    ),
    (
        "二、公司介绍",
        [
            "公司名称",
            "国家或地区",
            "官网",
            "地址",
            "电话",
            "邮箱",
            "社媒主页",
            "线上店铺",
            "线下门店数量",
            "销售渠道",
            "公司简介",
            "信息可靠性说明",
        ],
    ),
    (
        "三、公司实力分析",
        [
            "成立时间",
            "经营年限",
            "人员规模",
            "营业额或销售规模线索",
            "品牌影响力",
            "市场覆盖区域",
            "资质证书",
            "展会参与情况",
            "媒体报道情况",
            "Google Maps评分或客户评价",
            "LinkedIn员工数量",
            "社媒活跃度",
            "综合实力判断",
        ],
    ),
    (
        "四、产品线与销售能力",
        [
            "主营产品类型",
            "产品系列",
            "产品材质",
            "产品工艺",
            "产品功能",
            "产品价格区间",
            "折算人民币价格",
            "销售渠道",
            "面向客户群体",
            "产品定位",
            "与我司产品的匹配点",
            "可能存在的产品缺口",
            "可切入的产品机会",
        ],
    ),
    (
        "五、关键决策人员与联络建议",
        [
            "姓名",
            "职位",
            "LinkedIn或社媒主页",
            "公开邮箱",
            "公开电话",
            "可信来源",
            "是否可能参与采购决策",
            "建议联系优先级",
            "替代联系路径",
        ],
    ),
    (
        "六、近3年进口相关数据与采购环境",
        [
            "公司层面进口数据是否找到",
            "近3年进口记录",
            "进口产品",
            "HS Code线索",
            "供应商国家",
            "供应商公司",
            "采购频率",
            "采购规模",
            "可能合作供应链",
            "与我司产品的匹配程度",
            "国家和行业进口趋势",
            "主要进口来源国",
            "中国供应商竞争地位",
            "当地采购特点",
            "价格敏感度",
            "认证、关税、合规要求",
        ],
    ),
    (
        "七、风险与注意事项",
        [
            "公司真实性风险",
            "规模过小风险",
            "采购能力不足风险",
            "价格敏感风险",
            "付款风险",
            "认证合规风险",
            "供应商替换难度",
            "竞争激烈程度",
            "是否可能只是询价比价",
            "是否可能是中间商或采购代理",
            "信息不透明风险",
        ],
    ),
    (
        "八、对【我司】的合作价值与切入建议",
        [
            "合作价值评分（10分制）",
            "推荐开发优先级",
            "最适合切入的产品",
            "最适合使用的卖点",
            "首封开发信角度",
            "报价策略",
            "样品策略",
            "认证或资料准备建议",
            "适合推荐的产品组合",
            "后续跟进节奏",
            "最应该问客户的3到5个问题",
        ],
    ),
    (
        "九、开发话术建议",
        [
            "中文开发思路",
            "英文开发信",
        ],
    ),
    (
        "十、主要来源",
        [
            "来源名称",
            "链接",
            "用于支持什么信息",
            "可信度评价",
        ],
    ),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
OK_FILL = PatternFill("solid", fgColor="D9EAD3")
PRIORITY_FILL = PatternFill("solid", fgColor="E2F0D9")
PRIORITY_BADGE_FILL = PatternFill("solid", fgColor="FFE699")
DEFAULT_FONT = Font(name="Arial", size=10, color="333333")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Arial", size=10, bold=True, color="1F4E78")
PRIORITY_FONT = Font(name="Arial", size=10, bold=True, color="7F6000")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def parse_args(argv: list[str]) -> argparse.Namespace:
    """Parse command-line arguments for report generation.

    Args:
        argv: Command-line arguments without the Python executable name.

    Returns:
        Parsed arguments containing the input JSON path and output XLSX path.

    Raises:
        SystemExit: Raised by argparse when required arguments are missing.
    """

    parser = argparse.ArgumentParser(
        description="Build a company background-check XLSX report."
    )
    parser.add_argument("--input", required=True, help="Normalized report JSON path.")
    parser.add_argument("--output", required=True, help="Output .xlsx path.")
    parser.add_argument(
        "--log",
        help="Optional log file path. Defaults to <output>.log beside the workbook.",
    )
    return parser.parse_args(argv)


def configure_logging(output_path: Path, log_path: str | None) -> None:
    """Configure file logging for repeatable troubleshooting.

    Args:
        output_path: Final workbook path. Used to derive the default log path.
        log_path: Optional explicit log file path.

    Returns:
        None.

    Raises:
        OSError: If the log directory cannot be created or the log file cannot
            be opened by Python's logging module.
    """

    target = Path(log_path) if log_path else output_path.with_suffix(".log")
    target.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        filename=target,
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    logging.info("Starting company background-check XLSX build.")


def require_openpyxl() -> None:
    """Stop early when openpyxl is unavailable.

    Args:
        None.

    Returns:
        None.

    Raises:
        RuntimeError: If openpyxl could not be imported.
    """

    if OPENPYXL_IMPORT_ERROR is not None:
        raise RuntimeError(
            "openpyxl is required before generating .xlsx files."
        ) from OPENPYXL_IMPORT_ERROR


def load_payload(path: Path) -> JsonDict:
    """Load the normalized JSON payload from disk.

    Args:
        path: Path to the JSON file prepared by the agent.

    Returns:
        The parsed JSON object.

    Raises:
        FileNotFoundError: If the input file does not exist.
        json.JSONDecodeError: If the input file is not valid JSON.
        ValueError: If the JSON root is not an object.
    """

    logging.info("Loading payload from %s", path)
    with path.open("r", encoding="utf-8") as file_obj:
        payload = json.load(file_obj)
    if not isinstance(payload, dict):
        raise ValueError("Input JSON root must be an object.")
    return payload


def clean_text(value: Any) -> str:
    """Convert any JSON value into readable single-line cell text.

    Args:
        value: Any JSON-compatible value from the report payload.

    Returns:
        A compact string. Missing values become "未返回" so the workbook never
        implies that absent data was verified.

    Raises:
        No intentional exceptions.
    """

    if value is None:
        return MISSING
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        parts = [clean_text(item) for item in value]
        parts = [part for part in parts if part and part != MISSING]
        return "；".join(parts) if parts else MISSING
    if isinstance(value, dict):
        parts = [
            f"{clean_text(key)}: {clean_text(val)}"
            for key, val in value.items()
            if clean_text(val) != MISSING
        ]
        return "；".join(parts) if parts else MISSING
    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return MISSING
    # Keep untrusted website/Actor text literal instead of letting Excel
    # interpret formula-like prefixes.
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def shorten_text(text: str, max_chars: int) -> str:
    """Shorten very long text so worksheet sections stay readable.

    Args:
        text: Already normalized text.
        max_chars: Maximum number of characters to keep before adding an
            ellipsis-style note.

    Returns:
        Original text when short enough, otherwise a concise preview.

    Raises:
        No intentional exceptions.
    """

    if len(text) <= max_chars:
        return text
    return f"{text[:max_chars].rstrip()}..."


def as_list(value: Any) -> list[Any]:
    """Normalize a payload field into a list.

    Args:
        value: A list, dict, scalar, or null value.

    Returns:
        A list version of the value. Null becomes an empty list, and scalar
        values become a one-item list.

    Raises:
        No intentional exceptions.
    """

    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def get_dict(payload: JsonDict, key: str) -> JsonDict:
    """Return a dictionary field without forcing callers to type-check each use.

    Args:
        payload: Parent JSON object.
        key: Field name to read.

    Returns:
        The dictionary at `key`, or an empty dictionary when the field is absent
        or has another type.

    Raises:
        No intentional exceptions.
    """

    value = payload.get(key)
    return value if isinstance(value, dict) else {}


def append_rows(sheet: Worksheet, rows: list[Row], header: bool = False) -> None:
    """Append rows and apply the shared workbook style.

    Args:
        sheet: Target worksheet.
        rows: Two-dimensional row data.
        header: Whether the first row should be styled as a header row.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if the worksheet cannot accept data.
    """

    for row_index, row in enumerate(rows, start=1):
        sheet.append([clean_text(cell) for cell in row])
        # `sheet.append()` writes after the last existing row. Styling by
        # `row_index` alone would accidentally restyle title rows on sheets
        # that already have a header band, so use the real appended row number.
        appended_row = sheet.max_row
        for cell in sheet[appended_row]:
            cell.font = DEFAULT_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header and row_index == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_column_widths(sheet: Worksheet, widths: dict[str, int]) -> None:
    """Set stable column widths for a worksheet.

    Args:
        sheet: Target worksheet.
        widths: Mapping from Excel column letter to width.

    Returns:
        None.

    Raises:
        No intentional exceptions.
    """

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def write_title(sheet: Worksheet, title: str, subtitle: str = "", end_column: int = 6) -> None:
    """Write a merged title band at the top of a worksheet.

    Args:
        sheet: Target worksheet.
        title: Main sheet title.
        subtitle: Optional subtitle shown below the title.
        end_column: Last column included in the merged title band.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if merging or writing fails.
    """

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.fill = HEADER_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 26
    if subtitle:
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
        subtitle_cell = sheet.cell(row=2, column=1, value=subtitle)
        subtitle_cell.fill = SUBHEADER_FILL
        subtitle_cell.font = DEFAULT_FONT
        subtitle_cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[2].height = 22


def write_key_value_block(sheet: Worksheet, start_row: int, pairs: list[tuple[str, Any]]) -> int:
    """Write a two-column label/value block.

    Args:
        sheet: Target worksheet.
        start_row: First row where the block should be written.
        pairs: Label and value pairs to render.

    Returns:
        The next empty row after the block.

    Raises:
        openpyxl exceptions may propagate if cell writes fail.
    """

    row = start_row
    for label, value in pairs:
        label_cell = sheet.cell(row=row, column=1, value=label)
        value_cell = sheet.cell(row=row, column=2, value=clean_text(value))
        label_cell.font = LABEL_FONT
        label_cell.fill = SUBHEADER_FILL
        value_cell.font = DEFAULT_FONT
        for cell in (label_cell, value_cell):
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    return row


def write_wide_key_value_block(
    sheet: Worksheet,
    start_row: int,
    pairs: list[tuple[str, Any]],
    end_column: int = 7,
) -> int:
    """Write label/value rows where the value spans the sheet width.

    Args:
        sheet: Target worksheet.
        start_row: First row where the block should be written.
        pairs: Label and value pairs to render.
        end_column: Last column used by the merged value cell.

    Returns:
        The next empty row after the block.

    Raises:
        openpyxl exceptions may propagate if cells cannot be merged or written.
    """

    row = start_row
    for label, value in pairs:
        label_cell = sheet.cell(row=row, column=1, value=label)
        value_cell = sheet.cell(row=row, column=2, value=clean_text(value))
        if end_column > 2:
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end_column)
        label_cell.font = LABEL_FONT
        label_cell.fill = SUBHEADER_FILL
        value_cell.font = DEFAULT_FONT
        value_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        label_cell.border = THIN_BORDER
        value_cell.border = THIN_BORDER
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)
        text_length = len(clean_text(value))
        if text_length > 180:
            sheet.row_dimensions[row].height = 52
        elif text_length > 80:
            sheet.row_dimensions[row].height = 34
        row += 1
    return row


def write_section_header(sheet: Worksheet, row: int, title: str, end_column: int = 6) -> int:
    """Write a visual section header inside a worksheet.

    Args:
        sheet: Target worksheet.
        row: Row number where the section header should appear.
        title: Header text shown to the business user.
        end_column: Last column included in the merged header band.

    Returns:
        The next row after the section header.

    Raises:
        openpyxl exceptions may propagate if merging or styling fails.
    """

    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = SUBHEADER_FILL
    cell.font = LABEL_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 22
    return row + 1


def status_fill(value: Any) -> PatternFill | None:
    """Choose a fill color for confidence, status, or risk labels.

    Args:
        value: Cell value to inspect.

    Returns:
        An openpyxl PatternFill for known labels, otherwise None.

    Raises:
        No intentional exceptions.
    """

    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return None
    if any(label in text for label in ("FAILED", "高风险", "失败", "冲突")):
        return ERROR_FILL
    if any(label in text for label in ("中风险", "待判断", "未返回")) or text == "中":
        return WARNING_FILL
    if "优先联系" in text:
        return PRIORITY_BADGE_FILL
    if any(label in text for label in ("高置信", "高匹配", "SUCCEEDED", "低风险")) or text == "低":
        return OK_FILL
    if text == "高":
        return OK_FILL
    return None


def apply_status_styles(sheet: Worksheet) -> None:
    """Apply status colors to cells that contain risk or confidence labels.

    Args:
        sheet: Worksheet to style.

    Returns:
        None.

    Raises:
        No intentional exceptions.
    """

    for row in sheet.iter_rows():
        for cell in row:
            fill = status_fill(cell.value)
            if fill is not None:
                cell.fill = fill


def normalize_label(value: Any) -> str:
    """Normalize section and field labels for tolerant matching.

    Args:
        value: A label from the report payload.

    Returns:
        A compact lowercase label without Chinese section numbering.

    Raises:
        No intentional exceptions.
    """

    text = clean_text(value)
    text = re.sub(r"^[一二三四五六七八九十]+[、.．]\s*", "", text)
    return re.sub(r"\s+", "", text).lower()


def read_report_item(value: Any) -> tuple[str, str, str]:
    """Extract content, basis, and information type from one report item.

    Args:
        value: A scalar, list, or dict from `report_sections`.

    Returns:
        `(content, basis, info_type)` ready for worksheet rendering.

    Raises:
        No intentional exceptions.
    """

    if isinstance(value, dict):
        content = (
            value.get("answer")
            or value.get("value")
            or value.get("content")
            or value.get("analysis")
            or value.get("text")
        )
        basis = (
            value.get("basis")
            or value.get("evidence")
            or value.get("source")
            or value.get("source_url")
            or value.get("url")
        )
        info_type = (
            value.get("info_type")
            or value.get("information_type")
            or value.get("certainty")
            or value.get("confidence")
        )
        if content is None:
            content = {
                key: val
                for key, val in value.items()
                if key
                not in {
                    "basis",
                    "evidence",
                    "source",
                    "source_url",
                    "url",
                    "info_type",
                    "information_type",
                    "certainty",
                    "confidence",
                }
            }
        return clean_text(content), clean_text(basis), clean_text(info_type)
    return clean_text(value), MISSING, MISSING


def item_list_to_mapping(items: Any) -> dict[str, tuple[str, str, str]]:
    """Convert section item data into a field mapping.

    Args:
        items: Either a list of item dicts or a dict keyed by field name.

    Returns:
        Mapping from normalized field label to `(content, basis, info_type)`.

    Raises:
        No intentional exceptions.
    """

    mapping: dict[str, tuple[str, str, str]] = {}
    if isinstance(items, dict):
        items = items.get("items") if isinstance(items.get("items"), list) else items
    if isinstance(items, list):
        for item in items:
            if isinstance(item, dict):
                label = item.get("item") or item.get("field") or item.get("name") or item.get("label")
                if label is None:
                    continue
                mapping[normalize_label(label)] = read_report_item(item)
        return mapping
    if isinstance(items, dict):
        for label, value in items.items():
            mapping[normalize_label(label)] = read_report_item(value)
    return mapping


def normalize_report_sections(payload: JsonDict) -> dict[str, dict[str, tuple[str, str, str]]]:
    """Normalize flexible report payload shapes into section mappings.

    Args:
        payload: Full normalized report payload.

    Returns:
        Mapping from normalized section title to normalized item mappings.

    Raises:
        No intentional exceptions.
    """

    sections: dict[str, dict[str, tuple[str, str, str]]] = {}
    report_sections = payload.get("report_sections")
    if isinstance(report_sections, list):
        for section in report_sections:
            if not isinstance(section, dict):
                continue
            title = section.get("title") or section.get("section")
            if title is None:
                continue
            sections[normalize_label(title)] = item_list_to_mapping(section.get("items") or section)

    structured_report = payload.get("structured_report")
    if isinstance(structured_report, dict):
        for title, items in structured_report.items():
            sections[normalize_label(title)] = item_list_to_mapping(items)
    return sections


def get_apify_items(payload: JsonDict) -> list[JsonDict]:
    """Return raw Apify items as dictionaries.

    Args:
        payload: Full normalized report payload.

    Returns:
        Apify rows that can be used for fallback report rendering.

    Raises:
        No intentional exceptions.
    """

    apify = get_dict(payload, "apify")
    items: list[JsonDict] = []
    for item in as_list(apify.get("raw_items")):
        if isinstance(item, dict):
            items.append(item)
    return items


def make_report_source_summary(payload: JsonDict) -> str:
    """Summarize source URLs for the main report sheet.

    Args:
        payload: Full normalized report payload.

    Returns:
        A compact source list.

    Raises:
        No intentional exceptions.
    """

    sources: list[str] = []
    for item in as_list(payload.get("web_evidence")):
        if isinstance(item, dict):
            title = clean_text(item.get("title") or item.get("source_type"))
            url = clean_text(item.get("url"))
            if url != MISSING:
                sources.append(f"{title}: {url}")
    return "；".join(sources[:8]) if sources else MISSING


def fallback_report_sections(payload: JsonDict) -> dict[str, dict[str, tuple[str, str, str]]]:
    """Create a basic ten-section report when no explicit report is supplied.

    Args:
        payload: Full normalized report payload.

    Returns:
        Section mapping compatible with `make_report_sheet()`.

    Raises:
        No intentional exceptions.
    """

    report_input = get_dict(payload, "input")
    confirmed = get_dict(payload, "confirmed_company")
    summary = get_dict(payload, "summary")
    profile = get_dict(payload, "company_profile")
    apify_items = get_apify_items(payload)
    company_pairs = dict(extract_apify_company_pairs(apify_items)) if apify_items else {}
    contacts = build_apify_contact_records(apify_items)[:5] if apify_items else []
    source_summary = make_report_source_summary(payload)
    contact_summary = "；".join(
        f"{row['full_name']} - {row['job_title']} - {row['email']} - {row['bd_priority']}"
        for row in contacts
    )
    import_note = "未找到公司层面进口数据；当前仅可基于国家、行业和公开资料做谨慎判断。"

    raw_sections: dict[str, dict[str, Any]] = {
        "一、客户类型判断": {
            "客户类型": summary.get("customer_type") or "未找到可靠公开信息",
            "公司规模初判": company_pairs.get("员工规模") or profile.get("scale") or "未找到可靠公开信息",
            "主营产品": profile.get("business") or "未找到可靠公开信息",
            "销售模式": profile.get("sales_model") or "未找到可靠公开信息",
            "店铺数量或渠道数量": profile.get("channels") or "未找到可靠公开信息",
            "社媒粉丝数量": profile.get("social_followers") or "未找到可靠公开信息",
            "产品价格定位": profile.get("price_positioning") or "未找到可靠公开信息",
            "主要竞争对手或同类企业": summary.get("competitors") or "未找到可靠公开信息",
            "采购可能性": summary.get("purchase_probability") or "需要结合我司产品和客户产品线进一步判断",
            "对我司的初步价值": summary.get("value_to_our_company") or "未找到我司信息，需补充我司产品后判断",
        },
        "二、公司介绍": {
            "公司名称": confirmed.get("official_name") or report_input.get("company_name"),
            "国家或地区": report_input.get("country"),
            "官网": confirmed.get("official_domain") or company_pairs.get("公司官网"),
            "地址": profile.get("address") or company_pairs.get("公司地址"),
            "电话": profile.get("phone") or company_pairs.get("公司电话"),
            "邮箱": profile.get("email") or "未找到可靠公开信息",
            "社媒主页": profile.get("social_profiles") or company_pairs.get("公司LinkedIn"),
            "线上店铺": profile.get("online_stores") or "未找到可靠公开信息",
            "线下门店数量": profile.get("offline_stores") or "未找到可靠公开信息",
            "销售渠道": profile.get("sales_channels") or "未找到可靠公开信息",
            "公司简介": profile.get("notes") or company_pairs.get("公司描述") or summary.get("conclusion"),
            "信息可靠性说明": "官网、公开网页证据和 Apify 返回字段交叉整理；缺失项已标记。",
        },
        "三、公司实力分析": {
            "成立时间": company_pairs.get("成立年份") or "未找到可靠公开信息",
            "经营年限": "可由成立年份推算；若未找到成立年份则不可判断",
            "人员规模": company_pairs.get("员工规模") or "未找到可靠公开信息",
            "营业额或销售规模线索": company_pairs.get("年收入") or "未找到可靠公开信息",
            "品牌影响力": summary.get("brand_influence") or "未找到可靠公开信息",
            "市场覆盖区域": profile.get("locations") or "未找到可靠公开信息",
            "资质证书": profile.get("certifications") or "未找到可靠公开信息",
            "展会参与情况": profile.get("trade_shows") or "未找到可靠公开信息",
            "媒体报道情况": profile.get("media_coverage") or "未找到可靠公开信息",
            "Google Maps评分或客户评价": profile.get("google_reviews") or "未找到可靠公开信息",
            "LinkedIn员工数量": company_pairs.get("员工规模") or "未找到可靠公开信息",
            "社媒活跃度": profile.get("social_activity") or "未找到可靠公开信息",
            "综合实力判断": summary.get("strength_judgment") or summary.get("conclusion") or "未找到可靠公开信息",
        },
        "四、产品线与销售能力": {
            "主营产品类型": profile.get("business") or "未找到可靠公开信息",
            "产品系列": profile.get("product_series") or "未找到可靠公开信息",
            "产品材质": profile.get("materials") or "未找到可靠公开信息",
            "产品工艺": profile.get("process") or "未找到可靠公开信息",
            "产品功能": profile.get("features") or "未找到可靠公开信息",
            "产品价格区间": profile.get("price_range") or "未找到可靠公开信息",
            "折算人民币价格": profile.get("price_cny") or "未找到可靠公开信息",
            "销售渠道": profile.get("sales_channels") or "未找到可靠公开信息",
            "面向客户群体": profile.get("target_customers") or "未找到可靠公开信息",
            "产品定位": profile.get("positioning") or "未找到可靠公开信息",
            "与我司产品的匹配点": "未找到我司信息，需补充我司产品后判断",
            "可能存在的产品缺口": "未找到可靠公开信息",
            "可切入的产品机会": "未找到我司信息，需补充我司产品后判断",
        },
        "五、关键决策人员与联络建议": {
            "姓名": contact_summary or "未找到可靠公开信息",
            "职位": f"见 {SHEET_CONTACTS} Sheet 的联系人明细和优先级",
            "LinkedIn或社媒主页": f"见 {SHEET_CONTACTS} Sheet",
            "公开邮箱": f"见 {SHEET_CONTACTS} Sheet；没有公开邮箱的联系人不要编造",
            "公开电话": profile.get("phone") or company_pairs.get("公司电话") or "未找到可靠公开信息",
            "可信来源": "Apify Actor 输出、LinkedIn/公开网页证据",
            "是否可能参与采购决策": f"根据职位、层级和职能做合理推测，见 {SHEET_CONTACTS} Sheet 的建议原因",
            "建议联系优先级": f"见 {SHEET_CONTACTS} Sheet 的 BD优先级",
            "替代联系路径": profile.get("contacts") or "官网表单、公司邮箱、LinkedIn 私信、社媒主页、电话总机",
        },
        "六、近3年进口相关数据与采购环境": {
            "公司层面进口数据是否找到": "未找到可靠公开信息",
            "近3年进口记录": import_note,
            "进口产品": "未找到可靠公开信息",
            "HS Code线索": "未找到可靠公开信息",
            "供应商国家": "未找到可靠公开信息",
            "供应商公司": "未找到可靠公开信息",
            "采购频率": "未找到可靠公开信息",
            "采购规模": "未找到可靠公开信息",
            "可能合作供应链": "未找到可靠公开信息",
            "与我司产品的匹配程度": "未找到我司信息，需补充我司产品后判断",
            "国家和行业进口趋势": "需要额外 web_search / web_fetch 获取国家与行业层面数据",
            "主要进口来源国": "未找到可靠公开信息",
            "中国供应商竞争地位": "未找到可靠公开信息",
            "当地采购特点": "未找到可靠公开信息",
            "价格敏感度": "未找到可靠公开信息",
            "认证、关税、合规要求": "未找到可靠公开信息",
        },
        "七、风险与注意事项": {
            "公司真实性风险": summary.get("risk_level") or "待判断",
            "规模过小风险": "需结合人员规模、销售渠道、进口数据判断",
            "采购能力不足风险": "需结合营业额、门店/渠道、进口记录判断",
            "价格敏感风险": "需结合产品价格定位和市场竞争判断",
            "付款风险": "未找到可靠公开信息",
            "认证合规风险": "需核对目标国家和产品认证要求",
            "供应商替换难度": "未找到可靠公开信息",
            "竞争激烈程度": "需结合竞品和同类企业判断",
            "是否可能只是询价比价": "待业务沟通验证",
            "是否可能是中间商或采购代理": "需结合销售模式和公司介绍判断",
            "信息不透明风险": "公开信息缺失项越多，风险越高",
        },
        "八、对【我司】的合作价值与切入建议": {
            "合作价值评分（10分制）": "未找到我司信息，需 ask_user 补充后评分",
            "推荐开发优先级": "未找到我司信息，需 ask_user 补充后判断",
            "最适合切入的产品": "未找到我司信息",
            "最适合使用的卖点": "未找到我司信息",
            "首封开发信角度": "先围绕客户主营业务和现有渠道做轻量触达",
            "报价策略": "未找到我司信息",
            "样品策略": "未找到我司信息",
            "认证或资料准备建议": "按目标国家和客户产品类别准备",
            "适合推荐的产品组合": "未找到我司信息",
            "后续跟进节奏": "首封后 3-5 个工作日跟进；再根据回复调整",
            "最应该问客户的3到5个问题": "当前采购品类、目标价格、认证要求、采购量、现有供应商痛点",
        },
        "九、开发话术建议": {
            "中文开发思路": "围绕客户主营业务、渠道和可能产品缺口切入；避免泛泛推销。",
            "英文开发信": "需要结合我司产品信息生成更精准版本；未找到我司信息时不可过度承诺。",
        },
        "十、主要来源": {
            "来源名称": source_summary,
            "链接": source_summary,
            "用于支持什么信息": "官网确认、公司介绍、公开证据和联系人线索",
            "可信度评价": "官网/官方社媒高；第三方目录中；未交叉验证来源低",
        },
    }
    return {
        normalize_label(title): {
            normalize_label(field): (clean_text(value), MISSING, "公开信息/合理推测/待确认")
            for field, value in fields.items()
        }
        for title, fields in raw_sections.items()
    }


def make_report_sheet(workbook: Workbook, payload: JsonDict) -> None:
    """Create the ten-section business-development report sheet.

    Args:
        workbook: Workbook being built.
        payload: Normalized report payload.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if worksheet operations fail.
    """

    sheet = workbook.active
    sheet.title = SHEET_REPORT
    write_title(sheet, "客户背调报告", "按 B2B 开发流程整理：公开信息、合理推测和来源分开标注", end_column=4)
    set_column_widths(sheet, {"A": 24, "B": 82, "C": 52, "D": 18})

    explicit_sections = normalize_report_sections(payload)
    sections = explicit_sections or fallback_report_sections(payload)
    row = 4
    for title, fields in REPORT_SECTION_DEFINITIONS:
        row = write_section_header(sheet, row, title, end_column=4)
        header_row = row
        for column, header in enumerate(["项目", "内容", "依据/来源", "信息属性"], start=1):
            cell = sheet.cell(row=header_row, column=column, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

        section_mapping = sections.get(normalize_label(title), {})
        rendered_keys: set[str] = set()
        for field in fields:
            key = normalize_label(field)
            content, basis, info_type = section_mapping.get(
                key,
                ("未找到可靠公开信息", "未找到可靠公开信息", "未找到可靠公开信息"),
            )
            rendered_keys.add(key)
            values = [field, content, basis, info_type]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row, column=column, value=clean_text(value))
                cell.font = LABEL_FONT if column == 1 else DEFAULT_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            max_text = max(len(clean_text(value)) for value in values)
            if max_text > 180:
                sheet.row_dimensions[row].height = 58
            elif max_text > 80:
                sheet.row_dimensions[row].height = 38
            row += 1

        for raw_key, (content, basis, info_type) in section_mapping.items():
            if raw_key in rendered_keys:
                continue
            values = [raw_key, content, basis, info_type]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row, column=column, value=clean_text(value))
                cell.font = LABEL_FONT if column == 1 else DEFAULT_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
        row += 1

    sheet.freeze_panes = "A4"
    apply_status_styles(sheet)


def make_summary_sheet(workbook: Workbook, payload: JsonDict) -> None:
    """Create the report summary sheet.

    Args:
        workbook: Workbook being built.
        payload: Normalized report payload.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if worksheet operations fail.
    """

    sheet = workbook.create_sheet(SHEET_SUMMARY)
    report_input = get_dict(payload, "input")
    confirmed = get_dict(payload, "confirmed_company")
    summary = get_dict(payload, "summary")
    apify = get_dict(payload, "apify")
    write_title(sheet, "公司背调摘要", "公开网页证据 + Apify Actor 结果合并")
    next_row = write_key_value_block(
        sheet,
        4,
        [
            ("用户输入国家/地区", report_input.get("country")),
            ("用户输入公司名", report_input.get("company_name")),
            ("确认公司名", confirmed.get("official_name")),
            ("确认官网", confirmed.get("official_domain")),
            ("官网确认置信度", confirmed.get("confidence")),
            ("国家匹配度", confirmed.get("country_match")),
            ("Apify状态", apify.get("status")),
            ("风险等级", summary.get("risk_level")),
            ("一句话结论", summary.get("conclusion")),
        ],
    )

    findings = as_list(summary.get("key_findings"))
    if findings:
        sheet.cell(row=next_row + 1, column=1, value="重点发现").font = LABEL_FONT
        sheet.cell(row=next_row + 1, column=1).fill = SUBHEADER_FILL
        for offset, finding in enumerate(findings, start=1):
            sheet.cell(row=next_row + 1 + offset, column=1, value=offset)
            sheet.cell(row=next_row + 1 + offset, column=2, value=clean_text(finding))

    set_column_widths(sheet, {"A": 22, "B": 80, "C": 18, "D": 18, "E": 18, "F": 18})
    apply_status_styles(sheet)


def make_profile_sheet(workbook: Workbook, payload: JsonDict) -> None:
    """Create the company profile sheet.

    Args:
        workbook: Workbook being built.
        payload: Normalized report payload.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if worksheet operations fail.
    """

    sheet = workbook.create_sheet(SHEET_PROFILE)
    confirmed = get_dict(payload, "confirmed_company")
    profile = get_dict(payload, "company_profile")
    report_input = get_dict(payload, "input")
    write_title(sheet, "公司画像", "只展示有来源支撑的信息，缺失项标为未返回")
    write_key_value_block(
        sheet,
        4,
        [
            ("公司名", confirmed.get("official_name") or report_input.get("company_name")),
            ("国家/地区", report_input.get("country")),
            ("官网", confirmed.get("official_domain")),
            ("主营业务", profile.get("business")),
            ("地区/办公室", profile.get("locations")),
            ("联系方式", profile.get("contacts")),
            ("管理层/团队", profile.get("management")),
            ("客户/项目/案例", profile.get("projects_or_clients")),
            ("其他备注", profile.get("notes")),
        ],
    )
    set_column_widths(sheet, {"A": 22, "B": 90, "C": 18, "D": 18, "E": 18, "F": 18})


def make_web_sheet(workbook: Workbook, payload: JsonDict) -> None:
    """Create the web evidence sheet.

    Args:
        workbook: Workbook being built.
        payload: Normalized report payload.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if worksheet operations fail.
    """

    sheet = workbook.create_sheet(SHEET_WEB)
    write_title(sheet, "网页证据", "web_search / web_fetch 收集到的公开来源")
    headers = ["来源类型", "标题", "URL", "事实摘要", "置信度"]
    rows = [headers]
    for item in as_list(payload.get("web_evidence")):
        item_dict = item if isinstance(item, dict) else {"fact": item}
        rows.append(
            [
                item_dict.get("source_type"),
                item_dict.get("title"),
                item_dict.get("url"),
                item_dict.get("fact") or item_dict.get("summary"),
                item_dict.get("confidence"),
            ]
        )
    if len(rows) == 1:
        rows.append([MISSING, MISSING, MISSING, "没有返回可用网页证据", MISSING])
    table_start = sheet.max_row + 1
    append_rows(sheet, rows, header=True)
    set_column_widths(sheet, {"A": 18, "B": 34, "C": 52, "D": 80, "E": 12})
    sheet.auto_filter.ref = f"A{table_start}:E{sheet.max_row}"
    sheet.freeze_panes = "A4"
    apply_status_styles(sheet)


def flatten_dict(value: Any, prefix: str = "") -> JsonDict:
    """Flatten nested JSON so arbitrary Apify output can fit into a table.

    Args:
        value: Raw JSON value from Apify output.
        prefix: Internal key prefix used during recursion.

    Returns:
        A flat dictionary where nested keys use dot notation.

    Raises:
        No intentional exceptions.
    """

    if isinstance(value, dict):
        flattened: JsonDict = {}
        for key, child in value.items():
            child_key = f"{prefix}.{key}" if prefix else str(key)
            flattened.update(flatten_dict(child, child_key))
        return flattened
    if isinstance(value, list):
        if all(not isinstance(item, (dict, list)) for item in value):
            return {prefix: clean_text(value)}
        return {prefix: json.dumps(value, ensure_ascii=False)}
    return {prefix: clean_text(value)}


def analysis_text(value: Any) -> str:
    """Convert a value into plain text for scoring and deduplication.

    Args:
        value: A raw value from an Apify item.

    Returns:
        Plain text with missing values represented as an empty string. This is
        different from `clean_text()`, because ranking logic should not treat
        the Chinese missing-value marker as real source data.

    Raises:
        No intentional exceptions.
    """

    if value is None or value == "" or value == [] or value == {}:
        return ""
    if isinstance(value, (list, dict)):
        return clean_text(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def most_common_nonempty_value(items: list[JsonDict], field: str) -> str:
    """Return the most common non-empty value for one field.

    Args:
        items: Apify rows normalized as dictionaries.
        field: Field name to inspect.

    Returns:
        The most common non-empty value. Empty fields return "未返回".

    Raises:
        No intentional exceptions.
    """

    counts: dict[str, int] = {}
    first_seen_order: list[str] = []
    for item in items:
        text = analysis_text(item.get(field))
        if not text:
            continue
        if text not in counts:
            counts[text] = 0
            first_seen_order.append(text)
        counts[text] += 1
    if not counts:
        return MISSING
    return max(first_seen_order, key=lambda candidate: counts[candidate])


def extract_apify_company_pairs(items: list[JsonDict]) -> list[tuple[str, str]]:
    """Extract company-level fields once instead of repeating them per contact.

    Args:
        items: Raw Apify result rows. These rows usually repeat the same company
            profile values for every contact.

    Returns:
        Label/value pairs suitable for the top block of the Apify worksheet.

    Raises:
        No intentional exceptions.
    """

    pairs: list[tuple[str, str]] = []
    preview_limits = {
        "keywords": 260,
        "company_description": 360,
        "company_technologies": 220,
    }
    for field, label in APIFY_COMPANY_FIELDS:
        value = most_common_nonempty_value(items, field)
        if value != MISSING:
            if field in preview_limits:
                value = shorten_text(value, preview_limits[field])
            pairs.append((label, value))
    return pairs


def normalize_company_domain(value: Any) -> str:
    """Normalize a company domain or URL for business-email matching.

    Args:
        value: Domain-like value such as `https://www.example.com/`.

    Returns:
        Lowercase hostname without scheme, path, or a leading `www.`.

    Raises:
        No intentional exceptions.
    """

    text = analysis_text(value).lower()
    text = re.sub(r"^https?://", "", text)
    text = text.split("/", 1)[0]
    text = text.removeprefix("www.")
    return text


def score_b2b_contact(item: JsonDict, company_domain: str) -> tuple[str, int, str]:
    """Score one contact for B2B business-development outreach.

    Args:
        item: One Apify contact row.
        company_domain: Normalized company domain used to identify work email.

    Returns:
        A tuple of priority label, numeric score, and readable reason.

    Raises:
        No intentional exceptions.
    """

    score = 0
    reasons: list[str] = []
    email = analysis_text(item.get("email")).lower()
    linkedin = analysis_text(item.get("linkedin"))
    title = analysis_text(item.get("job_title")).lower()
    headline = analysis_text(item.get("headline")).lower()
    seniority = analysis_text(item.get("seniority_level")).lower()
    function = analysis_text(item.get("functional_level")).lower()
    title_blob = f"{title} {headline}"

    if email:
        score += 25
        reasons.append("有工作邮箱")
        if company_domain and company_domain in email.split("@")[-1]:
            score += 10
            reasons.append("邮箱域名匹配")
    else:
        score -= 20
        reasons.append("缺少工作邮箱")

    if linkedin:
        score += 10
        reasons.append("有LinkedIn")

    if seniority == "c_suite" or re.search(r"\b(ceo|founder|co-founder|chief|president)\b", title_blob):
        score += 40
        reasons.append("决策层/创始人")
    elif seniority in {"vp", "director", "head"} or re.search(r"\b(vp|vice president|director|head)\b", title_blob):
        score += 30
        reasons.append("负责人层级")
    elif seniority == "manager" or "manager" in title_blob:
        score += 20
        reasons.append("经理层级")
    elif seniority == "senior" or "senior" in title_blob:
        score += 10
        reasons.append("资深岗位")

    if function in {"sales", "business_development"}:
        score += 30
        reasons.append("销售/BD职能")
    elif function == "c_suite":
        score += 25
        reasons.append("公司决策职能")
    elif function in {"operations", "product", "finance"}:
        score += 18
        reasons.append("业务/运营相关职能")
    elif function == "engineering":
        score += 12
        reasons.append("技术项目相关职能")
    elif function in {"human_resources", "administrative"}:
        score -= 30
        reasons.append("非BD优先职能")

    if re.search(r"\b(business development|commercial|sales|partnership|partner|procurement|sourcing|purchasing|supply chain)\b", title_blob):
        score += 30
        reasons.append("职位与BD/采购/合作相关")
    if re.search(r"\b(project|operations|engineering|technical|design)\b", title_blob):
        score += 12
        reasons.append("项目/技术对接价值")
    if re.search(r"\b(hr|human resources|talent|recruiter|people)\b", title_blob):
        score -= 35
        reasons.append("HR类岗位暂不优先")
    if re.search(r"\b(intern|student|assistant)\b", title_blob):
        score -= 20
        reasons.append("岗位影响力较低")

    if score >= 80:
        priority = "优先联系"
    elif score >= 50:
        priority = "可作为备选"
    else:
        priority = "暂不优先"
    return priority, score, "；".join(reasons[:5]) if reasons else "未返回足够判断依据"


def build_apify_contact_records(items: list[JsonDict]) -> list[JsonDict]:
    """Build a compact contact table with BD priority columns.

    Args:
        items: Raw Apify result rows.

    Returns:
        Contact rows sorted by outreach priority. Company-level duplicate fields
        are intentionally omitted because they are shown once in the top block.

    Raises:
        No intentional exceptions.
    """

    company_domain = normalize_company_domain(most_common_nonempty_value(items, "company_domain"))
    records: list[JsonDict] = []
    for item in items:
        priority, score, reason = score_b2b_contact(item, company_domain)
        name = analysis_text(item.get("full_name"))
        if not name:
            name = " ".join(
                part
                for part in [analysis_text(item.get("first_name")), analysis_text(item.get("last_name"))]
                if part
            )
        records.append(
            {
                "bd_priority": priority,
                "bd_score": score,
                "bd_reason": reason,
                "full_name": name or MISSING,
                "job_title": analysis_text(item.get("job_title")) or MISSING,
                "email": analysis_text(item.get("email")) or MISSING,
                "linkedin": analysis_text(item.get("linkedin")) or MISSING,
                "seniority_level": analysis_text(item.get("seniority_level")) or MISSING,
                "functional_level": analysis_text(item.get("functional_level")) or MISSING,
                "city": analysis_text(item.get("city")) or MISSING,
                "state": analysis_text(item.get("state")) or MISSING,
                "country": analysis_text(item.get("country")) or MISSING,
            }
        )
    priority_order = {"优先联系": 0, "可作为备选": 1, "暂不优先": 2}
    return sorted(
        records,
        key=lambda row: (
            priority_order.get(str(row["bd_priority"]), 9),
            -int(row["bd_score"]),
            str(row["full_name"]),
        ),
    )


def report_section_value(
    sections: dict[str, dict[str, tuple[str, str, str]]],
    section_title: str,
    field_label: str,
    default: str = MISSING,
) -> str:
    """Read one rendered value from the normalized ten-section report.

    Args:
        sections: Normalized report sections produced from explicit report data
            or from fallback report generation.
        section_title: Human-readable section name, such as
            `八、对【我司】的合作价值与切入建议`.
        field_label: Human-readable field name inside the section.
        default: Value returned when the section or field is unavailable.

    Returns:
        The content cell for the requested report item.

    Raises:
        No intentional exceptions.
    """

    section = sections.get(normalize_label(section_title), {})
    content, _, _ = section.get(normalize_label(field_label), (default, MISSING, MISSING))
    return content if content and content != MISSING else default


def format_contact_for_action(row: JsonDict) -> str:
    """Format one contact record for the action-plan sheet.

    Args:
        row: Contact row generated by `build_apify_contact_records()`.

    Returns:
        A concise contact string containing name, title, email, LinkedIn, and
        the ranking reason when available.

    Raises:
        No intentional exceptions.
    """

    parts = [
        clean_text(row.get("full_name")),
        clean_text(row.get("job_title")),
        clean_text(row.get("email")),
        clean_text(row.get("linkedin")),
    ]
    visible_parts = [part for part in parts if part != MISSING]
    reason = clean_text(row.get("bd_reason"))
    if reason != MISSING:
        visible_parts.append(f"原因：{reason}")
    return " | ".join(visible_parts) if visible_parts else MISSING


def explicit_action_plan_rows(payload: JsonDict) -> list[Row]:
    """Convert optional `action_plan` payload rows into worksheet rows.

    Args:
        payload: Full normalized report payload. The optional `action_plan`
            field may be a list of dictionaries/lists, or a dictionary with a
            `rows` or `items` list.

    Returns:
        Rows following `ACTION_PLAN_HEADERS`. Empty means no explicit action
        plan was supplied and the script should build a fallback plan.

    Raises:
        No intentional exceptions.
    """

    action_plan = payload.get("action_plan")
    if isinstance(action_plan, dict):
        raw_rows = action_plan.get("rows") or action_plan.get("items") or []
    else:
        raw_rows = action_plan
    if not isinstance(raw_rows, list):
        return []

    rows: list[Row] = []
    for item in raw_rows:
        if isinstance(item, dict):
            rows.append(
                [
                    item.get("stage") or item.get("阶段"),
                    item.get("action") or item.get("行动项"),
                    item.get("recommendation") or item.get("建议内容") or item.get("suggestion"),
                    item.get("owner") or item.get("对象/负责人") or item.get("target"),
                    item.get("priority") or item.get("优先级"),
                    item.get("basis") or item.get("依据/备注") or item.get("note"),
                ]
            )
        elif isinstance(item, list):
            padded = item[: len(ACTION_PLAN_HEADERS)] + [MISSING] * len(ACTION_PLAN_HEADERS)
            rows.append(padded[: len(ACTION_PLAN_HEADERS)])
    return rows


def build_fallback_action_plan_rows(payload: JsonDict) -> list[Row]:
    """Build a practical BD action plan from report sections and contacts.

    Args:
        payload: Full normalized report payload.

    Returns:
        Rows for the `开发行动计划` sheet. The sheet is designed for business
        users who want to know what to do next after reading the full report.

    Raises:
        No intentional exceptions.
    """

    report_input = get_dict(payload, "input")
    confirmed = get_dict(payload, "confirmed_company")
    apify = get_dict(payload, "apify")
    sections = normalize_report_sections(payload) or fallback_report_sections(payload)
    apify_items = get_apify_items(payload)
    contacts = build_apify_contact_records(apify_items) if apify_items else []
    priority_contacts = [row for row in contacts if row.get("bd_priority") == "优先联系"]
    primary_contact = priority_contacts[0] if priority_contacts else (contacts[0] if contacts else {})
    backup_contacts = [
        format_contact_for_action(row)
        for row in contacts
        if row is not primary_contact and row.get("bd_priority") in {"优先联系", "可作为备选"}
    ][:3]

    official_name = (
        confirmed.get("official_name")
        or report_input.get("company_name")
        or report_section_value(sections, "二、公司介绍", "公司名称")
    )
    official_domain = confirmed.get("official_domain") or apify.get("input_domain") or report_section_value(
        sections,
        "二、公司介绍",
        "官网",
    )
    domain_attempts = clean_text(apify.get("domain_attempts"))
    domain_basis = f"确认主体：{official_name}；官网：{official_domain}"
    if domain_attempts != MISSING:
        domain_basis += f"；备用 domain 尝试：{domain_attempts}"

    entry_product = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "最适合切入的产品")
    if entry_product == MISSING or "未找到我司信息" in entry_product:
        entry_product = report_section_value(sections, "四、产品线与销售能力", "可切入的产品机会")
    selling_point = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "最适合使用的卖点")
    email_angle = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "首封开发信角度")
    quote_strategy = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "报价策略")
    sample_strategy = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "样品策略")
    certification = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "认证或资料准备建议")
    product_bundle = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "适合推荐的产品组合")
    follow_up = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "后续跟进节奏")
    questions = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "最应该问客户的3到5个问题")
    english_email = report_section_value(sections, "九、开发话术建议", "英文开发信")
    development_priority = report_section_value(sections, "八、对【我司】的合作价值与切入建议", "推荐开发优先级", "待确认")

    risks = []
    for item in as_list(payload.get("risks"))[:3]:
        item_dict = item if isinstance(item, dict) else {"risk": item}
        risk_text = clean_text(item_dict.get("risk"))
        reason_text = clean_text(item_dict.get("reason"))
        if risk_text != MISSING:
            risks.append(f"{risk_text}：{reason_text}")
    risk_summary = "；".join(risks) if risks else report_section_value(sections, "七、风险与注意事项", "信息不透明风险")

    contact_recommendation = format_contact_for_action(primary_contact) if primary_contact else "未返回联系人；先走官网表单、公司邮箱、LinkedIn 公司页或电话总机。"
    contact_priority = "高" if priority_contacts else "中"
    backup_recommendation = "；".join(backup_contacts) if backup_contacts else "未返回明确备选联系人；用公司公开邮箱/官网表单做备份触达。"

    return [
        ["1. 线索确认", "确认主体和官网", domain_basis, "业务员", "高", "来自 web_search/web_fetch、确认官网和 Actor 输入记录"],
        ["2. 首轮触达", "优先联系人", contact_recommendation, "业务员", contact_priority, f"来自 {SHEET_CONTACTS} 的 BD 评分和建议原因"],
        ["2. 首轮触达", "备用联系路径", backup_recommendation, "业务员", "中", "避免单一联系人无回复导致线索中断"],
        ["3. 产品切入", "推荐切入产品", entry_product, "业务员/产品", development_priority, "来自报告第八部分和产品线分析"],
        ["3. 产品切入", "首封卖点和角度", f"{selling_point}；{email_angle}", "业务员", development_priority, "来自报告第八、九部分"],
        ["4. 报价与样品", "报价策略", quote_strategy, "业务员", "中", "根据客户定位、采购可能性和我司信息谨慎判断"],
        ["4. 报价与样品", "样品和资料", f"{sample_strategy}；{certification}；产品组合：{product_bundle}", "业务员/产品", "中", "先降低客户回复门槛，再推进规格和认证确认"],
        ["5. 跟进节奏", "后续跟进", follow_up, "业务员", "中", "开发信发送后需要节奏化跟进"],
        ["6. 待问问题", "首轮必须确认的问题", questions, "客户", "高", "用于判断真实需求、价格带、认证、采购量和现有供应痛点"],
        ["7. 风险控制", "优先核验风险", risk_summary, "业务员", "高", "来自背调报告风险部分和公开信息缺口"],
        ["8. 邮件草稿", "英文开发信", english_email, "业务员", "中", "可直接作为首封邮件草稿，发送前按我司产品再微调"],
    ]


def build_action_plan_rows(payload: JsonDict) -> list[Row]:
    """Return explicit action-plan rows or a generated fallback plan.

    Args:
        payload: Full normalized report payload.

    Returns:
        Rows ready to be written under `ACTION_PLAN_HEADERS`.

    Raises:
        No intentional exceptions.
    """

    return explicit_action_plan_rows(payload) or build_fallback_action_plan_rows(payload)


def write_table(
    sheet: Worksheet,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
    highlight_priority: bool = False,
) -> int:
    """Write a styled table and optionally highlight priority contacts.

    Args:
        sheet: Target worksheet.
        start_row: Row where the table header should be written.
        headers: Header labels.
        rows: Data rows.
        highlight_priority: When true, rows whose first cell is `优先联系`
            receive a stronger fill so business users can scan them quickly.

    Returns:
        The next empty row after the table.

    Raises:
        openpyxl exceptions may propagate if cell writes fail.
    """

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_offset, row_values in enumerate(rows, start=1):
        row_number = start_row + row_offset
        is_priority = highlight_priority and clean_text(row_values[0]) == "优先联系"
        for column, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=clean_text(value))
            cell.font = PRIORITY_FONT if is_priority else DEFAULT_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_priority:
                cell.fill = PRIORITY_BADGE_FILL if column == 1 else PRIORITY_FILL
    last_column = get_excel_column_name(len(headers))
    sheet.auto_filter.ref = f"A{start_row}:{last_column}{sheet.max_row}"
    return sheet.max_row + 1


def make_apify_sheet(workbook: Workbook, payload: JsonDict) -> None:
    """Create the contact-information sheet from Apify output.

    Args:
        workbook: Workbook being built.
        payload: Normalized report payload.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if worksheet operations fail.
    """

    sheet = workbook.create_sheet(SHEET_APIFY)
    apify = get_dict(payload, "apify")
    write_title(sheet, SHEET_CONTACTS, "Actor 状态、固定公司字段和 B2B 联系建议", end_column=7)
    next_row = write_wide_key_value_block(
        sheet,
        4,
        [
            ("Actor页面", apify.get("actor_url")),
            ("输入Domain", apify.get("input_domain")),
            ("备用Domain尝试", apify.get("domain_attempts")),
            ("运行状态", apify.get("status")),
            ("运行ID", apify.get("run_id")),
            ("Dataset", apify.get("dataset_id")),
            ("失败/重试说明", apify.get("failure_reason") or apify.get("retry_note")),
        ],
    )

    raw_items = as_list(apify.get("raw_items"))
    normalized_items = [
        item if isinstance(item, dict) else flatten_dict(item)
        for item in raw_items
        if item is not None
    ]
    if normalized_items:
        next_row = write_section_header(sheet, next_row + 1, "公司固定字段（从联系人结果去重提取）", end_column=7)
        company_pairs = extract_apify_company_pairs(normalized_items)
        if company_pairs:
            next_row = write_wide_key_value_block(sheet, next_row, company_pairs)
        else:
            next_row = write_wide_key_value_block(sheet, next_row, [("公司固定字段", "Apify 未返回可用公司字段")])

        contact_records = build_apify_contact_records(normalized_items)
        recommended = [row for row in contact_records if row["bd_priority"] == "优先联系"][:8]
        if not recommended:
            recommended = contact_records[:5]

        next_row = write_section_header(
            sheet,
            next_row + 1,
            "B2B Business Development 联系建议",
            end_column=7,
        )
        recommendation_rows = [
            [
                row["bd_priority"],
                row["bd_score"],
                row["full_name"],
                row["job_title"],
                row["email"],
                row["linkedin"],
                row["bd_reason"],
            ]
            for row in recommended
        ]
        if not recommendation_rows:
            recommendation_rows = [["未返回", "未返回", "未返回", "未返回", "未返回", "未返回", "没有可排序联系人"]]
        next_row = write_table(
            sheet,
            next_row,
            ["优先级", "评分", "姓名", "职位", "工作邮箱", "LinkedIn", "建议原因"],
            recommendation_rows,
            highlight_priority=True,
        )

        next_row = write_section_header(
            sheet,
            next_row + 1,
            "联系人明细（已移除重复公司字段）",
            end_column=len(APIFY_CONTACT_FIELDS),
        )
        contact_headers = [label for _, label in APIFY_CONTACT_FIELDS]
        contact_rows = [
            [row.get(field) for field, _ in APIFY_CONTACT_FIELDS]
            for row in contact_records
        ]
        next_row = write_table(
            sheet,
            next_row,
            contact_headers,
            contact_rows,
            highlight_priority=True,
        )
        last_column = get_excel_column_name(len(APIFY_CONTACT_FIELDS))
        sheet.auto_filter.ref = f"A{next_row - len(contact_rows) - 1}:{last_column}{next_row - 1}"
        # 不在深处冻结窗格。冻结到第 40 行左右会让 Excel 视图像被锁住，
        # 用户滚轮很难正常浏览上方固定字段和下方联系人列表。
        sheet.freeze_panes = None
    else:
        write_key_value_block(
            sheet,
            next_row + 1,
            [("返回数据", "Apify 未返回可用结构化数据")],
        )
    set_column_widths(
        sheet,
        {
            "A": 14,
            "B": 8,
            "C": 24,
            "D": 36,
            "E": 34,
            "F": 52,
            "G": 56,
            "H": 14,
            "I": 18,
            "J": 18,
            "K": 18,
            "L": 20,
        },
    )
    apply_status_styles(sheet)


def make_action_plan_sheet(workbook: Workbook, payload: JsonDict) -> None:
    """Create the business-development action plan sheet.

    Args:
        workbook: Workbook being built.
        payload: Normalized report payload.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if worksheet operations fail.
    """

    sheet = workbook.create_sheet(SHEET_ACTION_PLAN)
    write_title(sheet, "开发行动计划", "把背调结论落到下一步开发动作", end_column=len(ACTION_PLAN_HEADERS))
    rows = build_action_plan_rows(payload)
    table_start = 4
    next_row = write_table(sheet, table_start, ACTION_PLAN_HEADERS, rows, highlight_priority=False)
    last_column = get_excel_column_name(len(ACTION_PLAN_HEADERS))
    sheet.auto_filter.ref = f"A{table_start}:{last_column}{next_row - 1}"
    sheet.freeze_panes = "A4"

    for row_idx in range(table_start + 1, next_row):
        values = [clean_text(sheet.cell(row=row_idx, column=column).value) for column in range(1, len(ACTION_PLAN_HEADERS) + 1)]
        max_length = max(len(value) for value in values) if values else 0
        if max_length > 260:
            sheet.row_dimensions[row_idx].height = 94
        elif max_length > 150:
            sheet.row_dimensions[row_idx].height = 70
        elif max_length > 80:
            sheet.row_dimensions[row_idx].height = 46

    set_column_widths(
        sheet,
        {
            "A": 15,
            "B": 22,
            "C": 70,
            "D": 18,
            "E": 12,
            "F": 54,
        },
    )
    apply_status_styles(sheet)


def make_risk_sheet(workbook: Workbook, payload: JsonDict) -> None:
    """Create the risks and open questions sheet.

    Args:
        workbook: Workbook being built.
        payload: Normalized report payload.

    Returns:
        None.

    Raises:
        openpyxl exceptions may propagate if worksheet operations fail.
    """

    sheet = workbook.create_sheet(SHEET_RISKS)
    write_title(sheet, "风险待确认", "冲突、缺失、异常和需要人工确认的问题")
    rows = [["风险/待确认项", "原因", "来源URL"]]
    for item in as_list(payload.get("risks")):
        item_dict = item if isinstance(item, dict) else {"risk": item}
        rows.append(
            [
                item_dict.get("risk"),
                item_dict.get("reason"),
                item_dict.get("source_url"),
            ]
        )
    if len(rows) == 1:
        rows.append(["暂无明显风险", "当前公开来源没有返回明显冲突；仍建议人工复核关键交易信息。", MISSING])
    table_start = sheet.max_row + 1
    append_rows(sheet, rows, header=True)
    set_column_widths(sheet, {"A": 42, "B": 80, "C": 60})
    sheet.auto_filter.ref = f"A{table_start}:C{sheet.max_row}"
    sheet.freeze_panes = "A4"
    apply_status_styles(sheet)


def get_excel_column_name(index: int) -> str:
    """Convert a 1-based column index into an Excel column name.

    Args:
        index: One-based column index, such as 1 for A or 28 for AB.

    Returns:
        Excel column letters.

    Raises:
        ValueError: If `index` is less than 1.
    """

    if index < 1:
        raise ValueError("Column index must be >= 1.")
    name = ""
    current = index
    while current:
        current, remainder = divmod(current - 1, 26)
        name = chr(65 + remainder) + name
    return name


def finalize_common_layout(workbook: Workbook) -> None:
    """Apply common layout settings after all sheets are created.

    Args:
        workbook: Workbook to finalize.

    Returns:
        None.

    Raises:
        No intentional exceptions.
    """

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        # Individual writer functions already style populated cells. Avoid
        # reassigning existing openpyxl style proxies here because those proxy
        # objects are not hashable in some openpyxl versions.
        for row_idx in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row_idx].height = max(sheet.row_dimensions[row_idx].height or 18, 18)


def build_workbook(payload: JsonDict, output: Path) -> None:
    """Build and save the XLSX workbook before package sanitization.

    Args:
        payload: Normalized report JSON.
        output: Destination workbook path.

    Returns:
        None.

    Raises:
        OSError: If the output directory cannot be created.
        openpyxl exceptions may propagate if saving fails.
    """

    logging.info("Rendering workbook to %s", output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    make_report_sheet(workbook, payload)
    make_apify_sheet(workbook, payload)
    make_action_plan_sheet(workbook, payload)
    finalize_common_layout(workbook)
    workbook.save(output)


def find_libreoffice() -> str:
    """Find a LibreOffice executable for the required re-save step.

    Args:
        None.

    Returns:
        Path or command name for the LibreOffice/soffice executable.

    Raises:
        RuntimeError: If LibreOffice cannot be found.
    """

    candidates = [
        "soffice",
        "libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return candidate
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
    raise RuntimeError("LibreOffice/soffice not found; XLSX safety flow cannot finish.")


def libreoffice_resave(path: Path) -> None:
    """Re-save the workbook through LibreOffice headless.

    Args:
        path: Workbook path to re-save in place.

    Returns:
        None.

    Raises:
        RuntimeError: If LibreOffice conversion fails or does not produce an
            XLSX file.
    """

    logging.info("Running LibreOffice headless re-save.")
    soffice = find_libreoffice()
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        input_dir = tmp_path / "input"
        output_dir = tmp_path / "output"
        profile_dir = tmp_path / "profile"
        input_dir.mkdir()
        output_dir.mkdir()
        profile_dir.mkdir()
        input_copy = input_dir / path.name
        shutil.copy2(path, input_copy)
        result = subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(output_dir),
                str(input_copy),
            ],
            text=True,
            capture_output=True,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice re-save failed: {result.stderr or result.stdout}")
        candidates = list(output_dir.glob("*.xlsx"))
        if not candidates:
            raise RuntimeError("LibreOffice did not produce an .xlsx file.")
        shutil.copy2(candidates[0], path)


def sanitize_xml_text(xml_text: str) -> str:
    """Remove table/drawing XML references that can trigger Excel repair prompts.

    Args:
        xml_text: XML file content decoded as UTF-8.

    Returns:
        Sanitized XML text.

    Raises:
        No intentional exceptions.
    """

    xml_text = re.sub(r"<tableParts\b[^>]*/>", "", xml_text)
    xml_text = re.sub(r"<drawing\b[^>]*/>", "", xml_text)
    xml_text = re.sub(
        r"<Relationship\b[^>]*(?:/tables/|/drawings/)[^>]*/>",
        "",
        xml_text,
        flags=re.IGNORECASE,
    )
    xml_text = re.sub(
        r"<Override\b[^>]*(?:/tables/|/drawings/)[^>]*/>",
        "",
        xml_text,
        flags=re.IGNORECASE,
    )
    return xml_text


def sanitize_xlsx_package(path: Path) -> None:
    """Remove table/drawing package residue from an XLSX zip archive.

    Args:
        path: Workbook path to sanitize in place.

    Returns:
        None.

    Raises:
        zipfile.BadZipFile: If the workbook is not a valid ZIP package.
        OSError: If the sanitized workbook cannot be written.
    """

    logging.info("Sanitizing XLSX package.")
    temp_path = path.with_suffix(".sanitized.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            lowered = item.filename.lower()
            if lowered.startswith("xl/tables/") or lowered.startswith("xl/drawings/"):
                continue
            data = source.read(item.filename)
            if item.filename.endswith(".xml") or item.filename.endswith(".rels"):
                try:
                    data = sanitize_xml_text(data.decode("utf-8")).encode("utf-8")
                except UnicodeDecodeError:
                    pass
            target.writestr(item, data)
    temp_path.replace(path)


def run_unzip_test(path: Path) -> None:
    """Run `unzip -t` against the workbook package.

    Args:
        path: Workbook path to validate.

    Returns:
        None.

    Raises:
        RuntimeError: If unzip reports a corrupt package.
    """

    logging.info("Running unzip -t validation.")
    result = subprocess.run(["unzip", "-t", str(path)], text=True, capture_output=True, check=False)
    if result.returncode != 0:
        raise RuntimeError(f"unzip -t failed: {result.stderr or result.stdout}")


def validate_xlsx(path: Path) -> None:
    """Validate that the workbook opens and has no table/drawing residue.

    Args:
        path: Workbook path to inspect.

    Returns:
        None.

    Raises:
        RuntimeError: If forbidden package parts remain or openpyxl cannot load
            the workbook.
        zipfile.BadZipFile: If the XLSX package is invalid.
    """

    logging.info("Running openpyxl and package residue validation.")
    with zipfile.ZipFile(path, "r") as archive:
        forbidden: list[str] = []
        for name in archive.namelist():
            if name.startswith("xl/tables/") or name.startswith("xl/drawings/"):
                forbidden.append(name)
                continue
            if name.endswith((".xml", ".rels")):
                xml_text = archive.read(name).decode("utf-8", errors="ignore")
                if "tableParts" in xml_text or "/drawings/" in xml_text or "/tables/" in xml_text:
                    forbidden.append(name)
    if forbidden:
        raise RuntimeError(f"XLSX still contains table/drawing residue: {forbidden[:5]}")
    workbook = load_workbook(path, data_only=False)
    required = {SHEET_REPORT, SHEET_CONTACTS, SHEET_ACTION_PLAN}
    missing_sheets = sorted(required - set(workbook.sheetnames))
    if missing_sheets:
        raise RuntimeError(f"Workbook missing required sheets: {missing_sheets}")
    unexpected_sheets = sorted(set(workbook.sheetnames) - required)
    if unexpected_sheets:
        raise RuntimeError(f"Workbook contains unexpected sheets: {unexpected_sheets}")


def run_safety_flow(path: Path) -> None:
    """Run the full Mac Excel safety flow required before delivery.

    Args:
        path: Workbook path to process.

    Returns:
        None.

    Raises:
        RuntimeError: If any required safety step fails.
    """

    libreoffice_resave(path)
    sanitize_xlsx_package(path)
    run_unzip_test(path)
    validate_xlsx(path)
    logging.info("XLSX safety flow finished successfully.")


def main(argv: list[str]) -> int:
    """Program entry point.

    Args:
        argv: Command-line arguments without the executable name.

    Returns:
        Process exit code. Zero means success; non-zero means generation failed.

    Raises:
        No exceptions are intentionally allowed to escape; failures are logged
        and printed for the calling agent.
    """

    args = parse_args(argv)
    output = Path(args.output)
    configure_logging(output, args.log)
    try:
        require_openpyxl()
        payload = load_payload(Path(args.input))
        build_workbook(payload, output)
        run_safety_flow(output)
    except Exception as exc:  # noqa: BLE001 - CLI should report any failure clearly.
        logging.exception("Failed to build company background-check workbook.")
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(str(output))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

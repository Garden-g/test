#!/usr/bin/env python3
"""Build the boss-facing Alibaba shop operation report as a safe XLSX workbook.

This script is intentionally conservative: it uses only ordinary worksheet
cells, fills, freeze panes, and filters. It does not create Excel Tables,
charts, drawings, images, or shapes, because those package parts are the most
common source of Mac Excel repair prompts in this project.
"""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


JsonDict = dict[str, Any]

EXPECTED_SHEETS = [
    "本周结论",
    "经营问题诊断",
    "行动与复查",
]

ALLOWED_ROLES = {"", "运营", "业务"}
ALLOWED_PRIORITIES = {"先做", "随后做", "持续观察"}
FORBIDDEN_DIAGNOSIS_TEXT = (
    "未返回",
    "待确认",
    "不可判断",
    "待补充",
    "TODO",
    "加强运营",
    "持续优化",
)

FORBIDDEN_BUSINESS_TEXT = (
    "m" + "cp",
    "br" + "idge",
    "accio-" + "mcp-cli",
    "errorCode",
    "errorMsg",
    "-32002",
    "Agent 类型不允许",
    "Traceback",
    "ECONNREFUSED",
)

TITLE_FILL = "1F4E78"
SECTION_FILL = "DDEBF7"
HEADER_FILL = "5B9BD5"
WARNING_FILL = "FFF2CC"
MISSING_FILL = "F2F2F2"
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def load_json(path: Path) -> JsonDict:
    """Read a JSON file and return a dictionary.

    Args:
        path: JSON file path.

    Returns:
        Parsed dictionary. Non-dict roots are wrapped into ``{"value": root}``
        so downstream rendering can still degrade gracefully.

    Raises:
        FileNotFoundError: If the file does not exist.
        json.JSONDecodeError: If the file is not valid JSON.
    """

    value = json.loads(path.read_text(encoding="utf-8"))
    return value if isinstance(value, dict) else {"value": value}


def pick(data: Any, *keys: str, default: Any = "未返回") -> Any:
    """Return the first present key from a dictionary.

    Args:
        data: Candidate dictionary.
        *keys: Keys to try in order.
        default: Value used when none of the keys are present.

    Returns:
        The first non-empty value, or ``default``.

    Raises:
        None: The helper deliberately degrades instead of failing report builds.
    """

    if not isinstance(data, dict):
        return default
    for key in keys:
        value = data.get(key)
        if value not in (None, ""):
            return value
    return default


def as_list(value: Any) -> list[Any]:
    """Normalize common list wrappers into a plain list.

    Args:
        value: A list, a wrapper dict containing ``data/list/rows/items``, or
            any other object.

    Returns:
        A list suitable for table rendering.

    Raises:
        None.
    """

    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        for key in ("data", "list", "rows", "items", "values"):
            child = value.get(key)
            if isinstance(child, list):
                return child
    return []


def percent_text(value: Any) -> str:
    """Render a decimal ratio as a readable percentage when possible.

    Args:
        value: Decimal ratio, percentage text, or a missing value.

    Returns:
        Percentage text for numeric ratios; otherwise the normal safe text.

    Raises:
        None. Invalid numeric strings are displayed without conversion.
    """

    if value in (None, ""):
        return "未返回"
    if isinstance(value, str) and value.strip().endswith("%"):
        return value.strip()
    try:
        return f"{float(value):.2%}"
    except (TypeError, ValueError):
        return text(value)


def data_quality_text(value: Any) -> str:
    """Summarize structured data-quality output without leaking raw field names.

    Args:
        value: Analysis or normalized data-quality object.

    Returns:
        A short Chinese description suitable for the boss page.

    Raises:
        None.
    """

    if not isinstance(value, dict):
        return text(value)
    status_map = {
        "ok": "可判断",
        "partial": "部分可判断",
        "degraded": "降级判断",
        "blocked": "不可判断",
    }
    status = status_map.get(str(value.get("status", "")).lower(), text(value.get("status")))
    coverage = value.get("coverage_rate")
    coverage_text = percent_text(coverage) if coverage not in (None, "") else "未返回"
    collection = value.get("collection") if isinstance(value.get("collection"), dict) else {}
    calls = collection.get("calls") if isinstance(collection, dict) else None
    call_text = f"，记录调用 {calls} 次" if isinstance(calls, int) and calls > 0 else ""
    return f"{status}；模块覆盖 {coverage_text}{call_text}"


def star_level_text(value: Any) -> str:
    """Render an ability level with an explicit ``星`` unit.

    Args:
        value: Raw level such as ``"4.0000"``.

    Returns:
        A concise business-facing value such as ``"4星"``.

    Raises:
        No exceptions are intentionally raised.
    """

    if value in (None, "", "未返回"):
        return "未返回"
    try:
        numeric = float(value)
        rendered = str(int(numeric)) if numeric.is_integer() else f"{numeric:g}"
        return f"{rendered}星"
    except (TypeError, ValueError):
        return text(value)


def text(value: Any) -> str:
    """Convert a workbook value to safe boss-facing text.

    Args:
        value: Any value from normalized data or analysis.

    Returns:
        A display string. Missing values are shown as ``未返回``.

    Raises:
        None.
    """

    if value in (None, ""):
        return "未返回"
    if isinstance(value, dict):
        parts = [f"{key}：{text(item)}" for key, item in list(value.items())[:8]]
        rendered = "；".join(parts) or "未返回"
    elif isinstance(value, list):
        rendered = "；".join(text(item) for item in value[:8]) or "未返回"
    else:
        rendered = str(value)

    # Excel treats leading =, +, -, and @ as formulas. Prefixing an apostrophe
    # preserves visible text while preventing external data from being executed.
    return f"'{rendered}" if rendered.startswith(("=", "+", "-", "@")) else rendered


def protect_formula_cells(workbook: Workbook) -> None:
    """Convert formula-like strings in every worksheet to literal text.

    Args:
        workbook: Workbook that is about to be saved.

    Returns:
        None.

    Raises:
        None. Only string cells are inspected and changed.
    """

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@")):
                    cell.value = f"'{cell.value}"


def compact(value: Any, limit: int = 160) -> str:
    """Shorten long strings so rows remain readable in Excel.

    Args:
        value: Raw value.
        limit: Maximum number of characters.

    Returns:
        Compact display text.

    Raises:
        None.
    """

    raw = text(value).replace("\n", " ").strip()
    return raw if len(raw) <= limit else raw[: limit - 1] + "..."


def sheet_title(ws: Worksheet, title: str, subtitle: str = "") -> int:
    """Write a merged title band and return the next row index.

    Args:
        ws: Target worksheet.
        title: Sheet title.
        subtitle: Optional context text.

    Returns:
        The next writable row number.

    Raises:
        None.
    """

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell = ws.cell(1, 1, f"{title}  |  {subtitle}" if subtitle else title)
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 26
    return 3


def write_kv_rows(ws: Worksheet, row: int, rows: list[tuple[str, Any]]) -> int:
    """Write label/value rows.

    Args:
        ws: Target worksheet.
        row: Starting row.
        rows: ``(label, value)`` pairs.

    Returns:
        Next writable row number.

    Raises:
        None.
    """

    for label, value in rows:
        ws.cell(row, 1, label)
        ws.cell(row, 2, compact(value, 300))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
        ws.cell(row, 1).font = Font(name="Arial", size=10, bold=True, color="1F4E78")
        for col in range(1, 9):
            ws.cell(row, col).border = THIN_BORDER
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    return row + 1


def write_table(ws: Worksheet, row: int, title: str, headers: list[str], rows: list[list[Any]]) -> int:
    """Write a section title and a normal-cell table.

    Args:
        ws: Target worksheet.
        row: Starting row.
        title: Section title.
        headers: Column headers.
        rows: Table body rows.

    Returns:
        Next writable row number.

    Raises:
        None.
    """

    width = max(len(headers), 1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    ws.cell(row, 1, title)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    ws.cell(row, 1).font = Font(name="Arial", size=11, bold=True, color="1F4E78")
    row += 1

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    header_row = row
    row += 1

    body = rows or [["未返回"] + [""] * (len(headers) - 1)]
    for values in body:
        padded = list(values)[: len(headers)] + [""] * max(0, len(headers) - len(values))
        for col, value in enumerate(padded, start=1):
            cell = ws.cell(row, col, compact(value))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value in {"未返回", "不可判断", "仅覆盖已获取范围", "需用户确认"}:
                cell.fill = PatternFill("solid", fgColor=MISSING_FILL)
                cell.font = Font(name="Arial", size=10, italic=True, color="808080")
            elif str(cell.value).startswith("P0"):
                cell.fill = PatternFill("solid", fgColor="F4CCCC")
            elif str(cell.value).startswith("P1"):
                cell.fill = PatternFill("solid", fgColor=WARNING_FILL)
        row += 1

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(width)}{row - 1}"
    ws.freeze_panes = f"A{header_row + 1}"
    return row + 2


def finish_sheet(ws: Worksheet) -> None:
    """Apply widths and base styles to a worksheet.

    Args:
        ws: Worksheet to polish.

    Returns:
        None.

    Raises:
        None.
    """

    widths = [18, 24, 24, 24, 24, 24, 24, 28]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            updated_font = copy(cell.font)
            updated_font.name = "Arial"
            updated_font.sz = cell.font.sz or 10
            cell.font = updated_font
            updated_alignment = copy(cell.alignment)
            updated_alignment.wrap_text = True
            updated_alignment.vertical = cell.alignment.vertical or "top"
            cell.alignment = updated_alignment


def diagnosis_text(value: Any) -> str:
    """Convert diagnosed content to safe display text without placeholders.

    Args:
        value: Scalar or short list supplied by the diagnosis JSON.

    Returns:
        Formula-safe text. Missing optional values become an empty string.

    Raises:
        ValueError: A dictionary reaches the user-facing rendering layer.
    """

    if value in (None, ""):
        return ""
    if isinstance(value, dict):
        raise ValueError("Nested objects cannot be rendered into compact workbook cells.")
    if isinstance(value, list):
        rendered = "、".join(
            str(item).strip()
            for item in value[:3]
            if item not in (None, "") and str(item).strip()
        )
    else:
        rendered = str(value).strip()
    if rendered.startswith(("=", "+", "-", "@")):
        return f"'{rendered}"
    return rendered


def validate_management_diagnosis(payload: JsonDict) -> None:
    """Validate that LLM output is specific enough for boss-facing delivery.

    Args:
        payload: Canonical operations-expert diagnosis JSON.

    Returns:
        None. A valid payload completes without returning a value.

    Raises:
        ValueError: Required diagnosis, action, ownership, or wording rules are
            violated. The caller must stop XLSX delivery on this error.
    """

    if not isinstance(payload, dict):
        raise ValueError("Diagnosis root must be an object.")
    conclusion = str(payload.get("executive_conclusion") or "").strip()
    top = payload.get("top_diagnoses") or []
    details = payload.get("detail_diagnoses") or []
    actions = payload.get("actions") or []
    limitations = payload.get("data_limitations") or []

    if not conclusion:
        raise ValueError("executive_conclusion is required.")
    if not isinstance(top, list) or not 1 <= len(top) <= 3:
        raise ValueError("top_diagnoses must contain 1 to 3 rows.")
    if not isinstance(details, list) or not 1 <= len(details) <= 15:
        raise ValueError("detail_diagnoses must contain 1 to 15 rows.")
    if not isinstance(actions, list) or not 1 <= len(actions) <= 8:
        raise ValueError("actions must contain 1 to 8 rows.")
    if not isinstance(limitations, list) or len(limitations) > 3:
        raise ValueError("data_limitations must contain no more than 3 notes.")

    for index, row in enumerate(top):
        if not isinstance(row, dict):
            raise ValueError(f"top_diagnoses[{index}] must be an object.")
        for key in ("what_happened", "root_cause", "solution", "review_standard"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"top_diagnoses[{index}].{key} is required.")
        if str(row.get("role") or "") not in ALLOWED_ROLES:
            raise ValueError(f"top_diagnoses[{index}].role is invalid.")

    diagnosis_pairs: set[tuple[str, str]] = set()
    for index, row in enumerate(details):
        if not isinstance(row, dict):
            raise ValueError(f"detail_diagnoses[{index}] must be an object.")
        for key in ("issue_group", "evidence", "expert_diagnosis", "root_cause", "solution"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"detail_diagnoses[{index}].{key} is required.")
        objects = row.get("objects") or []
        if not isinstance(objects, list) or len(objects) > 3:
            raise ValueError(
                f"detail_diagnoses[{index}].objects must contain at most 3 examples."
            )
        pair = (
            str(row.get("root_cause") or "").strip(),
            str(row.get("solution") or "").strip(),
        )
        if pair in diagnosis_pairs:
            raise ValueError("Duplicate root-cause and solution rows must be grouped.")
        diagnosis_pairs.add(pair)

    action_texts: set[str] = set()
    for index, row in enumerate(actions):
        if not isinstance(row, dict):
            raise ValueError(f"actions[{index}] must be an object.")
        for key in ("priority", "action", "due", "review_standard", "source_issue_group"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"actions[{index}].{key} is required.")
        if str(row.get("priority") or "") not in ALLOWED_PRIORITIES:
            raise ValueError(f"actions[{index}].priority is invalid.")
        if str(row.get("role") or "") not in ALLOWED_ROLES:
            raise ValueError(f"actions[{index}].role is invalid.")
        action = str(row.get("action") or "").strip()
        if action in action_texts:
            raise ValueError("Duplicate action rows must be grouped.")
        action_texts.add(action)

    visible_text = json.dumps(payload, ensure_ascii=False)
    matched = [
        token
        for token in FORBIDDEN_DIAGNOSIS_TEXT
        if token.lower() in visible_text.lower()
    ]
    if matched:
        raise ValueError(f"Placeholder or generic wording is forbidden: {matched}")
    internal_terms = [
        term
        for term in FORBIDDEN_BUSINESS_TEXT
        if term.lower() in visible_text.lower()
    ]
    if internal_terms:
        raise ValueError(f"Internal implementation wording is forbidden: {internal_terms}")


def owner_text(item: JsonDict) -> str:
    """Return a verified owner name or the allowed fallback business role.

    Args:
        item: One top diagnosis or action row.

    Returns:
        Verified ``owner_name`` when present; otherwise ``运营``/``业务`` or
        an empty string when the role cannot be determined.

    Raises:
        None. Role validity is enforced before rendering.
    """

    owner_name = str(item.get("owner_name") or "").strip()
    return owner_name or str(item.get("role") or "").strip()


def write_compact_title(
    sheet: Worksheet,
    title: str,
    period: str,
    conclusion: str | None = None,
    column_count: int = 6,
) -> None:
    """Write the compact report title, period, and optional conclusion.

    Args:
        sheet: Target worksheet.
        title: Natural-language sheet title.
        period: Business period shown once near the top.
        conclusion: Optional executive conclusion for the first sheet.
        column_count: Number of columns spanned by the title band.

    Returns:
        None.

    Raises:
        ValueError: Unsafe nested data reaches ``diagnosis_text``.
    """

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    sheet.cell(1, 1, diagnosis_text(title))
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=TITLE_FILL)
    sheet.cell(1, 1).font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    sheet.cell(1, 1).alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 28

    sheet.cell(2, 1, "数据周期")
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=column_count)
    sheet.cell(2, 2, diagnosis_text(period))
    if conclusion is not None:
        sheet.cell(3, 1, "一句话结论")
        sheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=column_count)
        sheet.cell(3, 2, diagnosis_text(conclusion))
        sheet.row_dimensions[3].height = 42
    for row_number in (2, 3):
        for column_number in range(1, column_count + 1):
            cell = sheet.cell(row_number, column_number)
            if column_number == 1:
                cell.fill = PatternFill("solid", fgColor=SECTION_FILL)
                cell.font = Font(name="Arial", size=10, bold=True, color="1F4E78")
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_compact_table(
    sheet: Worksheet,
    header_row: int,
    headers: list[str],
    rows: list[list[str]],
) -> int:
    """Write one compact normal-cell table.

    Args:
        sheet: Target worksheet.
        header_row: One-based header row.
        headers: Natural-language column labels.
        rows: Already diagnosed table body.

    Returns:
        The first row after the table.

    Raises:
        ValueError: Row width does not match the declared headers.
    """

    for column_number, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_number, header)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    sheet.row_dimensions[header_row].height = 28

    for row_offset, values in enumerate(rows, start=1):
        if len(values) != len(headers):
            raise ValueError("Compact table row width does not match its headers.")
        row_number = header_row + row_offset
        for column_number, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column_number, diagnosis_text(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
        sheet.row_dimensions[row_number].height = 54

    end_row = header_row + len(rows)
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, end_row)}"
    )
    sheet.freeze_panes = f"A{header_row + 1}"
    return end_row + 1


def style_compact_sheet(sheet: Worksheet, widths: list[int]) -> None:
    """Apply restrained widths and typography to a compact worksheet.

    Args:
        sheet: Worksheet to style.
        widths: Column widths in display order.

    Returns:
        None.

    Raises:
        None.
    """

    sheet.sheet_view.showGridLines = False
    for column_number, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_number)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            updated_font = copy(cell.font)
            updated_font.name = "Arial"
            updated_font.sz = cell.font.sz or 10
            cell.font = updated_font
            updated_alignment = copy(cell.alignment)
            updated_alignment.wrap_text = True
            updated_alignment.vertical = cell.alignment.vertical or "top"
            cell.alignment = updated_alignment


def build_compact_workbook(
    diagnosis: JsonDict,
    context: JsonDict | None = None,
) -> Workbook:
    """Build the diagnosis-first three-sheet weekly management workbook.

    Args:
        diagnosis: Canonical operations-expert LLM diagnosis.
        context: Optional runtime display context such as company name.

    Returns:
        An in-memory workbook containing only decision-useful diagnosed content.

    Raises:
        ValueError: Diagnosis quality or table structure is invalid.
    """

    validate_management_diagnosis(diagnosis)
    context = context or {}
    meta = diagnosis.get("report_meta") or {}
    title = str(meta.get("title") or "老板经营周报").strip()
    company = str(context.get("company") or "").strip()
    period = str(meta.get("period") or "").strip()
    display_title = f"{title}｜{company}" if company else title

    workbook = Workbook()
    home = workbook.active
    home.title = EXPECTED_SHEETS[0]
    diagnosis_sheet = workbook.create_sheet(EXPECTED_SHEETS[1])
    action_sheet = workbook.create_sheet(EXPECTED_SHEETS[2])

    top_rows = diagnosis["top_diagnoses"]
    show_home_owner = any(owner_text(item) for item in top_rows)
    home_headers = ["发生了什么", "核心原因", "解决方案"]
    if show_home_owner:
        home_headers.append("负责角色")
    home_headers.append("复查标准")
    rendered_top_rows: list[list[str]] = []
    for item in top_rows:
        row = [
            diagnosis_text(item.get("what_happened")),
            diagnosis_text(item.get("root_cause")),
            diagnosis_text(item.get("solution")),
        ]
        if show_home_owner:
            row.append(owner_text(item))
        row.append(diagnosis_text(item.get("review_standard")))
        rendered_top_rows.append(row)

    write_compact_title(
        home,
        display_title,
        period,
        str(diagnosis.get("executive_conclusion") or ""),
        len(home_headers),
    )
    next_row = write_compact_table(home, 5, home_headers, rendered_top_rows)
    limitations = [
        diagnosis_text(item)
        for item in diagnosis.get("data_limitations") or []
        if diagnosis_text(item)
    ]
    if limitations:
        home.cell(next_row + 1, 1, "数据限制")
        home.merge_cells(
            start_row=next_row + 1,
            start_column=2,
            end_row=next_row + 1,
            end_column=len(home_headers),
        )
        home.cell(next_row + 1, 2, "；".join(limitations))
        home.cell(next_row + 1, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
        home.cell(next_row + 1, 1).font = Font(bold=True, color="1F4E78")
        home.cell(next_row + 1, 2).alignment = Alignment(vertical="top", wrap_text=True)
    home_widths = [31, 30, 31] + ([14] if show_home_owner else []) + [27]
    style_compact_sheet(home, home_widths)

    detail_headers = [
        "问题类型",
        "重点对象",
        "关键证据",
        "专家判断",
        "核心原因",
        "解决方案",
    ]
    rendered_details = [
        [
            diagnosis_text(item.get("issue_group")),
            diagnosis_text(item.get("objects") or []),
            diagnosis_text(item.get("evidence")),
            diagnosis_text(item.get("expert_diagnosis")),
            diagnosis_text(item.get("root_cause")),
            diagnosis_text(item.get("solution")),
        ]
        for item in diagnosis["detail_diagnoses"]
    ]
    write_compact_title(
        diagnosis_sheet,
        EXPECTED_SHEETS[1],
        period,
        column_count=len(detail_headers),
    )
    write_compact_table(diagnosis_sheet, 4, detail_headers, rendered_details)
    style_compact_sheet(diagnosis_sheet, [16, 24, 30, 30, 28, 30])

    actions = diagnosis["actions"]
    show_action_owner = any(owner_text(item) for item in actions)
    action_headers = ["优先级", "行动"]
    if show_action_owner:
        action_headers.append("负责角色")
    action_headers.extend(["完成时点", "复查标准"])
    rendered_actions: list[list[str]] = []
    for item in actions:
        row = [
            diagnosis_text(item.get("priority")),
            diagnosis_text(item.get("action")),
        ]
        if show_action_owner:
            row.append(owner_text(item))
        row.extend(
            [
                diagnosis_text(item.get("due")),
                diagnosis_text(item.get("review_standard")),
            ]
        )
        rendered_actions.append(row)
    write_compact_title(
        action_sheet,
        EXPECTED_SHEETS[2],
        period,
        column_count=len(action_headers),
    )
    write_compact_table(action_sheet, 4, action_headers, rendered_actions)
    action_widths = [13, 38] + ([14] if show_action_owner else []) + [16, 32]
    style_compact_sheet(action_sheet, action_widths)

    return workbook


def build_boss_summary(wb: Workbook, report_data: JsonDict, analysis: JsonDict) -> None:
    """Create the boss conclusion sheet.

    Args:
        wb: Workbook being generated.
        report_data: Normalized business data.
        analysis: Deterministic analysis output.

    Returns:
        None.

    Raises:
        None.
    """

    meta = report_data.get("meta") or {}
    ws = wb.active
    ws.title = "老板结论"
    subtitle = f"{pick(meta, 'company_name')} | {pick(meta, 'period_start')} 至 {pick(meta, 'period_end')}"
    row = sheet_title(ws, "老板结论", subtitle)
    row = write_kv_rows(
        ws,
        row,
        [
            ("报告类型", "老板经营月报" if meta.get("mode") == "monthly" else "老板经营周报"),
            ("公司", pick(meta, "company_name")),
            ("行业", pick(meta, "industry")),
            ("周期", f"{pick(meta, 'period_start')} 至 {pick(meta, 'period_end')}"),
            ("老板结论", pick(analysis, "one_liner", "summary", default="需结合工作簿明细判断。")),
            (
                "数据可信度",
                data_quality_text(
                    pick(
                        analysis,
                        "data_quality",
                        "coverage_status",
                        default=report_data.get("data_quality") or "仅覆盖已返回数据。",
                    )
                ),
            ),
        ],
    )
    actions = as_list(analysis.get("top3_actions") or analysis.get("top_actions"))
    action_rows = []
    for index, item in enumerate(actions[:5], start=1):
        if isinstance(item, dict):
            action_rows.append([
                index,
                pick(item, "priority", default=f"P{min(index, 3)}"),
                pick(item, "action", "title", "name"),
                pick(item, "why", "reason", "evidence"),
                pick(item, "owner", default="待确认"),
            ])
        else:
            action_rows.append([index, f"P{min(index, 3)}", item, "来自分析输出", "待确认"])
    write_table(ws, row, "本期必抓动作", ["序号", "优先级", "动作", "依据", "建议负责人"], action_rows)
    finish_sheet(ws)


def build_overview(wb: Workbook, report_data: JsonDict) -> None:
    """Create the operation overview sheet.

    Args:
        wb: Workbook being generated.
        report_data: Normalized business data.

    Returns:
        None.

    Raises:
        None.
    """

    ws = wb.create_sheet("经营总览")
    row = sheet_title(ws, "经营总览")
    summary = report_data.get("summary") or {}
    indicators = as_list(summary.get("indicators"))
    rows = []
    for item in indicators:
        if isinstance(item, dict):
            rows.append([
                pick(item, "name"),
                pick(item, "value"),
                pick(item, "cycle_crc", "cycleCrc"),
                pick(item, "rival_avg", "rivalAvg"),
                pick(item, "rival_good", "rivalGood"),
                pick(item, "vs_avg", "vsAvg"),
            ])
    write_table(ws, row, "核心指标", ["指标", "本店", "环比", "同行均值", "同行优秀", "判断"], rows)
    finish_sheet(ws)


def build_star_sheet(wb: Workbook, report_data: JsonDict) -> None:
    """Create the star-level diagnosis sheet.

    Args:
        wb: Workbook being generated.
        report_data: Normalized business data.

    Returns:
        None.

    Raises:
        None.
    """

    ws = wb.create_sheet("星级保星诊断")
    diagnosis = report_data.get("diagnosis") or {}
    star_overview = diagnosis.get("star_overview") if isinstance(diagnosis, dict) else {}
    if not isinstance(star_overview, dict):
        star_overview = {}
    row = sheet_title(ws, "星级保星诊断")
    row = write_kv_rows(
        ws,
        row,
        [
            (
                "当前星级",
                pick(
                    star_overview,
                    "star_level",
                    "current_star",
                    default=pick(diagnosis, "current_star", "currentStar"),
                ),
            ),
            (
                "预测星级",
                pick(
                    star_overview,
                    "predicted_star",
                    "predictedStar",
                    default=pick(diagnosis, "predicted_star", "predictedStar"),
                ),
            ),
            ("风险摘要", pick(diagnosis, "conclusion", "risk_summary", "summary", default="未返回")),
        ],
    )
    advice_map: dict[str, str] = {}
    for advice in as_list(diagnosis.get("advices")):
        if not isinstance(advice, dict):
            continue
        indicator = text(pick(advice, "indicator", "name"))
        details = "；".join(text(item) for item in as_list(advice.get("details")))
        if indicator != "未返回" and details:
            advice_map[indicator] = details
    ability_rows = []
    for ability in as_list(diagnosis.get("abilities")):
        if not isinstance(ability, dict):
            continue
        kpis = as_list(ability.get("kpis"))
        gap_text = "；".join(
            f"{pick(kpi, 'name')}：当前 {pick(kpi, 'value')}，进阶参考 {pick(kpi, 'next_level_avg', 'target')}"
            for kpi in kpis
            if isinstance(kpi, dict)
        )
        suggestion = "；".join(
            advice_map.get(text(pick(kpi, "name")), "")
            for kpi in kpis
            if isinstance(kpi, dict) and advice_map.get(text(pick(kpi, "name")))
        )
        ability_rows.append([
            pick(ability, "ability", "name"),
            pick(ability, "score", "value"),
            star_level_text(pick(ability, "star", "status")),
            gap_text or pick(ability, "gap", "next_level_gap"),
            compact(suggestion or pick(ability, "suggestion", "action", default="未返回")),
        ])
    write_table(ws, row, "能力项", ["能力项", "当前值/分数", "能力等级", "差距", "建议动作"], ability_rows)
    finish_sheet(ws)


def build_flow_sheet(wb: Workbook, report_data: JsonDict, analysis: JsonDict) -> None:
    """Create the traffic and funnel sheet.

    Args:
        wb: Workbook being generated.
        report_data: Normalized business data.
        analysis: Deterministic analysis output.

    Returns:
        None.

    Raises:
        None.
    """

    ws = wb.create_sheet("流量结构与漏斗")
    row = sheet_title(ws, "流量结构与漏斗")
    funnel = analysis.get("funnel_diagnosis") or {}
    row = write_kv_rows(
        ws,
        row,
        [
            ("漏斗判断", pick(funnel, "summary", "finding", default="未返回")),
            ("主要漏点", pick(funnel, "bottleneck", "main_issue", default="未返回")),
            ("建议动作", pick(funnel, "action", "suggestion", default="未返回")),
        ],
    )
    channel = report_data.get("channel") or report_data.get("channels") or {}
    channel_rows = []
    for item in as_list(channel):
        if isinstance(item, dict):
            trend = (
                f"访客 {percent_text(pick(item, 'detail_uv_chg', default=None))}；"
                f"TM {percent_text(pick(item, 'tm_uv_chg', default=None))}；"
                f"询盘 {percent_text(pick(item, 'fb_uv_chg', default=None))}"
            )
            channel_rows.append([
                pick(item, "name", "channel"),
                pick(item, "detail_uv", "uv", "visitors"),
                pick(item, "tm_uv", "tm", "chat_visitors"),
                pick(item, "fb_uv", "inquiry", "business"),
                percent_text(pick(item, "uv_ab_rate", "business_rate", default=None)),
                trend,
            ])
    write_table(ws, row, "渠道结构", ["渠道", "访客", "TM人数", "询盘人数", "商机率", "环比变化"], channel_rows)
    finish_sheet(ws)


def build_product_sheet(wb: Workbook, report_data: JsonDict, analysis: JsonDict) -> None:
    """Create the product structure sheet.

    Args:
        wb: Workbook being generated.
        report_data: Normalized business data.
        analysis: Deterministic analysis output.

    Returns:
        None.

    Raises:
        None.
    """

    ws = wb.create_sheet("商品结构与清单")
    row = sheet_title(ws, "商品结构与清单")
    product_rows = []
    quadrants = analysis.get("products_quadrant") or (analysis.get("quadrants") or {}).get("products") or {}
    if isinstance(quadrants, dict):
        for group, items in quadrants.items():
            for item in as_list(items)[:20]:
                if isinstance(item, dict):
                    product_rows.append([
                        group,
                        pick(item, "product_id", "id"),
                        pick(item, "title", "subject", "name"),
                        pick(item, "imps"),
                        pick(item, "fb_num", "inquiry"),
                        pick(item, "why", "evidence"),
                        "; ".join(text(v) for v in as_list(item.get("actions"))[:3]),
                    ])
    if not product_rows:
        for item in as_list(report_data.get("products_top") or report_data.get("products"))[:50]:
            if isinstance(item, dict):
                product_rows.append([
                    "待判断",
                    pick(item, "product_id", "id"),
                    pick(item, "subject", "title", "name"),
                    pick(item, "imps"),
                    pick(item, "fb_num", "inquiry"),
                    pick(item, "fb_rate"),
                    "需结合商品详情复查",
                ])
    write_table(ws, row, "商品分层与处理清单", ["分层", "商品ID", "商品", "曝光", "询盘", "证据", "建议动作"], product_rows)
    finish_sheet(ws)


def build_service_sheet(wb: Workbook, report_data: JsonDict) -> None:
    """Create the service warning sheet.

    Args:
        wb: Workbook being generated.
        report_data: Normalized business data.

    Returns:
        None.

    Raises:
        None.
    """

    ws = wb.create_sheet("服务力预警")
    row = sheet_title(ws, "服务力预警")
    service = report_data.get("service") or report_data.get("communication") or {}
    metric_rows = []
    if isinstance(service, dict):
        service_labels = {
            "first_5min_reply_rate_30d": "30天首次5分钟回复率",
            "avg_reply_time_30d": "30天平均回复时长",
            "reply_over_12h_count": "超过12小时未回复",
            "offline_msg_count": "离线消息",
            "not_follow_count": "未跟进",
            "repeat_reply_count": "重复回复",
            "status": "服务状态",
        }
        for key, value in service.items():
            if isinstance(value, (dict, list)):
                continue
            metric_rows.append([
                service_labels.get(key, "补充服务指标"),
                value,
                "需按同行均值或历史趋势判断",
                "未展开聊天内容",
            ])
        for warning in as_list(service.get("warnings")):
            metric_rows.append(["服务预警", warning, "需要当日复查", "未展开聊天内容"])
    write_table(ws, row, "服务指标摘要", ["指标", "数值", "判断", "备注"], metric_rows)
    finish_sheet(ws)


def build_backlog_sheet(wb: Workbook, analysis: JsonDict) -> None:
    """Create the action backlog sheet.

    Args:
        wb: Workbook being generated.
        analysis: Deterministic analysis output.

    Returns:
        None.

    Raises:
        None.
    """

    ws = wb.create_sheet("行动Backlog")
    row = sheet_title(ws, "行动Backlog")
    backlog_rows = []
    backlog = analysis.get("backlog")
    prioritized_items: list[tuple[str, Any]] = []
    if isinstance(backlog, dict):
        for priority in ("P0", "P1", "P2", "P3"):
            prioritized_items.extend((priority, item) for item in as_list(backlog.get(priority)))
    else:
        prioritized_items.extend(("P2", item) for item in as_list(backlog))
    for priority, item in prioritized_items:
        if isinstance(item, dict):
            backlog_rows.append([
                pick(item, "priority", default=priority),
                pick(item, "topic", "object", "title"),
                pick(item, "evidence", "why"),
                pick(item, "action", "suggestion"),
                pick(item, "owner", default="待确认"),
                pick(item, "deadline", "review_time", default="待确认"),
                pick(item, "metric", "review_metric", default="待确认"),
            ])
        else:
            backlog_rows.append([priority, item, "来自经营分析", item, "待确认", "待确认", "待确认"])
    write_table(ws, row, "P0/P1/P2/P3 行动清单", ["优先级", "对象", "证据", "动作", "负责人", "截止/复查", "验收指标"], backlog_rows)
    finish_sheet(ws)


def build_appendix_sheet(wb: Workbook, report_data: JsonDict, analysis: JsonDict) -> None:
    """Create the appendix and collection status sheet.

    Args:
        wb: Workbook being generated.
        report_data: Normalized business data.
        analysis: Deterministic analysis output.

    Returns:
        None.

    Raises:
        None.
    """

    ws = wb.create_sheet("附录与采集状态")
    row = sheet_title(ws, "附录与采集状态")
    coverage_rows = []
    coverage = report_data.get("coverage") or analysis.get("coverage")
    for item in as_list(coverage):
        if isinstance(item, dict):
            coverage_rows.append([
                pick(item, "source", "name"),
                pick(item, "range", "period"),
                pick(item, "status"),
                pick(item, "note", "impact"),
            ])
    data_quality = report_data.get("data_quality") or analysis.get("data_quality") or {}
    if not coverage_rows and isinstance(data_quality, dict):
        check_labels = {
            "collection_trace": "本次采集记录",
            "summary_indicators": "经营大盘",
            "funnel": "流量漏斗",
            "region": "区域市场",
            "ads": "广告数据",
            "products": "商品数据",
            "risk": "风险诊断",
            "market_keywords": "市场与关键词",
        }
        checks = data_quality.get("checks") if isinstance(data_quality.get("checks"), dict) else {}
        for key, label in check_labels.items():
            available = bool(checks.get(key))
            coverage_rows.append([
                label,
                pick(report_data.get("meta") or {}, "period_start", default="未返回")
                + " 至 "
                + pick(report_data.get("meta") or {}, "period_end", default="未返回"),
                "已返回" if available else "未返回",
                "可用于经营判断" if available else "对应结论需降级",
            ])
    row = write_table(ws, row, "采集状态", ["数据源", "周期", "状态", "影响"], coverage_rows)
    keywords = analysis.get("keywords_quadrant") or (analysis.get("quadrants") or {}).get("keywords") or {}
    keyword_rows = []
    if isinstance(keywords, dict):
        for group, items in keywords.items():
            for item in as_list(items)[:30]:
                if isinstance(item, dict):
                    keyword_rows.append([
                        group,
                        pick(item, "keyword", "word"),
                        pick(item, "why", "evidence"),
                        "; ".join(text(v) for v in as_list(item.get("actions"))[:2]),
                    ])
    write_table(ws, row, "关键词与资源附录", ["分层", "关键词/资源", "依据", "建议"], keyword_rows)
    finish_sheet(ws)


def build_workbook(report_data: JsonDict, analysis: JsonDict) -> Workbook:
    """Build the workbook in memory.

    Args:
        report_data: Normalized report data.
        analysis: Analysis result.

    Returns:
        An openpyxl workbook ready to save.

    Raises:
        None.
    """

    wb = Workbook()
    build_boss_summary(wb, report_data, analysis)
    build_overview(wb, report_data)
    build_star_sheet(wb, report_data)
    build_flow_sheet(wb, report_data, analysis)
    build_product_sheet(wb, report_data, analysis)
    build_service_sheet(wb, report_data)
    build_backlog_sheet(wb, analysis)
    build_appendix_sheet(wb, report_data, analysis)
    return wb


def find_soffice() -> str:
    """Locate LibreOffice for the required safety re-save.

    Args:
        None.

    Returns:
        Path to a LibreOffice executable.

    Raises:
        RuntimeError: If LibreOffice cannot be found.
    """

    candidates = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)
    raise RuntimeError("LibreOffice/soffice not found; XLSX safety flow cannot finish.")


def libreoffice_resave(path: Path) -> None:
    """Re-save the workbook through LibreOffice headless.

    Args:
        path: Workbook path to replace with the re-saved package.

    Returns:
        None.

    Raises:
        RuntimeError: If LibreOffice conversion fails.
    """

    soffice = find_soffice()
    with tempfile.TemporaryDirectory(prefix="shop_report_lo_") as tmp:
        tmp_dir = Path(tmp)
        input_dir = tmp_dir / "input"
        output_dir = tmp_dir / "output"
        profile_dir = tmp_dir / "profile"
        input_dir.mkdir()
        output_dir.mkdir()
        profile_dir.mkdir()
        source = input_dir / path.name
        shutil.copy2(path, source)
        result = subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(output_dir),
                str(source),
            ],
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice re-save failed: {result.stderr or result.stdout}")
        candidates = list(output_dir.glob("*.xlsx"))
        converted = candidates[0] if candidates else output_dir / path.name
        if not converted.exists():
            raise RuntimeError("LibreOffice did not produce an .xlsx file.")
        shutil.copy2(converted, path)


def strip_xml_residue(xml_bytes: bytes) -> bytes:
    """Remove table/drawing references from XML text.

    Args:
        xml_bytes: Original XML bytes.

    Returns:
        XML bytes with common table/drawing references removed.

    Raises:
        None.
    """

    try:
        text_data = xml_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return xml_bytes
    text_data = re.sub(r"<tableParts\b[^>]*/>", "", text_data)
    text_data = re.sub(r"<drawing\b[^>]*/>", "", text_data)
    text_data = re.sub(r"<Relationship\b[^>]*(?:table|drawing)[^>]*/>", "", text_data, flags=re.IGNORECASE)
    text_data = re.sub(r"<Override\b[^>]*(?:/tables/|/drawings/)[^>]*/>", "", text_data, flags=re.IGNORECASE)
    return text_data.encode("utf-8")


def remove_table_drawing_residue(path: Path) -> None:
    """Remove table and drawing package parts from an XLSX file.

    Args:
        path: Workbook path to sanitize in place.

    Returns:
        None.

    Raises:
        zipfile.BadZipFile: If the workbook is not a valid zip package.
    """

    sanitized = path.with_suffix(".sanitized.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(sanitized, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            lowered = info.filename.lower()
            if lowered.startswith("xl/tables/") or lowered.startswith("xl/drawings/"):
                continue
            data = zin.read(info.filename)
            if lowered.endswith(".xml") or lowered.endswith(".rels"):
                data = strip_xml_residue(data)
            zout.writestr(info, data)
    sanitized.replace(path)


def run_unzip_test(path: Path) -> None:
    """Run ``unzip -t`` to catch corrupt XLSX packages.

    Args:
        path: Workbook path.

    Returns:
        None.

    Raises:
        RuntimeError: If unzip reports corruption.
    """

    result = subprocess.run(["unzip", "-t", str(path)], text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=False)
    if result.returncode != 0:
        raise RuntimeError(f"unzip -t failed: {result.stderr or result.stdout}")


def validate_xlsx(path: Path) -> None:
    """Validate package integrity and absence of table/drawing residue.

    Args:
        path: Workbook path.

    Returns:
        None.

    Raises:
        ValueError: If package residue remains.
        RuntimeError: If zip/openpyxl validation fails.
    """

    run_unzip_test(path)
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        residue = [
            name
            for name in names
            if name.startswith("xl/tables/")
            or name.startswith("xl/drawings/")
            or "tableParts" in zf.read(name).decode("utf-8", errors="ignore")
            if name.endswith((".xml", ".rels"))
        ]
        drawing_rels = [name for name in names if "drawing" in name.lower() and name.endswith(".rels")]
        if residue or drawing_rels:
            raise ValueError(f"Unexpected table/drawing residue found: {residue + drawing_rels}")
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise ValueError(
            f"Workbook sheets must be exactly {EXPECTED_SHEETS}, got {workbook.sheetnames}."
        )
    for worksheet in workbook.worksheets:
        meaningful_rows = [
            row
            for row in worksheet.iter_rows(values_only=True)
            if any(value not in (None, "") for value in row)
        ]
        if len(meaningful_rows) < 5:
            raise ValueError(f"Worksheet has too little business content: {worksheet.title}")
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise ValueError(f"Unexpected formula in {worksheet.title}!{cell.coordinate}")
                if not isinstance(cell.value, str):
                    continue
                lowered = cell.value.lower()
                if any(term.lower() in lowered for term in FORBIDDEN_BUSINESS_TEXT):
                    raise ValueError(
                        f"Forbidden internal wording in {worksheet.title}!{cell.coordinate}: {cell.value}"
                    )
                stripped = cell.value.strip()
                if (
                    (stripped.startswith("{") and stripped.endswith("}"))
                    or (stripped.startswith("[") and stripped.endswith("]"))
                ):
                    raise ValueError(
                        f"Raw JSON-like text in {worksheet.title}!{cell.coordinate}: {cell.value}"
                    )
    workbook.close()


def main(argv: list[str]) -> int:
    """CLI entrypoint.

    Args:
        argv: Command line arguments without the executable name.

    Returns:
        Process exit code.

    Raises:
        None: Errors are printed and converted to non-zero exit codes.
    """

    if len(argv) != 4:
        print(
            "Usage: build_xlsx.py report_data.json analysis.json "
            "management_diagnosis.json output.xlsx",
            file=sys.stderr,
        )
        return 2
    report_data_path = Path(argv[0])
    analysis_path = Path(argv[1])
    diagnosis_path = Path(argv[2])
    output = Path(argv[3])
    output.parent.mkdir(parents=True, exist_ok=True)

    try:
        report_data = load_json(report_data_path)
        analysis = load_json(analysis_path)
        diagnosis = load_json(diagnosis_path)
        # The deterministic files remain required inputs so the caller cannot
        # silently skip data collection. They are intentionally not rendered:
        # every visible row must come from the operations-expert diagnosis.
        if not report_data or not analysis:
            raise ValueError("Report facts and deterministic analysis are required.")
        workbook = build_compact_workbook(
            diagnosis,
            {
                "company": str(
                    (report_data.get("meta") or {}).get("company")
                    or (report_data.get("meta") or {}).get("company_name")
                    or ""
                ),
            },
        )
        protect_formula_cells(workbook)
        workbook.save(output)
        workbook.close()
        libreoffice_resave(output)
        remove_table_drawing_residue(output)
        validate_xlsx(output)
    except Exception as exc:  # noqa: BLE001 - CLI must convert all failures into a clear message.
        print(f"[error] failed to build safe XLSX: {exc}", file=sys.stderr)
        return 1
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

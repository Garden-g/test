#!/usr/bin/env python3
"""Build a safe XLSX workbook for the boss daily pulse skill.

The script turns a normalized JSON file into a boss-facing morning meeting
workbook. It intentionally uses ordinary worksheet cells only. It does not use
Excel Tables, charts, drawings, images, or shapes, because those package parts
are the common source of Excel repair warnings in this project.
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import shutil
import subprocess
import tempfile
import unicodedata
import zipfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


JsonDict = dict[str, Any]
ColumnSpec = tuple[str, tuple[str, ...]]
SheetSpec = dict[str, Any]

TITLE_FILL = "1F4E78"
HEADER_FILL = "5B9BD5"
SECTION_FILL = "DDEBF7"
WARNING_FILL = "FFF2CC"
MISSING_FILL = "F2F2F2"

LOGGER = logging.getLogger("boss_daily_pulse_xlsx")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


SHEETS: list[SheetSpec] = [
    {
        "name": "老板结论",
        "paths": ("boss_conclusion", "boss_summary", "summary", "conclusions"),
        "columns": [
            ("项目", ("item", "name", "module")),
            ("结论", ("conclusion", "summary", "judgement")),
            ("证据", ("evidence", "proof", "data")),
            ("优先级", ("priority", "level")),
            ("负责人", ("owner", "responsible", "assignee")),
            ("今日动作", ("today_action", "action", "next_step")),
            ("复查指标", ("review_metric", "metric", "check")),
        ],
        "fallback": "老板结论未返回；请先补充店铺诊断、大盘、风控、广告、商品或服务数据。",
    },
    {
        "name": "数据覆盖与缺口",
        "paths": ("data_coverage", "coverage", "data_gaps", "access_issues"),
        "columns": [
            ("模块", ("module", "area")),
            ("时间口径", ("period", "date_range", "time_scope")),
            ("状态", ("status", "data_status")),
            ("缺口", ("gap", "issue", "missing")),
            ("影响", ("impact", "business_impact")),
            ("补齐方式", ("how_to_fix", "fix", "next_step")),
        ],
        "fallback": "未提供数据覆盖说明；交付前必须补齐各模块时间口径和缺口影响。",
    },
    {
        "name": "P0今日必须拍板",
        "paths": ("p0_actions", "p0", "must_decide_today"),
        "columns": [
            ("异常/问题", ("issue", "risk", "problem")),
            ("证据", ("evidence", "proof")),
            ("业务影响", ("impact", "business_impact")),
            ("可能原因", ("reason", "hypothesis", "cause")),
            ("负责人", ("owner", "responsible", "assignee")),
            ("今天动作", ("today_action", "action", "next_step")),
            ("复查指标", ("review_metric", "metric", "check")),
        ],
        "fallback": "未识别到 P0；若风控、成交、广告、服务或 Top 商品数据缺失，需要先写明不可判断。",
    },
    {
        "name": "P1今日跟进",
        "paths": ("p1_actions", "p1", "follow_today"),
        "columns": [
            ("异常/机会", ("issue", "opportunity", "problem")),
            ("证据", ("evidence", "proof")),
            ("可能原因", ("reason", "hypothesis", "cause")),
            ("负责人", ("owner", "responsible", "assignee")),
            ("下一步", ("next_step", "action")),
            ("复查时间", ("review_time", "deadline")),
        ],
        "fallback": "未识别到 P1；如样本小或周期不完整，应降级为 P2 或写入数据缺口。",
    },
    {
        "name": "机会窗口",
        "paths": ("opportunities", "opportunity_windows", "growth_windows"),
        "columns": [
            ("机会", ("opportunity", "name")),
            ("证据", ("evidence", "proof")),
            ("为什么值得抓", ("why", "reason", "impact")),
            ("建议承接动作", ("action", "next_step")),
            ("负责人", ("owner", "responsible", "assignee")),
            ("复查指标", ("review_metric", "metric", "check")),
        ],
        "fallback": "未返回机会窗口；不要用经验值硬造区域、商品、访客或关键词机会。",
    },
    {
        "name": "分模块晨会看板",
        "paths": ("module_dashboard", "dashboard", "modules"),
        "columns": [
            ("模块", ("module", "area")),
            ("状态", ("status", "data_status")),
            ("关键指标", ("metric", "kpi", "indicator")),
            ("信号", ("signal", "finding")),
            ("管理判断", ("judgement", "conclusion")),
            ("下一步", ("next_step", "action")),
        ],
        "fallback": "未返回分模块看板；至少应覆盖店铺、流量、商品、区域、广告、订单、服务、团队和风控。",
    },
    {
        "name": "15分钟晨会议程",
        "paths": ("meeting_agenda", "agenda"),
        "columns": [
            ("顺序", ("order", "seq")),
            ("议题", ("topic", "name")),
            ("负责人", ("owner", "responsible", "assignee")),
            ("要拍板的问题", ("decision", "question")),
            ("预计用时", ("duration", "time")),
        ],
        "fallback": "未返回晨会议程；需要按用户确认的晨会时长排列 P0/P1 讨论顺序。",
    },
    {
        "name": "今日行动清单",
        "paths": ("action_items", "actions", "todo"),
        "columns": [
            ("优先级", ("priority", "level")),
            ("动作", ("action", "task")),
            ("负责人", ("owner", "responsible", "assignee")),
            ("截止/复查时间", ("deadline", "review_time")),
            ("验收证据", ("acceptance", "evidence", "proof")),
            ("状态", ("status",)),
        ],
        "fallback": "未返回今日行动清单；每条动作必须有负责人、截止或复查时间、验收证据。",
    },
]

COMPACT_SHEET_NAMES = ["今天先看", "异常原因", "今天行动"]
ALLOWED_ROLES = {"", "运营", "业务"}
ALLOWED_PRIORITIES = {"先做", "随后做", "持续观察"}
FORBIDDEN_DIAGNOSIS_TEXT = (
    "未返回",
    "待确认",
    "不可判断",
    "待补充",
    "TODO",
    "加强运营",
    "持续优化",
)


def configure_logging(verbose: bool) -> None:
    """Configure console logging for repeatable local execution.

    Args:
        verbose: Whether to emit debug-level details.

    Returns:
        None.

    Raises:
        None.
    """

    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )


def load_json(path: Path) -> JsonDict:
    """Load a JSON file as a dictionary.

    Args:
        path: Input JSON path. The file should contain normalized report data.

    Returns:
        A dictionary. Non-dictionary roots are wrapped so the script can still
        create a visible data-gap workbook.

    Raises:
        FileNotFoundError: If the path does not exist.
        json.JSONDecodeError: If the file is not valid JSON.
    """

    LOGGER.info("Reading normalized data: %s", path)
    value = json.loads(path.read_text(encoding="utf-8"))
    return value if isinstance(value, dict) else {"value": value}


def choose_first(data: JsonDict, paths: Sequence[str]) -> Any:
    """Return the first non-empty top-level value from ``data``.

    Args:
        data: Normalized report dictionary.
        paths: Candidate top-level keys, ordered by preference.

    Returns:
        The first present value, or ``None``.

    Raises:
        None.
    """

    for path in paths:
        value = data.get(path)
        if value not in (None, "", [], {}):
            return value
    return None


def as_records(value: Any) -> list[JsonDict]:
    """Normalize arbitrary input into a list of row dictionaries.

    Args:
        value: A list, a wrapper dictionary, a single dictionary, or any other
            value.

    Returns:
        A list of dictionaries ready for table rendering.

    Raises:
        None.
    """

    if isinstance(value, list):
        records: list[JsonDict] = []
        for item in value:
            records.append(item if isinstance(item, dict) else {"value": item})
        return records
    if isinstance(value, dict):
        for key in ("rows", "items", "list", "data", "records"):
            child = value.get(key)
            if isinstance(child, list):
                return as_records(child)
        return [value]
    if value in (None, ""):
        return []
    return [{"value": value}]


def pick(row: JsonDict, keys: Iterable[str], default: str = "未返回") -> Any:
    """Pick the first useful value from a row.

    Args:
        row: Source row dictionary.
        keys: Candidate key names.
        default: Value used when no key has data.

    Returns:
        A raw value for display.

    Raises:
        None.
    """

    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return default


def display(value: Any, limit: int = 260) -> str:
    """Convert a cell value into concise, safe text.

    Args:
        value: Any value from the normalized JSON.
        limit: Maximum characters to keep in a single cell.

    Returns:
        Display string. Missing values become ``未返回``.

    Raises:
        None.
    """

    if value in (None, ""):
        raw = "未返回"
    elif isinstance(value, dict):
        raw = "；".join(f"{display(key, 80)}: {display(item, 160)}" for key, item in value.items())
    elif isinstance(value, list):
        raw = "；".join(display(item, 160) for item in value)
    else:
        raw = str(value)
    raw = raw.replace("\n", " ").strip()
    if raw.startswith(("=", "+", "-", "@")):
        raw = f"'{raw}"
    return raw if len(raw) <= limit else raw[: limit - 1] + "..."


def write_title(ws: Worksheet, title: str, subtitle: str) -> int:
    """Write a merged title band and return the next row.

    Args:
        ws: Target worksheet.
        title: Sheet title.
        subtitle: Period, data status, and generation time.

    Returns:
        Next writable row index.

    Raises:
        None.
    """

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell = ws.cell(1, 1, f"{title}  |  {subtitle}")
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 27
    return 3


def write_summary(ws: Worksheet, row: int, data: JsonDict, args: argparse.Namespace) -> int:
    """Write the common workbook context block.

    Args:
        ws: Target worksheet.
        row: Starting row.
        data: Normalized input data.
        args: Parsed command line arguments.

    Returns:
        Next writable row index.

    Raises:
        None.
    """

    summary_rows = [
        ("对象范围", args.company or display(data.get("company") or data.get("scope"))),
        ("周期", args.period or display(data.get("period") or data.get("date_range"))),
        ("数据状态", display(data.get("data_status") or data.get("status") or "仅覆盖已返回数据")),
        ("生成时间", args.generated_at),
        ("老板结论", display(data.get("one_liner") or data.get("boss_conclusion") or data.get("conclusion"))),
    ]
    for label, value in summary_rows:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
        ws.cell(row, 1).font = Font(name="Arial", size=10, bold=True, color="1F4E78")
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    return row + 1


def write_table(ws: Worksheet, row: int, spec: SheetSpec, data: JsonDict) -> int:
    """Write a normal-cell table for one sheet.

    Args:
        ws: Target worksheet.
        row: Starting row.
        spec: Sheet configuration.
        data: Normalized input data.

    Returns:
        Next writable row index.

    Raises:
        None.
    """

    columns: list[ColumnSpec] = spec["columns"]
    selected = choose_first(data, spec["paths"])
    # `boss_conclusion` is intentionally allowed to be a single sentence. A
    # generic scalar becomes {"value": ...}, which none of the boss columns
    # understands, so normalize it to the explicit business columns here.
    if spec["name"] == "老板结论" and not isinstance(selected, (dict, list)) and selected not in (None, ""):
        records = [{
            "item": "老板结论",
            "conclusion": selected,
            "evidence": data.get("one_liner") or "详见各业务模块",
            "priority": "P0/P1",
            "owner": "老板/对应负责人",
            "today_action": "按今日行动清单执行",
            "review_metric": "按复查指标验收",
        }]
    else:
        records = as_records(selected)
    if not records:
        records = [{"item": "未返回", "conclusion": spec["fallback"], "status": "不可判断"}]

    for col_index, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row, col_index, header)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    for record in records:
        missing_row = all(display(value) == "未返回" for value in record.values())
        for col_index, (_, keys) in enumerate(columns, 1):
            value = pick(record, keys)
            cell = ws.cell(row, col_index, display(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if missing_row or cell.value in {"未返回", "不可判断", "需用户确认"}:
                cell.fill = PatternFill("solid", fgColor=MISSING_FILL)
            if col_index == 1 and str(record.get("priority", "")).upper() in {"P0", "P1"}:
                cell.fill = PatternFill("solid", fgColor=WARNING_FILL)
        row += 1
    return row + 1


def format_worksheet(ws: Worksheet) -> None:
    """Apply safe worksheet formatting.

    Args:
        ws: Worksheet to format.

    Returns:
        None.

    Raises:
        None.
    """

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    max_col = max(ws.max_column, 1)
    max_row = max(ws.max_row, 1)
    ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{max_row}"
    for col_index in range(1, max_col + 1):
        column_letter = get_column_letter(col_index)
        width = 16 if col_index <= 2 else 24
        ws.column_dimensions[column_letter].width = width
    # A merged summary cell spans several physical columns. Build a lookup for
    # its effective width so row-height estimation does not treat it as a
    # narrow single cell.
    merged_widths: dict[str, float] = {}
    for merged_range in ws.merged_cells.ranges:
        start_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        total_width = 0.0
        for col_index in range(merged_range.min_col, merged_range.max_col + 1):
            total_width += 16 if col_index <= 2 else 24
        merged_widths[start_cell.coordinate] = total_width

    for row_index in range(1, max_row + 1):
        if row_index <= 3:
            continue
        estimated_lines = 1
        for cell in ws[row_index]:
            if cell.value in (None, ""):
                continue
            rendered = str(cell.value)
            visual_width = sum(
                2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1
                for char in rendered
            )
            column_width = merged_widths.get(
                cell.coordinate,
                16 if cell.column <= 2 else 24,
            )
            # Excel/LibreOffice wrap Chinese text more aggressively than a
            # simple "column width × 1.25" estimate.  Use a conservative
            # factor so long boss conclusions keep their final line after the
            # workbook is opened and re-saved by a real office application.
            chars_per_line = max(6, int(column_width * 0.95))
            lines = max(1, (visual_width + chars_per_line - 1) // chars_per_line)
            estimated_lines = max(estimated_lines, lines)
        ws.row_dimensions[row_index].height = min(180, max(34, 17 * estimated_lines + 6))


def diagnosis_text(value: Any) -> str:
    """Convert diagnosed content to formula-safe display text.

    Args:
        value: Scalar or short list from the operations-expert diagnosis.

    Returns:
        Safe text. Missing optional values become blank.

    Raises:
        ValueError: Nested objects reach the user-facing renderer.
    """

    if value in (None, ""):
        return ""
    if isinstance(value, dict):
        raise ValueError("Nested objects cannot be rendered into compact workbook cells.")
    if isinstance(value, list):
        rendered = "、".join(
            str(item).strip()
            for item in value[:3]
            if item not in (None, "") and str(item).strip()
        )
    else:
        rendered = re.sub(r"\s+", " ", str(value)).strip()
    return f"'{rendered}" if rendered.startswith(("=", "+", "-", "@")) else rendered


def validate_management_diagnosis(payload: JsonDict) -> None:
    """Validate daily operations-expert diagnosis before delivery.

    Args:
        payload: Canonical diagnosis JSON generated from the daily facts package.

    Returns:
        None when the diagnosis satisfies the compact report contract.

    Raises:
        ValueError: Required content, ownership, grouping, or wording is invalid.
    """

    if not isinstance(payload, dict):
        raise ValueError("Diagnosis root must be an object.")
    conclusion = str(payload.get("executive_conclusion") or "").strip()
    top = payload.get("top_diagnoses") or []
    details = payload.get("detail_diagnoses") or []
    actions = payload.get("actions") or []
    limitations = payload.get("data_limitations") or []

    if not conclusion:
        raise ValueError("executive_conclusion is required.")
    if not isinstance(top, list) or not 1 <= len(top) <= 3:
        raise ValueError("top_diagnoses must contain 1 to 3 rows.")
    if not isinstance(details, list) or not 1 <= len(details) <= 15:
        raise ValueError("detail_diagnoses must contain 1 to 15 rows.")
    if not isinstance(actions, list) or not 1 <= len(actions) <= 8:
        raise ValueError("actions must contain 1 to 8 rows.")
    if not isinstance(limitations, list) or len(limitations) > 3:
        raise ValueError("data_limitations must contain no more than 3 notes.")

    for index, row in enumerate(top):
        if not isinstance(row, dict):
            raise ValueError(f"top_diagnoses[{index}] must be an object.")
        for key in ("what_happened", "root_cause", "solution", "review_standard"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"top_diagnoses[{index}].{key} is required.")
        if str(row.get("role") or "") not in ALLOWED_ROLES:
            raise ValueError(f"top_diagnoses[{index}].role is invalid.")

    diagnosis_pairs: set[tuple[str, str]] = set()
    for index, row in enumerate(details):
        if not isinstance(row, dict):
            raise ValueError(f"detail_diagnoses[{index}] must be an object.")
        for key in ("issue_group", "evidence", "expert_diagnosis", "root_cause", "solution"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"detail_diagnoses[{index}].{key} is required.")
        objects = row.get("objects") or []
        if not isinstance(objects, list) or len(objects) > 3:
            raise ValueError(
                f"detail_diagnoses[{index}].objects must contain at most 3 examples."
            )
        pair = (
            str(row.get("root_cause") or "").strip(),
            str(row.get("solution") or "").strip(),
        )
        if pair in diagnosis_pairs:
            raise ValueError("Duplicate root-cause and solution rows must be grouped.")
        diagnosis_pairs.add(pair)

    action_texts: set[str] = set()
    for index, row in enumerate(actions):
        if not isinstance(row, dict):
            raise ValueError(f"actions[{index}] must be an object.")
        for key in ("priority", "action", "due", "review_standard", "source_issue_group"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"actions[{index}].{key} is required.")
        if str(row.get("priority") or "") not in ALLOWED_PRIORITIES:
            raise ValueError(f"actions[{index}].priority is invalid.")
        if str(row.get("role") or "") not in ALLOWED_ROLES:
            raise ValueError(f"actions[{index}].role is invalid.")
        action = str(row.get("action") or "").strip()
        if action in action_texts:
            raise ValueError("Duplicate action rows must be grouped.")
        action_texts.add(action)

    visible_text = json.dumps(payload, ensure_ascii=False)
    matched = [
        token
        for token in FORBIDDEN_DIAGNOSIS_TEXT
        if token.lower() in visible_text.lower()
    ]
    if matched:
        raise ValueError(f"Placeholder or generic wording is forbidden: {matched}")
    if re.search(
        r"\b(gateway|authorization|access[_ -]?token|traceback|proxy|errorcode)\b",
        visible_text,
        re.I,
    ):
        raise ValueError("Internal execution wording is forbidden.")


def owner_text(item: JsonDict) -> str:
    """Return a verified owner name or the allowed fallback role.

    Args:
        item: One top diagnosis or action row.

    Returns:
        ``owner_name`` when present, otherwise ``运营``/``业务`` or blank.

    Raises:
        None. Role validity is checked before rendering.
    """

    owner_name = str(item.get("owner_name") or "").strip()
    return owner_name or str(item.get("role") or "").strip()


def write_compact_title(
    sheet: Worksheet,
    title: str,
    column_count: int,
    period: str = "",
    conclusion: str | None = None,
) -> None:
    """Write a title and optional first-sheet context.

    Args:
        sheet: Target worksheet.
        title: Natural-language title.
        column_count: Number of visible columns.
        period: Business period shown only when supplied.
        conclusion: Executive conclusion shown only on the first sheet.

    Returns:
        None.

    Raises:
        ValueError: Unsafe nested content reaches the renderer.
    """

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    sheet.cell(1, 1, diagnosis_text(title))
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=TITLE_FILL)
    sheet.cell(1, 1).font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    sheet.cell(1, 1).alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 28

    if period:
        sheet.cell(2, 1, "数据周期")
        sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=column_count)
        sheet.cell(2, 2, diagnosis_text(period))
        sheet.cell(2, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
        sheet.cell(2, 1).font = Font(name="Arial", size=10, bold=True, color="1F4E78")
    if conclusion is not None:
        sheet.cell(3, 1, "一句话结论")
        sheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=column_count)
        sheet.cell(3, 2, diagnosis_text(conclusion))
        sheet.cell(3, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
        sheet.cell(3, 1).font = Font(name="Arial", size=10, bold=True, color="1F4E78")
        sheet.row_dimensions[3].height = 42
    for row_number in (2, 3):
        for column_number in range(1, column_count + 1):
            sheet.cell(row_number, column_number).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def write_compact_table(
    sheet: Worksheet,
    header_row: int,
    headers: list[str],
    rows: list[list[str]],
) -> int:
    """Write one diagnosed table.

    Args:
        sheet: Target worksheet.
        header_row: One-based header row.
        headers: Natural-language labels.
        rows: Diagnosed user-facing rows.

    Returns:
        First row after the table.

    Raises:
        ValueError: A row width differs from the declared headers.
    """

    for column_number, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_number, header)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    sheet.row_dimensions[header_row].height = 28

    for row_offset, values in enumerate(rows, start=1):
        if len(values) != len(headers):
            raise ValueError("Compact table row width does not match its headers.")
        row_number = header_row + row_offset
        for column_number, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column_number, diagnosis_text(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
        sheet.row_dimensions[row_number].height = 54
    end_row = header_row + len(rows)
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, end_row)}"
    )
    sheet.freeze_panes = f"A{header_row + 1}"
    return end_row + 1


def style_compact_sheet(sheet: Worksheet, widths: list[int]) -> None:
    """Apply compact widths and consistent typography.

    Args:
        sheet: Worksheet to style.
        widths: Column widths in display order.

    Returns:
        None.

    Raises:
        None.
    """

    sheet.sheet_view.showGridLines = False
    for column_number, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_number)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            updated_font = copy(cell.font)
            updated_font.name = "Arial"
            updated_font.sz = cell.font.sz or 10
            cell.font = updated_font
            updated_alignment = copy(cell.alignment)
            updated_alignment.wrap_text = True
            updated_alignment.vertical = cell.alignment.vertical or "top"
            cell.alignment = updated_alignment


def build_compact_workbook(
    diagnosis: JsonDict,
    context: JsonDict | None = None,
) -> Workbook:
    """Build the diagnosis-first three-sheet boss morning workbook.

    Args:
        diagnosis: Canonical operations-expert LLM diagnosis.
        context: Optional display context such as company and period.

    Returns:
        In-memory workbook containing only diagnosed, decision-useful content.

    Raises:
        ValueError: Diagnosis quality or table structure is invalid.
    """

    validate_management_diagnosis(diagnosis)
    context = context or {}
    meta = diagnosis.get("report_meta") or {}
    report_title = str(meta.get("title") or "老板经营异常晨会").strip()
    company = str(context.get("company") or "").strip()
    period = str(meta.get("period") or context.get("period") or "").strip()
    display_title = f"{report_title}｜{company}" if company else report_title

    workbook = Workbook()
    home = workbook.active
    home.title = COMPACT_SHEET_NAMES[0]
    diagnosis_sheet = workbook.create_sheet(COMPACT_SHEET_NAMES[1])
    action_sheet = workbook.create_sheet(COMPACT_SHEET_NAMES[2])

    top_rows = diagnosis["top_diagnoses"]
    show_home_owner = any(owner_text(item) for item in top_rows)
    home_headers = ["发生了什么", "核心原因", "解决方案"]
    if show_home_owner:
        home_headers.append("负责角色")
    home_headers.append("复查标准")
    rendered_top: list[list[str]] = []
    for item in top_rows:
        row = [
            diagnosis_text(item.get("what_happened")),
            diagnosis_text(item.get("root_cause")),
            diagnosis_text(item.get("solution")),
        ]
        if show_home_owner:
            row.append(owner_text(item))
        row.append(diagnosis_text(item.get("review_standard")))
        rendered_top.append(row)
    write_compact_title(
        home,
        display_title,
        len(home_headers),
        period,
        str(diagnosis.get("executive_conclusion") or ""),
    )
    next_row = write_compact_table(home, 5, home_headers, rendered_top)
    limitations = [
        diagnosis_text(item)
        for item in diagnosis.get("data_limitations") or []
        if diagnosis_text(item)
    ]
    if limitations:
        home.cell(next_row + 1, 1, "数据限制")
        home.merge_cells(
            start_row=next_row + 1,
            start_column=2,
            end_row=next_row + 1,
            end_column=len(home_headers),
        )
        home.cell(next_row + 1, 2, "；".join(limitations))
        home.cell(next_row + 1, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
        home.cell(next_row + 1, 1).font = Font(bold=True, color="1F4E78")
    style_compact_sheet(
        home,
        [31, 30, 31] + ([14] if show_home_owner else []) + [27],
    )

    detail_headers = [
        "问题类型",
        "重点对象",
        "关键证据",
        "专家判断",
        "核心原因",
        "解决方案",
    ]
    rendered_details = [
        [
            diagnosis_text(item.get("issue_group")),
            diagnosis_text(item.get("objects") or []),
            diagnosis_text(item.get("evidence")),
            diagnosis_text(item.get("expert_diagnosis")),
            diagnosis_text(item.get("root_cause")),
            diagnosis_text(item.get("solution")),
        ]
        for item in diagnosis["detail_diagnoses"]
    ]
    write_compact_title(
        diagnosis_sheet,
        COMPACT_SHEET_NAMES[1],
        len(detail_headers),
    )
    write_compact_table(diagnosis_sheet, 3, detail_headers, rendered_details)
    style_compact_sheet(diagnosis_sheet, [16, 24, 30, 30, 28, 30])

    actions = diagnosis["actions"]
    show_action_owner = any(owner_text(item) for item in actions)
    action_headers = ["优先级", "行动"]
    if show_action_owner:
        action_headers.append("负责角色")
    action_headers.extend(["完成时点", "复查标准"])
    rendered_actions: list[list[str]] = []
    for item in actions:
        row = [
            diagnosis_text(item.get("priority")),
            diagnosis_text(item.get("action")),
        ]
        if show_action_owner:
            row.append(owner_text(item))
        row.extend(
            [
                diagnosis_text(item.get("due")),
                diagnosis_text(item.get("review_standard")),
            ]
        )
        rendered_actions.append(row)
    write_compact_title(
        action_sheet,
        COMPACT_SHEET_NAMES[2],
        len(action_headers),
    )
    write_compact_table(action_sheet, 3, action_headers, rendered_actions)
    style_compact_sheet(
        action_sheet,
        [13, 38] + ([14] if show_action_owner else []) + [16, 32],
    )
    return workbook


def build_workbook(data: JsonDict, args: argparse.Namespace) -> Workbook:
    """Create the boss daily pulse workbook.

    Args:
        data: Normalized input data.
        args: Parsed command line arguments.

    Returns:
        An openpyxl workbook.

    Raises:
        None.
    """

    wb = Workbook()
    subtitle = (
        f"周期：{args.period or display(data.get('period') or data.get('date_range'))}"
        f"；数据状态：{display(data.get('data_status') or data.get('status') or '仅覆盖已返回数据')}"
        f"；生成：{args.generated_at}"
    )
    for index, spec in enumerate(SHEETS):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = spec["name"]
        row = write_title(ws, spec["name"], subtitle)
        row = write_summary(ws, row, data, args)
        write_table(ws, row, spec, data)
        format_worksheet(ws)
    return wb


def resave_with_libreoffice(source: Path, output: Path) -> None:
    """Resave the workbook with LibreOffice headless.

    Args:
        source: Temporary XLSX generated by openpyxl.
        output: Final XLSX path.

    Returns:
        None.

    Raises:
        RuntimeError: If LibreOffice is not installed or conversion fails.
    """

    soffice = (
        shutil.which("soffice")
        or shutil.which("libreoffice")
        or (
            "/Applications/LibreOffice.app/Contents/MacOS/soffice"
            if Path("/Applications/LibreOffice.app/Contents/MacOS/soffice").exists()
            else None
        )
    )
    if not soffice:
        raise RuntimeError("LibreOffice/soffice not found; cannot complete required XLSX safety resave.")
    with tempfile.TemporaryDirectory(prefix="boss_daily_lo_") as temp_dir:
        temp_root = Path(temp_dir)
        out_dir = temp_root / "out"
        profile_dir = temp_root / "profile"
        out_dir.mkdir()
        profile_dir.mkdir()
        command = [soffice, f"-env:UserInstallation={profile_dir.resolve().as_uri()}", "--headless", "--convert-to", "xlsx", "--outdir", str(out_dir), str(source)]
        LOGGER.info("Resaving workbook through LibreOffice.")
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=False)
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice conversion failed: {result.stderr or result.stdout}")
        converted = out_dir / source.name
        if not converted.exists():
            raise RuntimeError("LibreOffice conversion did not create the expected XLSX file.")
        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(converted, output)


def clean_xml_payload(name: str, payload: bytes) -> bytes:
    """Remove table and drawing references from XML package parts.

    Args:
        name: Zip member name.
        payload: Raw member payload.

    Returns:
        Possibly cleaned payload.

    Raises:
        None.
    """

    if not name.endswith((".xml", ".rels")):
        return payload
    try:
        text = payload.decode("utf-8")
    except UnicodeDecodeError:
        return payload
    text = re.sub(r"<tableParts\b[^>]*>.*?</tableParts>", "", text, flags=re.DOTALL)
    text = re.sub(r"<drawing\b[^>]*/>", "", text)
    text = re.sub(r"<legacyDrawing\b[^>]*/>", "", text)
    text = re.sub(r"<Override\b[^>]*(?:/xl/tables/|/xl/drawings/)[^>]*/>", "", text)
    text = re.sub(r"<Relationship\b[^>]*Type=\"[^\"]*/(?:table|drawing)\"[^>]*/>", "", text)
    return text.encode("utf-8")


def clean_xlsx_package(path: Path) -> None:
    """Strip table/drawing package parts that this workbook never needs.

    Args:
        path: XLSX path to clean in place.

    Returns:
        None.

    Raises:
        zipfile.BadZipFile: If the workbook is not a valid zip package.
    """

    LOGGER.info("Cleaning XLSX package residue.")
    temp_path = path.with_suffix(".cleaning.xlsx")
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            name = info.filename
            if name.startswith("xl/tables/") or name.startswith("xl/drawings/"):
                continue
            payload = clean_xml_payload(name, src.read(info))
            dst.writestr(info, payload)
    temp_path.replace(path)


def verify_xlsx(path: Path) -> None:
    """Run the required delivery checks for the workbook.

    Args:
        path: XLSX path to verify.

    Returns:
        None.

    Raises:
        RuntimeError: If any required verification fails.
    """

    LOGGER.info("Verifying XLSX package.")
    unzip_result = subprocess.run(["unzip", "-t", str(path)], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if unzip_result.returncode != 0:
        raise RuntimeError(f"unzip -t failed: {unzip_result.stderr or unzip_result.stdout}")
    workbook = load_workbook(path, data_only=False)
    expected_sheets = COMPACT_SHEET_NAMES
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    forbidden_text = (
        "Gate" + "way", "local" + "host", "Author" + "ization", "cookie", "access " + "token",
        "/mcp" + "/proxy", "br" + "idge", "Trace" + "back", "error" + "Code", "send_" + "msg",
    )
    for sheet in workbook.worksheets:
        if sheet.max_row < 4:
            raise RuntimeError(f"Required business content is missing from sheet: {sheet.title}")
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise RuntimeError(f"Formula cell is not allowed: {sheet.title}!{cell.coordinate}")
                value = str(cell.value or "")
                if any(token.lower() in value.lower() for token in forbidden_text):
                    raise RuntimeError(f"Technical text leaked into {sheet.title}!{cell.coordinate}")
                if value.lstrip().startswith(("{\"", "[{\"")):
                    raise RuntimeError(f"Raw JSON leaked into {sheet.title}!{cell.coordinate}")
                if any(token.lower() in value.lower() for token in FORBIDDEN_DIAGNOSIS_TEXT):
                    raise RuntimeError(
                        f"Placeholder or generic diagnosis leaked into "
                        f"{sheet.title}!{cell.coordinate}"
                    )
    home_sheet = workbook[COMPACT_SHEET_NAMES[0]]
    conclusion_rows = [
        cell.row
        for row in home_sheet.iter_rows()
        for cell in row
        if cell.value == "一句话结论"
    ]
    if not conclusion_rows:
        raise RuntimeError("今天先看 sheet 缺少一句话结论，停止交付。")
    conclusion_value = home_sheet.cell(conclusion_rows[0], 2).value
    if not str(conclusion_value or "").strip():
        raise RuntimeError("今天先看 sheet 的一句话结论为空，停止交付。")
    workbook.close()
    with zipfile.ZipFile(path, "r") as workbook_zip:
        names = workbook_zip.namelist()
        leftovers = [name for name in names if name.startswith("xl/tables/") or name.startswith("xl/drawings/")]
        if leftovers:
            raise RuntimeError(f"Unexpected table/drawing package parts: {leftovers}")
        for name in names:
            if not name.endswith((".xml", ".rels")):
                continue
            payload = workbook_zip.read(name).decode("utf-8", errors="ignore")
            if "tableParts" in payload or "<drawing" in payload or "Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/table\"" in payload:
                raise RuntimeError(f"Unexpected table/drawing reference remains in {name}")


def parse_args() -> argparse.Namespace:
    """Parse command line arguments.

    Args:
        None.

    Returns:
        Parsed arguments.

    Raises:
        SystemExit: If required arguments are missing.
    """

    parser = argparse.ArgumentParser(description="Build a safe boss daily pulse XLSX workbook.")
    parser.add_argument(
        "input_json",
        type=Path,
        help="Operations-expert diagnosis JSON for the workbook.",
    )
    parser.add_argument("output_xlsx", type=Path, help="Destination .xlsx path.")
    parser.add_argument("--company", default="", help="Optional company or shop scope shown in the workbook.")
    parser.add_argument("--period", default="", help="Optional report period shown in each sheet.")
    parser.add_argument("--generated-at", default=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), help="Generation timestamp.")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
    return parser.parse_args()


def main() -> None:
    """Build, resave, clean, and verify the workbook.

    Args:
        None.

    Returns:
        None.

    Raises:
        Any exception from the build or verification steps is allowed to abort
        the command, because failed safety checks must block delivery.
    """

    args = parse_args()
    configure_logging(args.verbose)
    data = load_json(args.input_json)
    workbook = build_compact_workbook(
        data,
        {
            "company": args.company,
            "period": args.period,
            "generated_at": args.generated_at,
        },
    )
    with tempfile.TemporaryDirectory(prefix="boss_daily_xlsx_") as temp_dir:
        source = Path(temp_dir) / "boss_daily_pulse.xlsx"
        workbook.save(source)
        resave_with_libreoffice(source, args.output_xlsx)
    clean_xlsx_package(args.output_xlsx)
    verify_xlsx(args.output_xlsx)
    LOGGER.info("Safe XLSX ready: %s", args.output_xlsx)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Build and validate a plain-cell business workbook from normalized JSON.

The builder is deliberately generic so each Skill can define its own sheet
contract in SKILL.md while sharing the same safe rendering behavior. It never
creates Excel Table objects, charts, images, drawings, macros, or formulas.
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


JsonObject = dict[str, Any]
MISSING = "未返回"
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
FORMULA_PREFIXES = ("=", "+", "-", "@")

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
META_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10, color="333333")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments.

    Returns:
        Parsed input path, output path, and optional expected sheet names.

    Raises:
        SystemExit: If required arguments are missing.
    """

    parser = argparse.ArgumentParser(description="Build a validated business workbook.")
    parser.add_argument("--input", required=True, help="Normalized JSON payload path.")
    parser.add_argument("--output", required=True, help="Destination .xlsx path.")
    parser.add_argument(
        "--expected-sheets",
        default="",
        help="Comma-separated sheet names required by the calling Skill.",
    )
    return parser.parse_args()


def load_payload(path: Path) -> JsonObject:
    """Read and minimally validate the normalized JSON payload.

    Args:
        path: UTF-8 JSON file prepared by the agent.

    Returns:
        Parsed top-level JSON object.

    Raises:
        FileNotFoundError: If the payload does not exist.
        json.JSONDecodeError: If the payload is not valid JSON.
        ValueError: If the top-level object or sheet list is invalid.
    """

    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("输入 JSON 顶层必须是对象。")
    if not isinstance(payload.get("sheets"), list) or not payload["sheets"]:
        raise ValueError("输入 JSON 必须包含非空 sheets 数组。")
    return payload


def safe_text(value: Any) -> Any:
    """Convert missing values and neutralize spreadsheet formula injection.

    Args:
        value: A scalar cell value from the normalized payload.

    Returns:
        A safe scalar. Strings beginning with formula operators are prefixed
        with an apostrophe so Excel treats them as text.

    Raises:
        None.
    """

    if value is None or value == "":
        return MISSING
    if isinstance(value, str) and value.lstrip().startswith(FORMULA_PREFIXES):
        return "'" + value
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def normalize_sheet_name(value: Any) -> str:
    """Create a legal, non-empty Excel worksheet name.

    Args:
        value: Requested sheet name.

    Returns:
        A name no longer than Excel's 31-character limit.

    Raises:
        ValueError: If the normalized name is empty.
    """

    name = INVALID_SHEET_CHARS.sub("_", str(value or "")).strip()[:31]
    if not name:
        raise ValueError("每个 sheet 都必须有有效名称。")
    return name


def normalize_rows(sheet: JsonObject) -> tuple[list[str], list[list[Any]]]:
    """Normalize one sheet's columns and rows into a rectangular matrix.

    Args:
        sheet: Object with optional columns and rows arrays.

    Returns:
        A pair of column labels and row values.

    Raises:
        ValueError: If rows contain unsupported structures.
    """

    raw_rows = sheet.get("rows", [])
    if not isinstance(raw_rows, list):
        raise ValueError("sheet.rows 必须是数组。")
    raw_columns = sheet.get("columns", [])
    columns = [str(item) for item in raw_columns] if isinstance(raw_columns, list) else []

    if raw_rows and all(isinstance(item, dict) for item in raw_rows):
        if not columns:
            columns = list(raw_rows[0].keys())
        rows = [[safe_text(item.get(column)) for column in columns] for item in raw_rows]
    elif raw_rows and all(isinstance(item, list) for item in raw_rows):
        width = max(len(item) for item in raw_rows)
        if not columns:
            columns = [f"字段{index}" for index in range(1, width + 1)]
        rows = [[safe_text(value) for value in item] for item in raw_rows]
    elif raw_rows:
        raise ValueError("sheet.rows 只能由对象或数组行组成。")
    else:
        columns = columns or ["数据状态"]
        rows = [[MISSING]]

    width = len(columns)
    normalized = [row[:width] + [MISSING] * max(0, width - len(row)) for row in rows]
    return columns, normalized


def style_sheet(worksheet, max_row: int, max_column: int) -> None:
    """Apply readable business formatting without creating drawing parts.

    Args:
        worksheet: openpyxl worksheet to style.
        max_row: Last populated row.
        max_column: Last populated column.

    Returns:
        None.

    Raises:
        None.
    """

    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = f"A4:{worksheet.cell(max_row, max_column).coordinate}"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 34

    for row in worksheet.iter_rows(min_row=4, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in worksheet[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, column_cells in enumerate(
        worksheet.iter_cols(min_col=1, max_col=max_column, max_row=max_row),
        start=1,
    ):
        # Row 1 is merged across the title width, so columns after A begin with
        # MergedCell objects that do not expose column_letter. Derive the
        # letter from the stable loop index instead.
        letter = get_column_letter(column_index)
        observed = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(observed + 2, 12), 42)


def build_workbook(payload: JsonObject, expected_sheets: list[str]) -> Workbook:
    """Render the normalized payload as a plain-cell workbook.

    Args:
        payload: Validated input object.
        expected_sheets: Required sheet names declared by the calling Skill.

    Returns:
        An in-memory openpyxl workbook.

    Raises:
        ValueError: If sheet names are duplicated or required sheets are absent.
    """

    workbook = Workbook()
    workbook.remove(workbook.active)
    title = str(payload.get("title") or "业务工作簿")
    meta = payload.get("meta") if isinstance(payload.get("meta"), dict) else {}
    meta_text = "｜".join(
        f"{key}: {safe_text(value)}" for key, value in meta.items()
    ) or "数据状态: 未返回"

    seen: set[str] = set()
    for sheet in payload["sheets"]:
        if not isinstance(sheet, dict):
            raise ValueError("sheets 中的每一项必须是对象。")
        name = normalize_sheet_name(sheet.get("name"))
        if name in seen:
            raise ValueError(f"sheet 名称重复：{name}")
        seen.add(name)
        worksheet = workbook.create_sheet(name)
        columns, rows = normalize_rows(sheet)
        max_column = max(1, len(columns))

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
        worksheet.cell(1, 1, title).fill = TITLE_FILL
        worksheet.cell(1, 1).font = TITLE_FONT
        worksheet.cell(1, 1).alignment = Alignment(vertical="center")
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
        worksheet.cell(2, 1, meta_text).fill = META_FILL
        worksheet.cell(2, 1).font = BODY_FONT
        worksheet.cell(2, 1).alignment = Alignment(vertical="top", wrap_text=True)

        for column_index, label in enumerate(columns, start=1):
            worksheet.cell(4, column_index, safe_text(label))
        for row_index, values in enumerate(rows, start=5):
            for column_index, value in enumerate(values, start=1):
                worksheet.cell(row_index, column_index, safe_text(value))
        style_sheet(worksheet, 4 + len(rows), max_column)

    missing = [name for name in expected_sheets if name not in seen]
    if missing:
        raise ValueError(f"缺少必需 sheet：{', '.join(missing)}")
    return workbook


def validate_output(path: Path, expected_sheets: list[str]) -> None:
    """Verify workbook readability and reject unsafe package residue.

    Args:
        path: Generated workbook path.
        expected_sheets: Required sheet names.

    Returns:
        None.

    Raises:
        ValueError: If the ZIP package is corrupt, expected sheets are missing,
        or table/drawing/macro/formula residue is found.
    """

    with zipfile.ZipFile(path) as archive:
        corrupt_member = archive.testzip()
        if corrupt_member:
            raise ValueError(f"XLSX ZIP 损坏：{corrupt_member}")
        names = archive.namelist()
        blocked_parts = [
            name for name in names
            if name.startswith(("xl/tables/", "xl/drawings/", "xl/vbaProject"))
        ]
        if blocked_parts:
            raise ValueError(f"发现不允许的包内对象：{blocked_parts}")
        for name in names:
            if not name.endswith(".xml"):
                continue
            xml = archive.read(name)
            if b"tableParts" in xml or b"<drawing" in xml or b"<f>" in xml:
                raise ValueError(f"发现表格、绘图或公式残留：{name}")

    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        missing = [name for name in expected_sheets if name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"生成文件缺少必需 sheet：{', '.join(missing)}")
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        raise ValueError(f"发现意外公式：{worksheet.title}!{cell.coordinate}")
    finally:
        workbook.close()


def main() -> int:
    """Build the workbook, save it atomically enough for local delivery, and validate it.

    Returns:
        Process status code 0 on success.

    Raises:
        OSError: If input/output files cannot be read or written.
        ValueError: If payload or workbook validation fails.
    """

    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError("输出文件必须使用 .xlsx 扩展名。")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    expected = [item.strip() for item in args.expected_sheets.split(",") if item.strip()]

    payload = load_payload(input_path)
    workbook = build_workbook(payload, expected)
    workbook.save(output_path)
    workbook.close()
    validate_output(output_path, expected)
    logging.info("Workbook generated and validated: %s", output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

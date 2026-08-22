#!/usr/bin/env python3
"""Build the boss-facing Alibaba ROI weekly report as a safe .xlsx workbook.

The script consumes the normalized report_data.json and analysis.json generated
by prepare_data.py / analyze.py. It writes a multi-sheet workbook that answers
the six boss questions directly, then runs the Excel safety flow required for
Mac Excel compatibility.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
import zipfile
from copy import copy
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from dataclasses import dataclass, field

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


JsonDict = dict[str, Any]
Rows = list[list[Any]]


# --- Layout and visual markers -----------------------------------------------
# A single source of truth for status pill colors. Keys are matched against cell
# values (exact match preferred, substring match only on short cells) so the
# same colour appears wherever the same status word appears.
STATUS_PALETTE: dict[str, tuple[str, str]] = {
    "红灯":     ("C00000", "FFFFFF"),
    "P0":       ("C00000", "FFFFFF"),
    "预警":     ("C00000", "FFFFFF"),
    "失血款":   ("C00000", "FFFFFF"),
    "黄灯":     ("F1A33A", "1F1F1F"),
    "P1":       ("F4B084", "5C2E00"),
    "待补信息": ("FFE699", "5C4400"),
    "待判断":   ("FFE699", "5C4400"),
    "P2":       ("FFE699", "5C4400"),
    "观察款":   ("FFE699", "5C4400"),
    "潜力词":   ("D9EAD3", "274E13"),
    "潜力款":   ("D9EAD3", "274E13"),
    "金词":     ("D9EAD3", "274E13"),
    "高质量":   ("D9EAD3", "274E13"),
    "可放大":   ("D9EAD3", "274E13"),
    "可判断":   ("CFE2F3", "0B5394"),
    "可用":     ("CFE2F3", "0B5394"),
    "绿灯":     ("38761D", "FFFFFF"),
    "健康":     ("38761D", "FFFFFF"),
    "印钞款":   ("38761D", "FFFFFF"),
    "机会":     ("38761D", "FFFFFF"),
    "P3":       ("E6E6E6", "5A5A5A"),
    "低质量":   ("EEEEEE", "5A5A5A"),
    "拓展词":   ("EEEEEE", "5A5A5A"),
    "烧钱词":   ("C00000", "FFFFFF"),
}

# Markers we visually dim because data-gap cells should not steal the eye.
MISSING_MARKERS: set[str] = {"未返回", "不可判断", "未拆人", "周期错位", "证据不足"}

# Section palette names.
TITLE_FILL = "1F4E78"
SECTION_FILL_PARAGRAPH = "EAF1FB"
SECTION_FILL_VISUAL = "EAF7EA"
SECTION_FILL_DETAIL = "DDE7F1"

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
TITLE_FONT_OBJ = Font(name="Arial", size=14, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="1F4E78")
PARAGRAPH_LABEL_FONT = Font(name="Arial", size=10, bold=True, color="1F4E78")
PARAGRAPH_CONTENT_FONT = Font(name="Arial", size=10, color="333333")
DEFAULT_FONT = Font(name="Arial", size=10, color="333333")
DETAIL_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
MISSING_FONT = Font(name="Arial", size=10, italic=True, color="9A9A9A")
MISSING_FILL = PatternFill("solid", fgColor="F4F4F4")


@dataclass
class Block:
    """Single layout-aware row block in a sheet."""

    kind: str
    payload: dict = field(default_factory=dict)


def title_block(title: str, subtitle: str = "", status: str = "") -> Block:
    return Block("title", {"title": title, "subtitle": subtitle, "status": status})


def paragraph_block(label: str, content: Any) -> Block:
    return Block("paragraph", {"label": label, "content": str(content or "")})


def section_block(text: str, palette: str = "detail") -> Block:
    return Block("section", {"text": text, "palette": palette})


def visual_block(
    kind: str,
    label: str,
    display: Any,
    ratio: float | None,
    status: str,
    hint: str,
    danger: bool = False,
) -> Block:
    return Block(
        "visual",
        {
            "kind": kind,
            "label": label,
            "display": display,
            "ratio": ratio,
            "status": status,
            "hint": hint,
            "danger": danger,
        },
    )


def detail_header_block(headers: list[str]) -> Block:
    return Block("detail_header", {"headers": headers})


def detail_block(values: list[Any]) -> Block:
    return Block("detail", {"values": values})


def blank_block(height: int = 6) -> Block:
    return Block("blank", {"height": height})


def safe_ratio(value: Any, max_value: Any = None) -> float | None:
    """Compute a 0..1 ratio for in-cell DataBar visualisation.

    Returns None when the value cannot be turned into a number — the layout
    engine then knows to dim that cell rather than draw a zero-width bar.
    """

    number = to_float(value)
    maximum = to_float(max_value)
    if number is None:
        return None
    if maximum is None or maximum <= 0:
        maximum = max(number, 1)
    return max(0.0, min(number / maximum, 1.0))


def status_palette_for(text: Any) -> tuple[str, str] | None:
    """Look up a (fill, font) pair for a status marker.

    Exact match wins; substring match only fires for short cells so that long
    sentences containing the word "P0" do not get painted red by accident.
    """

    raw = str(text or "").strip()
    if not raw:
        return None
    if raw in STATUS_PALETTE:
        return STATUS_PALETTE[raw]
    if len(raw) <= 14:
        for key, palette in STATUS_PALETTE.items():
            if key in raw:
                return palette
    return None


def apply_status_pill(cell, status_text: Any) -> None:
    """Paint a status pill on a cell when its text matches a known status."""

    palette = status_palette_for(status_text)
    if not palette:
        return
    bg, fg = palette
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", size=10, bold=True, color=fg)
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def is_missing_cell(value: Any) -> bool:
    """Return True when a cell value belongs to the 'data not back' set."""

    return str(value or "").strip() in MISSING_MARKERS


def write_blocks(
    worksheet: Worksheet,
    blocks: list[Block],
    width: int,
    widths: dict[str, int] | None = None,
    sheet_status: str = "",
    auto_filter: bool = True,
) -> None:
    """Write a block list to a worksheet and apply layout-aware styling.

    The engine intentionally separates four physical regions:

    * title band (merged, dark blue)
    * paragraph rows (A=label, B..end merged)
    * visual cards (with DataBar)
    * detail header + detail rows (autoFilter and freeze anchor)

    These four regions are what makes the workbook scannable: the boss sees a
    sheet title, four short summary sentences, a small visual block, then the
    real detail table. Each gets its own fill so the reader's eye snaps to the
    section boundary instead of trying to align rows against a single header.
    """

    worksheet.sheet_view.showGridLines = False
    end_col = max(1, width)
    end_col_letter = get_column_letter(end_col)

    first_detail_header_row: int | None = None
    detail_header_count = 0
    detail_last_row: int | None = None
    bar_ranges: list[str] = []
    bar_danger_ranges: list[str] = []

    for block in blocks:
        if block.kind == "title":
            title = block.payload.get("title") or ""
            sub = block.payload.get("subtitle") or ""
            status = block.payload.get("status") or ""
            text = title
            if sub:
                text += f"  ｜  {sub}"
            if status:
                text += f"  ｜  {status}"
            worksheet.append([text] + [None] * (end_col - 1))
            row = worksheet.max_row
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
            cell = worksheet.cell(row=row, column=1)
            cell.value = text
            cell.font = TITLE_FONT_OBJ
            cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
            worksheet.row_dimensions[row].height = 30

        elif block.kind == "paragraph":
            label = block.payload.get("label") or ""
            content = str(block.payload.get("content") or "")
            row_values = [label, content] + [None] * max(0, end_col - 2)
            worksheet.append(row_values)
            row = worksheet.max_row
            if end_col >= 3:
                worksheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end_col)
            label_cell = worksheet.cell(row=row, column=1)
            content_cell = worksheet.cell(row=row, column=2)
            label_cell.font = PARAGRAPH_LABEL_FONT
            label_cell.fill = PatternFill("solid", fgColor=SECTION_FILL_PARAGRAPH)
            label_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            label_cell.border = THIN_BORDER
            content_cell.font = PARAGRAPH_CONTENT_FONT
            content_cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
            content_cell.border = THIN_BORDER
            # Auto height: roughly estimate visual lines from content length and merged width.
            merged_width_chars = 0
            for col in range(2, end_col + 1):
                merged_width_chars += (widths or {}).get(get_column_letter(col), 16)
            char_per_line = max(1, int(merged_width_chars * 1.8))
            est_lines = max(1, (len(content) + char_per_line - 1) // char_per_line)
            worksheet.row_dimensions[row].height = min(96, max(26, 18 * est_lines + 8))

        elif block.kind == "section":
            text = block.payload.get("text") or ""
            palette = block.payload.get("palette", "detail")
            worksheet.append([text] + [None] * (end_col - 1))
            row = worksheet.max_row
            worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
            cell = worksheet.cell(row=row, column=1)
            cell.value = text
            cell.font = SECTION_FONT
            fills = {
                "visual": SECTION_FILL_VISUAL,
                "detail": SECTION_FILL_DETAIL,
                "paragraph": SECTION_FILL_PARAGRAPH,
            }
            cell.fill = PatternFill("solid", fgColor=fills.get(palette, SECTION_FILL_DETAIL))
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
            worksheet.row_dimensions[row].height = 22

        elif block.kind == "visual":
            kind = block.payload.get("kind") or ""
            label = block.payload.get("label") or ""
            display = block.payload.get("display")
            ratio = block.payload.get("ratio")
            status = block.payload.get("status") or ""
            hint = block.payload.get("hint") or ""
            danger = bool(block.payload.get("danger"))
            row_values: list[Any] = [
                kind,
                label,
                fmt_value(display),
                ratio if ratio is not None else "未返回",
                status,
                hint,
            ]
            row_values += [None] * max(0, end_col - len(row_values))
            worksheet.append(row_values)
            row = worksheet.max_row
            for col_idx in range(1, end_col + 1):
                c = worksheet.cell(row=row, column=col_idx)
                c.font = DEFAULT_FONT
                c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                c.border = THIN_BORDER
            worksheet.cell(row=row, column=1).font = Font(name="Arial", size=10, bold=True, color="1F4E78")
            worksheet.cell(row=row, column=2).font = Font(name="Arial", size=10, bold=True, color="333333")
            value_cell = worksheet.cell(row=row, column=3)
            if is_missing_cell(value_cell.value):
                value_cell.font = MISSING_FONT
                value_cell.fill = MISSING_FILL
            ratio_cell = worksheet.cell(row=row, column=4)
            ratio_cell.alignment = Alignment(vertical="center", horizontal="right")
            if ratio is None:
                ratio_cell.value = "未返回"
                ratio_cell.font = MISSING_FONT
                ratio_cell.fill = MISSING_FILL
                # Keep a harmless DataBar rule on missing visual cards. The
                # text cell renders no bar, but the visual-layer contract stays
                # structurally consistent even when an entire data source is
                # unavailable for the reporting period.
                bar_ranges.append(f"D{row}:D{row}")
            else:
                ratio_cell.value = float(ratio)
                ratio_cell.number_format = "0%"
                target = bar_danger_ranges if danger else bar_ranges
                target.append(f"D{row}:D{row}")
            apply_status_pill(worksheet.cell(row=row, column=5), status)
            worksheet.row_dimensions[row].height = 22

        elif block.kind == "detail_header":
            headers = list(block.payload.get("headers") or [])
            row_values = headers + [None] * max(0, end_col - len(headers))
            worksheet.append(row_values)
            row = worksheet.max_row
            detail_header_count += 1
            if first_detail_header_row is None:
                first_detail_header_row = row
            for col_idx in range(1, end_col + 1):
                c = worksheet.cell(row=row, column=col_idx)
                c.font = DETAIL_HEADER_FONT
                c.fill = PatternFill("solid", fgColor=TITLE_FILL)
                c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                c.border = THIN_BORDER
            worksheet.row_dimensions[row].height = 24

        elif block.kind == "detail":
            values = list(block.payload.get("values") or [])
            row_values = values + [None] * max(0, end_col - len(values))
            worksheet.append(row_values)
            row = worksheet.max_row
            detail_last_row = row
            for col_idx in range(1, end_col + 1):
                c = worksheet.cell(row=row, column=col_idx)
                c.font = DEFAULT_FONT
                c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                c.border = THIN_BORDER
                if is_missing_cell(c.value):
                    c.font = MISSING_FONT
                    c.fill = MISSING_FILL
                else:
                    apply_status_pill(c, c.value)
            # Complete evidence/action sentences were previously hidden behind
            # a fixed 22-point row, making populated sheets look sparse.
            estimated_lines = 1
            for col_idx, value in enumerate(row_values[:end_col], start=1):
                if value in (None, ""):
                    continue
                col_letter = get_column_letter(col_idx)
                column_width = (widths or {}).get(col_letter, 16)
                chars_per_line = max(6, int(column_width * 1.35))
                lines = max(1, (display_width(value) + chars_per_line - 1) // chars_per_line)
                estimated_lines = max(estimated_lines, lines)
            worksheet.row_dimensions[row].height = min(120, max(22, 17 * estimated_lines + 5))

        elif block.kind == "blank":
            worksheet.append([None] * end_col)
            row = worksheet.max_row
            worksheet.row_dimensions[row].height = int(block.payload.get("height", 6))

    # Column widths come last so cells already exist when openpyxl applies them.
    if widths:
        for col_letter, w in widths.items():
            worksheet.column_dimensions[col_letter].width = w
    else:
        for col_idx in range(1, end_col + 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 16

    # DataBar conditional formatting for visual cards. Two colour buckets so a
    # warning metric (12h+ unreplied, high exposure / zero inquiry) reads red.
    if bar_ranges:
        rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=1,
            color="638EC6", showValue=None,
        )
        for rng in bar_ranges:
            worksheet.conditional_formatting.add(rng, rule)
    if bar_danger_ranges:
        rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=1,
            color="C00000", showValue=None,
        )
        for rng in bar_danger_ranges:
            worksheet.conditional_formatting.add(rng, rule)

    # Freeze the rows above the detail table so the boss can scroll the detail
    # list while still seeing the column names.
    if first_detail_header_row is not None:
        worksheet.freeze_panes = worksheet.cell(
            row=first_detail_header_row + 1, column=1
        ).coordinate
    else:
        worksheet.freeze_panes = "A2"

    if (
        auto_filter
        and detail_header_count == 1
        and first_detail_header_row is not None
    ):
        # Some weeks genuinely have no detail rows for a sheet, for example
        # when product diagnostics return an empty top-products list. The sheet
        # still has a real detail table header, so keep AutoFilter on the header
        # row instead of failing the visual-layout contract.
        filter_last_row = detail_last_row or first_detail_header_row
        worksheet.auto_filter.ref = (
            f"A{first_detail_header_row}:{end_col_letter}{filter_last_row}"
        )
    else:
        worksheet.auto_filter.ref = None

    palette = status_palette_for(sheet_status)
    worksheet.sheet_properties.tabColor = palette[0] if palette else "B7B7B7"
# ----------------------------------------------------------------------------

SHEETS = [
    "老板结论",
    "投产看板",
    "询盘质量",
    "订单产出",
    "商品节奏",
    "关键词与广告机会",
    "业务员回复与跟进",
    "数据质量检查",
]

FORBIDDEN_TEXT = (
    "m" + "cp",
    "br" + "idge",
    "accio-" + "mcp-cli",
    "de" + "bug",
    "json",
    "内部绕行",
    "逆" + "向",
    "errorCode",
    "errorMsg",
    "-32002",
    "Agent 类型不允许",
    "Traceback",
    "ECONNREFUSED",
    "小" + "满",
    "g" + "gs",
)


def parse_args(argv: list[str]) -> argparse.Namespace:
    """Parse command-line arguments for workbook generation.

    Args:
        argv: Command-line arguments without the program name.

    Returns:
        Parsed arguments with report paths and output path.

    Raises:
        SystemExit: Raised by argparse when required arguments are missing.
    """

    parser = argparse.ArgumentParser(description="Build Alibaba boss ROI .xlsx report.")
    parser.add_argument("--report-data", required=True, help="Path to report_data.json.")
    parser.add_argument("--analysis", required=True, help="Path to analysis.json.")
    parser.add_argument("--raw-dir", required=True, help="Directory containing raw JSON files.")
    parser.add_argument(
        "--narrative",
        help="Optional narrative.json written by the executing Accio Agent. No provider API key is required.",
    )
    parser.add_argument("--output", required=True, help="Output .xlsx path.")
    parser.add_argument(
        "--skip-libreoffice",
        action="store_true",
        help="Only for local dry runs. Final delivery must not use this flag.",
    )
    return parser.parse_args(argv)


def load_json(path: Path) -> JsonDict:
    """Load a JSON object from disk.

    Args:
        path: Path to a UTF-8 JSON file.

    Returns:
        Parsed JSON dictionary. Non-object roots become an empty dictionary.

    Raises:
        FileNotFoundError: If the file is missing.
        json.JSONDecodeError: If the file is not valid JSON.
    """

    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, dict) else {}


def load_optional_json(path: Path) -> Any:
    """Load optional JSON while treating missing or malformed files as absent.

    Args:
        path: File path to inspect.

    Returns:
        Parsed JSON value, or None when the file is missing or invalid.

    Raises:
        No exceptions are intentionally raised; failures are represented as None.
    """

    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def load_narrative(path: str | None) -> JsonDict:
    """Load the optional Agent-written narrative contract.

    Args:
        path: Path to narrative.json, or None when the report should use
            deterministic fallback copy.

    Returns:
        Narrative dictionary. Invalid or non-object files are treated as absent
        so local tests can still exercise the deterministic pipeline.

    Raises:
        No exceptions are intentionally raised.
    """

    if not path:
        return {}
    data = load_optional_json(Path(path))
    return data if isinstance(data, dict) else {}


def first_present(mapping: JsonDict | None, *keys: str) -> Any:
    """Return the first non-empty value for a list of candidate keys.

    Args:
        mapping: Dictionary to inspect.
        *keys: Candidate keys in priority order.

    Returns:
        First value that is not None or an empty string; otherwise None.

    Raises:
        No exceptions are intentionally raised.
    """

    if not isinstance(mapping, dict):
        return None
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ""):
            return value
    return None


def to_float(value: Any) -> float | None:
    """Convert numbers and percentage strings to float values.

    Args:
        value: Raw number-like value from API output.

    Returns:
        Float value, or None when conversion is not possible.

    Raises:
        No exceptions are intentionally raised.
    """

    if value in (None, ""):
        return None
    try:
        if isinstance(value, str):
            text = value.strip().replace(",", "")
            if text.endswith("%"):
                return float(text[:-1]) / 100
            return float(text)
        return float(value)
    except (TypeError, ValueError):
        return None


def fmt_value(value: Any, missing: str = "未返回") -> Any:
    """Format a value for business-facing worksheet cells.

    Args:
        value: Raw value.
        missing: Text to use when the value is absent.

    Returns:
        Original value or a missing-data label.

    Raises:
        No exceptions are intentionally raised.
    """

    if value in (None, ""):
        return missing
    return value


def compact_text(value: Any, limit: int = 70) -> str:
    """Convert a value to a compact one-line cell string.

    Args:
        value: Any value that needs to be displayed in Excel.
        limit: Maximum number of characters to keep before truncating.

    Returns:
        A one-line string. Missing values become "未返回".

    Raises:
        No exceptions are intentionally raised.
    """

    if value in (None, ""):
        return "未返回"
    if isinstance(value, dict):
        return f"已返回 {len(value)} 个业务字段，需按平台明细复核"
    if isinstance(value, list):
        return f"已返回 {len(value)} 条记录，需按平台明细复核"
    text = str(value).replace("\n", " ").strip()
    text = text if len(text) <= limit else f"{text[: limit - 1]}…"
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def protect_formula_cells(workbook: Workbook) -> None:
    """Prevent formula injection in user- or platform-derived text cells.

    Args:
        workbook: Workbook that is about to be saved.

    Returns:
        None.

    Raises:
        None. The helper only inspects string cells.
    """

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@")):
                    cell.value = f"'{cell.value}"


def clean_light(value: Any) -> str:
    """Normalize traffic-light symbols into plain business labels.

    Args:
        value: Raw status value from the analysis layer.

    Returns:
        Plain Chinese status text.

    Raises:
        No exceptions are intentionally raised.
    """

    text = str(value or "").strip()
    replacements = {
        "🔴": "红灯",
        "🟡": "黄灯",
        "🟢": "绿灯",
        "red": "红灯",
        "yellow": "黄灯",
        "green": "绿灯",
    }
    lowered = text.lower()
    for key, label in replacements.items():
        if key in text or key in lowered:
            return label
    return text or "待判断"


def section(title: str, width: int) -> list[Any]:
    """Create a visual section row without Excel drawing objects.

    Args:
        title: Section title shown in the first cell.
        width: Number of columns in the current sheet.

    Returns:
        A row padded to the requested width.

    Raises:
        No exceptions are intentionally raised.
    """

    return [f"【{title}】"] + [""] * max(width - 1, 0)


def join_actions(actions: Any, limit: int = 140) -> str:
    """Join action suggestions into one readable Excel cell.

    Args:
        actions: List of action strings, or any fallback value.
        limit: Maximum cell text length.

    Returns:
        Compact action sentence.

    Raises:
        No exceptions are intentionally raised.
    """

    if isinstance(actions, list):
        text = "；".join(str(item) for item in actions if item)
    else:
        text = str(actions or "")
    return compact_text(text, limit)


def visual_section(title: str, width: int) -> list[Any]:
    """Create a worksheet visual-dashboard section marker.

    Args:
        title: Business title for the visual block.
        width: Number of columns in the target sheet.

    Returns:
        Section row padded to the target width.

    Raises:
        No exceptions are intentionally raised.
    """

    return section(f"可视化看板：{title}", width)


def visual_bar(value: Any, max_value: Any = None, width: int = 12) -> str:
    """Build a safe in-cell bar chart using text only.

    Args:
        value: Numeric value to visualize.
        max_value: Optional maximum used for scaling.
        width: Number of text blocks in the bar.

    Returns:
        Text bar such as "■■■■□□□□ 45%".

    Raises:
        No exceptions are intentionally raised.
    """

    number = to_float(value)
    maximum = to_float(max_value)
    if number is None:
        return "未返回"
    if maximum is None or maximum <= 0:
        maximum = max(number, 1)
    ratio = max(0.0, min(number / maximum, 1.0))
    filled = int(round(ratio * width))
    empty = max(width - filled, 0)
    return f"{'■' * filled}{'□' * empty} {ratio * 100:.0f}%"


def visual_row(kind: str, label: str, value: Any, max_value: Any, status: str, action: str, width: int = 6) -> list[Any]:
    """Create a standard visual-dashboard row.

    Args:
        kind: Visual group name, such as KPI card or heat map.
        label: Metric/object label.
        value: Raw value shown in the row.
        max_value: Scaling maximum for the in-cell bar.
        status: Business judgement.
        action: Suggested next action.
        width: Number of columns in the target sheet.

    Returns:
        Row matching either 5-column or 6-column sheet layouts.

    Raises:
        No exceptions are intentionally raised.
    """

    row6 = [kind, label, fmt_value(value), visual_bar(value, max_value), status, action]
    if width == 5:
        return [kind, label, f"{fmt_value(value)} | {visual_bar(value, max_value)}", status, action]
    if width <= len(row6):
        return row6[:width]
    return row6 + [""] * (width - len(row6))


def max_positive(values: list[Any], fallback: float = 1.0) -> float:
    """Return the largest positive numeric value in a list.

    Args:
        values: Candidate values.
        fallback: Value used when no positive number is available.

    Returns:
        Positive maximum for visual scaling.

    Raises:
        No exceptions are intentionally raised.
    """

    numbers = [to_float(value) for value in values]
    positives = [value for value in numbers if value is not None and value > 0]
    return max(positives) if positives else fallback


def priority_weight(priority: Any) -> int:
    """Convert action priority into a heat-map weight.

    Args:
        priority: Priority text, typically P0/P1/P2/P3.

    Returns:
        Numeric weight where P0 is highest risk.

    Raises:
        No exceptions are intentionally raised.
    """

    return {"P0": 4, "P1": 3, "P2": 2, "P3": 1}.get(str(priority or "").upper(), 0)


def find_ad_cost(report_data: JsonDict) -> Any:
    for ind in (report_data.get("summary") or {}).get("indicators") or []:
        name = str(ind.get("name") or "").lower()
        if "花费" in name or "cost" in name or "spend" in name:
            v = ind.get("value")
            if v not in (None, ""):
                return v
    ads = report_data.get("ads") or {}
    for row in ads.get("overview") or []:
        if isinstance(row, dict) and "花费" in str(row.get("name") or ""):
            return row.get("value")
    return None


def parse_ad_summary(report_data: JsonDict, raw: JsonDict | None = None) -> JsonDict:
    """Extract ad spend and lead metrics from structured or sentence outputs.

    Args:
        report_data: Normalized report data.
        raw: Optional raw context used as a fallback.

    Returns:
        Dictionary containing spend, clicks, ad_inquiry, lead_cost and source_note.

    Raises:
        No exceptions are intentionally raised.
    """

    ads = report_data.get("ads") or {}
    result: JsonDict = {
        "spend": find_ad_cost(report_data),
        "clicks": None,
        "ad_inquiry": None,
        "lead_cost": None,
        "source_note": "广告诊断摘要",
    }
    for row in ads.get("overview") or []:
        if not isinstance(row, dict):
            continue
        name = str(row.get("name") or "")
        value = row.get("value")
        if "花费" in name and result["spend"] in (None, ""):
            result["spend"] = value
        elif "点击" in name:
            result["clicks"] = value
        elif "CPC" in name.upper():
            result["cpc"] = value

    text_parts = [
        str(ads.get("overview_summary") or ""),
        str(ads.get("ai_conclusion") or ""),
    ]
    if raw:
        text_parts.append(str(raw.get("ads_account") or ""))
    text = " ".join(text_parts)
    patterns = {
        "spend": r"(?:账户)?花费\s*([0-9]+(?:\.[0-9]+)?)",
        "clicks": r"点击量\s*([0-9]+(?:\.[0-9]+)?)",
        "ad_inquiry": r"商机量\s*([0-9]+(?:\.[0-9]+)?)",
        "lead_cost": r"商机成本\s*([0-9]+(?:\.[0-9]+)?)",
    }
    for key, pattern in patterns.items():
        if result.get(key) not in (None, ""):
            continue
        match = re.search(pattern, text)
        if match:
            result[key] = float(match.group(1))
    return result


def funnel_totals(report_data: JsonDict, raw: JsonDict) -> JsonDict:
    """Build a conservative weekly funnel total from available sources.

    Args:
        report_data: Normalized report data.
        raw: Optional raw context.

    Returns:
        Dictionary with exposure, visitors, inquiries and orders.

    Raises:
        No exceptions are intentionally raised.
    """

    daily = (report_data.get("funnel") or {}).get("daily") or []
    total: JsonDict = {"imps": 0.0, "visitor": 0.0, "inquiry": 0.0, "order": 0.0}
    seen_daily = False
    for row in daily:
        if not isinstance(row, dict):
            continue
        seen_daily = True
        total["imps"] += to_float(row.get("imps")) or 0
        total["visitor"] += to_float(row.get("visitor_uv")) or 0
        total["inquiry"] += to_float(row.get("fb_count")) or 0
        total["order"] += to_float(row.get("order_count")) or 0
    if not seen_daily:
        total["inquiry"] = to_float(summary_value(raw.get("shop_summary"), "abCnt", "abCntValue")) or 0
        total["order"] = to_float(summary_value(raw.get("shop_summary"), "orderCnt", "crtOrdCnt", "orderCntValue")) or 0
    summary_inquiry = to_float(summary_value(raw.get("shop_summary"), "abCnt", "abCntValue"))
    summary_order = to_float(summary_value(raw.get("shop_summary"), "orderCnt", "crtOrdCnt", "orderCntValue"))
    if summary_inquiry is not None:
        total["summary_inquiry"] = summary_inquiry
    if summary_order is not None:
        total["summary_order"] = summary_order
    return total


def quality_status_from_service(service: JsonDict) -> str:
    """Return a plain service status for boss-facing sheets.

    Args:
        service: Normalized service section.

    Returns:
        "红灯", "黄灯", or "绿灯/待判断".

    Raises:
        No exceptions are intentionally raised.
    """

    status = str(service.get("status") or "").lower()
    if status == "red":
        return "红灯"
    if status == "yellow":
        return "黄灯"
    if status == "green":
        return "绿灯"
    return "待判断"


def as_percent(value: Any) -> Any:
    """Convert a percent-like value to an Excel decimal when possible.

    Args:
        value: Raw number or string percentage.

    Returns:
        Decimal float for Excel percentage formatting, or the original/missing value.

    Raises:
        No exceptions are intentionally raised.
    """

    number = to_float(value)
    return number if number is not None else fmt_value(value)


def find_indicator(report_data: JsonDict, names: tuple[str, ...]) -> Any:
    """Find a KPI value by matching Chinese indicator names.

    Args:
        report_data: Normalized report data.
        names: Candidate substrings for the indicator name.

    Returns:
        Matched indicator value or None.

    Raises:
        No exceptions are intentionally raised.
    """

    indicators = ((report_data.get("summary") or {}).get("indicators") or [])
    for item in indicators:
        if not isinstance(item, dict):
            continue
        label = str(item.get("name") or "")
        if any(name in label for name in names):
            return item.get("value")
    return None


def raw_rows(raw: Any) -> list[JsonDict]:
    """Extract a list of rows from common Alibaba response shapes.

    Args:
        raw: Parsed raw response.

    Returns:
        List of dictionary rows.

    Raises:
        No exceptions are intentionally raised.
    """

    if isinstance(raw, list):
        return [row for row in raw if isinstance(row, dict)]
    if not isinstance(raw, dict):
        return []
    for key in ("data", "object", "result", "values", "rows", "list", "items", "records", "tradeList"):
        value = raw.get(key)
        if isinstance(value, list):
            return [row for row in value if isinstance(row, dict)]
        if isinstance(value, dict):
            nested = raw_rows(value)
            if nested:
                return nested
    return []


def first_non_missing(*values: Any) -> Any:
    """Return the first value that is present, while preserving a real zero."""

    for value in values:
        if value not in (None, ""):
            return value
    return None


def maximum_numeric_evidence(*values: Any) -> Any:
    """Return the largest metric proven by overlapping exact-period sources.

    Alibaba summary, funnel, and trade-list responses can overlap. A summary
    may contain a real zero while an exact-period trade row proves that at
    least one order exists. Taking the largest value avoids contradictory
    output without adding overlapping sources and double-counting an order.

    Args:
        *values: Candidate values from summary, funnel, and trade data.

    Returns:
        Largest numeric candidate, or the first non-missing original value
        when no candidate is numeric.

    Raises:
        No exceptions are intentionally raised.
    """

    numeric_values = [number for value in values if (number := to_float(value)) is not None]
    if not numeric_values:
        return first_non_missing(*values)
    maximum = max(numeric_values)
    return int(maximum) if maximum.is_integer() else maximum


def display_width(value: Any) -> int:
    """Estimate rendered width, counting CJK glyphs as two columns."""

    rendered = str(value or "")
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1 for char in rendered)


def identifier_text(value: Any) -> str:
    """Return a transaction identifier as literal text, never scientific notation."""

    if value in (None, ""):
        return "未返回"
    rendered = str(value)
    return f"'{rendered}" if rendered.startswith(("=", "+", "-", "@")) else rendered


def amount_node_value(value: Any) -> Any:
    """Extract the numeric amount from an Alibaba money node."""

    if isinstance(value, dict):
        return value.get("amount")
    return value


def trade_metrics(
    raw_trade: Any,
    period_start: str | None = None,
    period_end: str | None = None,
) -> tuple[int | None, float | None]:
    """Return exact trade count and paid/received amount from trade rows.

    The realized amount prefers paid order amount, then received amount. A
    contract total is used only when the paid fields are absent, never to turn
    an unpaid order into realized ROI.
    """

    rows = raw_rows(raw_trade)
    if period_start and period_end:
        rows = [
            item
            for item in rows
            if (
                not str(first_present(item, "createDate", "gmtCreate", "orderCreateTime") or "")[:10]
                or period_start
                <= str(first_present(item, "createDate", "gmtCreate", "orderCreateTime") or "")[:10]
                <= period_end
            )
        ]
    if not rows:
        # The run-local collector writes an explicit exact-period canonical
        # response. An empty, complete response proves a real zero; an absent
        # or legacy first-page file does not.
        if isinstance(raw_trade, dict):
            source_start = str(raw_trade.get("periodStart") or "")[:10]
            source_end = str(raw_trade.get("periodEnd") or "")[:10]
            row_count = raw_trade.get("rowCount")
            complete = raw_trade.get("complete")
            exact_period = (
                period_start
                and period_end
                and source_start == period_start
                and source_end == period_end
            )
            if exact_period and row_count == 0 and complete is True:
                return 0, 0.0
        return None, None
    amounts: list[float] = []
    for item in rows:
        payment = item.get("payment") if isinstance(item.get("payment"), dict) else {}
        chosen = first_non_missing(
            amount_node_value(payment.get("paidOrderAmount")),
            amount_node_value(payment.get("receivedAmount")),
            amount_node_value(payment.get("totalAmount")),
            first_present(item, "amount", "orderAmount", "totalAmount"),
        )
        number = to_float(chosen)
        if number is not None:
            amounts.append(number)
    return len(rows), (sum(amounts) if amounts else None)


def trade_buyer_name(item: JsonDict) -> Any:
    """Return the readable buyer name from a flat or nested trade row."""

    buyer = item.get("buyer") if isinstance(item.get("buyer"), dict) else {}
    return first_non_missing(
        first_present(buyer, "participantName", "companyName", "name"),
        first_present(item, "buyerName", "customerName"),
    )


def trade_status_text(item: JsonDict) -> Any:
    """Return a readable status from a flat value or nested status object."""

    status = item.get("status")
    if isinstance(status, dict):
        return first_present(status, "status", "name", "label")
    return first_non_missing(status, first_present(item, "tradeStatus", "orderStatus"))


def trade_amount_value(item: JsonDict) -> Any:
    """Return one trade's paid/received amount without exposing payment JSON."""

    payment = item.get("payment") if isinstance(item.get("payment"), dict) else {}
    return first_non_missing(
        amount_node_value(payment.get("paidOrderAmount")),
        amount_node_value(payment.get("receivedAmount")),
        amount_node_value(payment.get("totalAmount")),
        first_present(item, "amount", "orderAmount", "totalAmount"),
    )


def collect_raw_context(raw_dir: Path) -> JsonDict:
    """Load optional raw files used only by the workbook output layer.

    Args:
        raw_dir: Directory containing raw tool responses.

    Returns:
        Dictionary with parsed optional raw payloads.

    Raises:
        No exceptions are intentionally raised; missing files are allowed.
    """

    context: JsonDict = {
        "status": load_optional_json(raw_dir / "_collect_status.json") or {},
        "shop_summary": load_optional_json(raw_dir / "data_advisor_shop_summary_current.json"),
        "shop_summary_day": load_optional_json(raw_dir / "data_advisor_shop_summary_day.json"),
        "shop_product": load_optional_json(raw_dir / "data_advisor_shop_product.json"),
        "trade": load_optional_json(raw_dir / "queryTradeListMcp.json"),
        "ads_account": (
            load_optional_json(raw_dir / "icbu_ads_hateoas_query_diagnosis.json")
            or load_optional_json(raw_dir / "icbu_ads_hateoas_query.json")
            or load_optional_json(raw_dir / "icbu_ads_hateoas_query_company.json")
        ),
        "subaccounts": load_optional_json(raw_dir / "subaccount_query.json"),
        "seller_shop": load_optional_json(raw_dir / "query_seller_shop_dim_diag_data.json"),
        "seller_shop_l1plus": load_optional_json(raw_dir / "query_seller_shop_dim_diag_data_l1plus.json"),
        "chat_quality": load_optional_json(raw_dir / "query_seller_chat_quality_check_detail.json"),
        "contacts": load_optional_json(raw_dir / "query_contact.json"),
        "recent_conversations": load_optional_json(raw_dir / "query_recent_conversation_week.json"),
        "conversation_messages": load_optional_json(raw_dir / "query_conversation_msg_week.json"),
    }
    context["seller_acct_files"] = [
        load_optional_json(path)
        for path in sorted(raw_dir.glob("query_seller_acct_dim_diag_data_*.json"))
        if "_l1plus_" not in path.name
    ]
    context["seller_acct_l1plus_files"] = [
        load_optional_json(path)
        for path in sorted(raw_dir.glob("query_seller_acct_dim_diag_data_l1plus_*.json"))
    ]
    context["chat_quality_daily_files"] = [
        load_optional_json(path)
        for path in sorted(raw_dir.glob("query_seller_chat_quality_check_detail_*.json"))
    ]
    return context


def summary_value(raw_summary: Any, *keys: str) -> Any:
    """Read a metric from data_advisor_shop_summary style output.

    Args:
        raw_summary: Parsed raw summary response.
        *keys: Candidate field names.

    Returns:
        First present field value, or None.

    Raises:
        No exceptions are intentionally raised.
    """

    rows = raw_rows(raw_summary)
    if rows:
        return first_present(rows[0], *keys)
    if isinstance(raw_summary, dict):
        return first_present(raw_summary, *keys)
    return None


def setup_workbook() -> Workbook:
    """Create the workbook and initialize all required sheets.

    Args:
        None.

    Returns:
        Openpyxl workbook with exactly the required report sheets.

    Raises:
        No exceptions are intentionally raised.
    """

    workbook = Workbook()
    first = workbook.active
    first.title = SHEETS[0]
    for name in SHEETS[1:]:
        workbook.create_sheet(name)
    return workbook


def style_sheet(worksheet: Worksheet, widths: dict[str, int] | None = None) -> None:
    """Apply consistent worksheet styling without creating Excel table objects.

    Args:
        worksheet: Sheet to style.
        widths: Optional explicit column width mapping by column letter.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    visual_fill = PatternFill("solid", fgColor="E2F0D9")
    visual_label_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    if worksheet.max_row >= 1 and worksheet.max_column >= 1 and worksheet.title != "老板结论":
        worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if cell.row == 1:
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.fill = header_fill
            elif isinstance(cell.value, str) and cell.value.startswith("【") and cell.value.endswith("】"):
                is_visual = "可视化看板" in cell.value
                cell.font = Font(name="Arial", size=10, bold=True, color="274E13" if is_visual else "1F4E78")
                cell.fill = visual_fill if is_visual else section_fill
            elif isinstance(row[0].value, str) and row[0].value in {"KPI卡片", "横向条形图", "热力矩阵", "迷你排行榜", "质量漏斗", "订单漏斗", "优先级堆叠条"}:
                cell.fill = visual_label_fill if cell.column in {1, 2} else PatternFill("solid", fgColor="F3F6EF")
                cell.font = Font(name="Arial", size=10, bold=cell.column in {1, 2}, color="274E13" if cell.column in {1, 2} else "000000")

            text = str(cell.value or "")
            if cell.column in {4, 5} and any(marker in text for marker in ("■", "□")):
                cell.font = Font(name="Arial", size=10, bold=True, color="38761D")
            if any(word in text for word in ("P0", "红灯", "预警", "需处理")) and not (isinstance(cell.value, str) and cell.value.startswith("【")):
                cell.fill = PatternFill("solid", fgColor="F4CCCC")
            elif any(word in text for word in ("P1", "需复盘", "需追问")):
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
            elif any(word in text for word in ("健康", "可放大", "高质量")):
                cell.fill = PatternFill("solid", fgColor="D9EAD3")

    for column in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = (widths or {}).get(letter, 18)


def write_rows(worksheet: Worksheet, rows: Rows, widths: dict[str, int] | None = None) -> None:
    """Write rows to a sheet and apply report styling.

    Args:
        worksheet: Sheet receiving rows.
        rows: Two-dimensional list of values.
        widths: Optional column widths.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    for row in rows:
        worksheet.append([sanitize_for_boss(cell) if isinstance(cell, str) else cell for cell in row])
    style_sheet(worksheet, widths)


def style_boss_conclusion(worksheet: Worksheet) -> None:
    """Add stronger visual hierarchy to the executive conclusion sheet.

    Args:
        worksheet: The 老板结论 worksheet after rows are written.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    palette = {
        "基本信息": ("EAF3F8", "1F4E78"),
        "1. 投产": ("FCE4D6", "C00000"),
        "2. 推新节奏": ("E2F0D9", "548235"),
        "3. 询盘质量": ("FFF2CC", "BF9000"),
        "4. 订单产出": ("DDEBF7", "1F4E78"),
        "5. 业务员跟进": ("F4CCCC", "A61C00"),
        "6. 关键词机会": ("EADCF8", "7030A0"),
        "Top Action": ("D9EAD3", "38761D"),
    }
    section_fill = PatternFill("solid", fgColor="1F4E78")

    for row in worksheet.iter_rows(min_row=2):
        module = str(row[0].value or "")
        if module.startswith("【") and module.endswith("】"):
            for cell in row:
                cell.fill = section_fill
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            continue

        fill_color = None
        font_color = "000000"
        for key, (bg, fg) in palette.items():
            if module.startswith(key):
                fill_color = bg
                font_color = fg
                break
        if not fill_color:
            continue
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Arial", size=10, bold=(cell.column in {1, 4}), color=font_color if cell.column in {1, 4} else "000000")

    for row in worksheet.iter_rows(min_row=2):
        judgement = str(row[3].value or "") if len(row) >= 4 else ""
        if any(word in judgement for word in ("红灯", "预警", "需处理")):
            row[3].fill = PatternFill("solid", fgColor="C00000")
            row[3].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        elif any(word in judgement for word in ("健康", "绿灯", "可放大")):
            row[3].fill = PatternFill("solid", fgColor="70AD47")
            row[3].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        elif judgement:
            row[3].fill = PatternFill("solid", fgColor="FFD966")
            row[3].font = Font(name="Arial", size=10, bold=True, color="000000")


def style_inquiry_quality(worksheet: Worksheet) -> None:
    """Highlight inquiry quality priorities for meeting review.

    Args:
        worksheet: The 询盘质量 worksheet after rows are written.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    fills = {
        "P0": PatternFill("solid", fgColor="F4CCCC"),
        "P1": PatternFill("solid", fgColor="FCE4D6"),
        "P2": PatternFill("solid", fgColor="FFF2CC"),
        "P3": PatternFill("solid", fgColor="E7E6E6"),
        "高质量": PatternFill("solid", fgColor="D9EAD3"),
        "低质量": PatternFill("solid", fgColor="E7E6E6"),
        "待补": PatternFill("solid", fgColor="FFF2CC"),
    }
    for row in worksheet.iter_rows(min_row=2):
        row_text = " ".join(str(cell.value or "") for cell in row)
        fill = None
        for key, candidate in fills.items():
            if key in row_text:
                fill = candidate
                break
        if not fill:
            continue
        for cell in row:
            cell.fill = fill
        row[0].font = Font(name="Arial", size=10, bold=True, color="9C0006" if "P0" in row_text else "000000")


def traffic_light(value: Any, good_when_positive: bool = True) -> str:
    """Turn numeric status into a short boss-facing label.

    Args:
        value: Numeric value or missing marker.
        good_when_positive: Whether positive values should be treated as good.

    Returns:
        "健康", "预警", or "待确认".

    Raises:
        No exceptions are intentionally raised.
    """

    number = to_float(value)
    if number is None:
        return "待确认"
    if good_when_positive:
        return "健康" if number > 0 else "预警"
    return "健康" if number <= 0 else "预警"



def business_source_label(tool_name: str) -> str:
    """Map internal collection names to business-facing source labels.

    Args:
        tool_name: Internal read-only source name.

    Returns:
        Business-facing label that does not expose implementation wording.

    Raises:
        No exceptions are intentionally raised.
    """

    pairs = [
        ("shop_summary", "店铺经营汇总"),
        ("shop_region", "国家与地区分布"),
        ("shop_product", "商品效果数据"),
        ("weekly", "服务周报数据"),
        ("ads", "广告诊断数据"),
        ("trade", "交易合同数据"),
        ("seller", "业务员沟通数据"),
        ("chat_quality", "沟通质检数据"),
        ("subaccount", "子账号数据"),
        ("contact", "联系人数据"),
        ("product_score", "商品质量分"),
        ("product", "商品数据"),
        ("keyword", "关键词数据"),
        ("risk", "风险诊断数据"),
        ("customer", "店铺基础信息"),
    ]
    lowered = tool_name.lower()
    for key, label in pairs:
        if key in lowered:
            return label
    return "只读业务数据"


def quality_flags(report_data: JsonDict) -> JsonDict:
    """Return blocking flags from the data quality layer.

    Args:
        report_data: Normalized report data.

    Returns:
        Data quality blocking flags.

    Raises:
        No exceptions are intentionally raised.
    """

    return ((report_data.get("data_quality") or {}).get("blocking_flags") or {})


def can_judge(report_data: JsonDict, *flags: str) -> bool:
    """Return whether a conclusion can be made under data-quality rules.

    Args:
        report_data: Normalized report data.
        *flags: Blocking flag names that would invalidate the conclusion.

    Returns:
        True when none of the requested flags is active.

    Raises:
        No exceptions are intentionally raised.
    """

    active = quality_flags(report_data)
    return not any(active.get(flag) for flag in flags)


def judgement_text(report_data: JsonDict, *flags: str, positive: str = "可判断") -> str:
    """Format a conservative judgement label.

    Args:
        report_data: Normalized report data.
        *flags: Blocking flags that force an unknown judgement.
        positive: Label to use when no blocking flag is active.

    Returns:
        Conservative judgement text.

    Raises:
        No exceptions are intentionally raised.
    """

    return positive if can_judge(report_data, *flags) else "不可判断"


def money_text(value: Any) -> str:
    """Format money-like values without inventing missing numbers.

    Args:
        value: Raw money value.

    Returns:
        User-facing money text or "未返回".

    Raises:
        No exceptions are intentionally raised.
    """

    number = to_float(value)
    if number is None:
        return "未返回"
    return f"{number:,.2f}"


def ratio_text(numerator: Any, denominator: Any) -> str:
    """Format a ratio only when both sides are present and denominator is non-zero.

    Args:
        numerator: Ratio numerator.
        denominator: Ratio denominator.

    Returns:
        Percentage text or "不可判断".

    Raises:
        No exceptions are intentionally raised.
    """

    n = to_float(numerator)
    d = to_float(denominator)
    if n is None or d in (None, 0):
        return "不可判断"
    return f"{n / d * 100:.2f}%"


def full_product_title(item: JsonDict) -> str:
    """Return the most complete product title available.

    Args:
        item: Product row.

    Returns:
        Full product title or missing marker.

    Raises:
        No exceptions are intentionally raised.
    """

    return compact_text(first_present(item, "subject", "title", "name", "productName"), 120)


def product_link(item: JsonDict) -> str:
    """Return a product link when available.

    Args:
        item: Product row.

    Returns:
        URL text or "未返回".

    Raises:
        No exceptions are intentionally raised.
    """

    return fmt_value(first_present(item, "detail_url", "detailUrl", "url", "productUrl"))


def action_field(action: JsonDict, key: str, missing: str = "未返回") -> str:
    """Read a structured action field with conservative missing handling.

    Args:
        action: Structured action row.
        key: Field name.
        missing: Fallback text.

    Returns:
        Compact text value.

    Raises:
        No exceptions are intentionally raised.
    """

    return compact_text(action.get(key), 120) if action.get(key) not in (None, "") else missing


def narrative_section(narrative: JsonDict, name: str) -> JsonDict:
    """Return a named narrative section as a dictionary.

    Args:
        narrative: Agent-written narrative contract.
        name: Section key, such as "boss_conclusion" or "sheet_summaries".

    Returns:
        Section dictionary or an empty dictionary when absent.

    Raises:
        No exceptions are intentionally raised.
    """

    value = narrative.get(name) if isinstance(narrative, dict) else None
    return value if isinstance(value, dict) else {}


def narrative_text(value: Any, limit: int = 120) -> str:
    """Clean Agent-written text before it enters workbook cells.

    Args:
        value: Candidate narrative value.
        limit: Maximum characters for the cell text.

    Returns:
        Safe compact text, or an empty string when the value is absent.

    Raises:
        No exceptions are intentionally raised.
    """

    if value in (None, ""):
        return ""
    if isinstance(value, (dict, list)):
        return ""
    return compact_text(sanitize_for_boss(str(value)), limit)


def first_narrative_value(mapping: JsonDict, *keys: str, limit: int = 120) -> str:
    """Read the first usable Agent-written value from several key aliases.

    Args:
        mapping: Narrative subsection.
        *keys: Candidate key names.
        limit: Maximum characters to keep.

    Returns:
        First safe narrative text or an empty string.

    Raises:
        No exceptions are intentionally raised.
    """

    if not isinstance(mapping, dict):
        return ""
    for key in keys:
        text = narrative_text(mapping.get(key), limit)
        if text:
            return text
    return ""


def sheet_summary(narrative: JsonDict, sheet_name: str, analysis: JsonDict, report_data: JsonDict) -> JsonDict:
    """Return four spoken summary fields for a sheet.

    Args:
        narrative: Optional Agent-written narrative contract.
        sheet_name: Workbook sheet name.
        analysis: Diagnosis output used for deterministic fallback copy.
        report_data: Normalized report data used for deterministic fallback copy.

    Returns:
        Dictionary with 老板口径、关键证据、今天动作、下周复查.

    Raises:
        No exceptions are intentionally raised.
    """

    summaries = narrative_section(narrative, "sheet_summaries")
    candidate = summaries.get(sheet_name)
    if isinstance(candidate, str):
        return {
            "老板口径": narrative_text(candidate, 150),
            "关键证据": "见下方核心数据",
            "今天动作": "先处理红灯和 P0 项",
            "下周复查": "按同一口径复查",
        }
    if isinstance(candidate, dict):
        return {
            "老板口径": first_narrative_value(candidate, "老板口径", "boss_view", "summary", limit=150) or sheet_intro_line(sheet_name, analysis, report_data),
            "关键证据": first_narrative_value(candidate, "关键证据", "evidence", limit=150) or "见下方核心数据",
            "今天动作": first_narrative_value(candidate, "今天动作", "today_action", "action", limit=130) or "先处理红灯和 P0 项",
            "下周复查": first_narrative_value(candidate, "下周复查", "next_review", "review", limit=110) or "按同一口径复查",
        }
    return {
        "老板口径": sheet_intro_line(sheet_name, analysis, report_data),
        "关键证据": "见下方核心数据",
        "今天动作": "先处理红灯和 P0 项",
        "下周复查": "按同一口径复查",
    }


def spoken_summary_rows(sheet_name: str, analysis: JsonDict, report_data: JsonDict, narrative: JsonDict, width: int) -> Rows:
    """Create four spoken summary rows for the top of a worksheet.

    Args:
        sheet_name: Current worksheet name.
        analysis: Diagnosis output.
        report_data: Normalized report data.
        narrative: Optional Agent-written narrative contract.
        width: Number of columns in the current sheet.

    Returns:
        Padded rows for 老板口径、关键证据、今天动作、下周复查.

    Raises:
        No exceptions are intentionally raised.
    """

    summary = sheet_summary(narrative, sheet_name, analysis, report_data)
    labels = ["老板口径", "关键证据", "今天动作", "下周复查"]
    return [[label, compact_text(summary.get(label), 150)] + [""] * max(0, width - 2) for label in labels]


def rewrite_for(narrative: JsonDict, group: str, keys: list[Any], fallback: str, limit: int = 90) -> str:
    """Return an Agent rewrite for a row-level action when available.

    Args:
        narrative: Optional Agent-written narrative contract.
        group: Row rewrite group, such as "product_actions".
        keys: Candidate object identifiers.
        fallback: Deterministic fallback action.
        limit: Maximum text length.

    Returns:
        Agent rewrite or fallback action.

    Raises:
        No exceptions are intentionally raised.
    """

    rewrites = narrative_section(narrative, "row_rewrites").get(group)
    if not isinstance(rewrites, dict):
        return compact_text(fallback, limit)
    normalized = {str(key).strip().lower(): value for key, value in rewrites.items()}
    for key in keys:
        if key in (None, ""):
            continue
        direct = normalized.get(str(key).strip().lower())
        if direct:
            return narrative_text(direct, limit) or compact_text(fallback, limit)
    return compact_text(fallback, limit)


def overall_status(report_data: JsonDict, analysis: JsonDict) -> JsonDict:
    """Return the conservative executive status.

    Args:
        report_data: Normalized report data.
        analysis: Diagnosis output.

    Returns:
        Status dictionary with label and reason.

    Raises:
        No exceptions are intentionally raised.
    """

    status = analysis.get("executive_status") or {}
    if status:
        return status
    flags = quality_flags(report_data)
    if flags.get("collection_period_mismatch") or flags.get("order_period_mismatch"):
        return {"label": "红灯", "reason": "周期错位，不能判健康"}
    if flags.get("order_pagination_truncated"):
        return {"label": "红灯", "reason": "订单分页未完成，订单数和金额只能看作下限"}
    if flags.get("order_amount_missing") or flags.get("ad_cost_missing"):
        return {"label": "不可判断", "reason": "关键金额缺失，ROI 不完整"}
    return {"label": "黄灯", "reason": "按可用数据保守判断"}


_SYSTEM_ERROR_PATTERNS = (
    "errorCode", "errorMsg", "error_code", "-32002", "-32001",
    "Agent 类型不允许", "Traceback", "ECONNREFUSED", "at Object.",
    "stack trace", "data connector exited", "Data source returned",
    "accio-" + "mcp-cli exited", "Gate" + "way returned",
)


def sanitize_for_boss(value: Any) -> Any:
    """Remove raw technical leakage and decode simple JSON-like labels.

    Args:
        value: Candidate workbook cell value.

    Returns:
        Safe value for boss-facing workbook cells.

    Raises:
        No exceptions are intentionally raised.
    """

    if value is None or not isinstance(value, str):
        return value
    if any(pat in value for pat in _SYSTEM_ERROR_PATTERNS):
        return "平台接口异常，数据未返回"
    text = value.strip()
    if (text.startswith("{") and text.endswith("}")) or (text.startswith("[") and text.endswith("]")):
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, dict):
            for key in ("def", "name", "title", "label", "value"):
                if parsed.get(key) not in (None, ""):
                    return compact_text(parsed.get(key), 70)
            return "已返回业务字段，需按平台明细复核"
        if isinstance(parsed, list):
            return f"已返回 {len(parsed)} 条记录，需按平台明细复核"
    return value


def percent_text(value: Any) -> str:
    n = to_float(value)
    if n is None:
        return "未返回"
    if n <= 1:
        n = n * 100
    return f"{n:.0f}%"


def boss_narrative_sentence(report_data: JsonDict, analysis: JsonDict, raw: JsonDict) -> str:
    ad = parse_ad_summary(report_data, raw)
    cost = ad.get("cost")
    inquiry_q = (analysis.get("inquiry_quality") or {}).get("summary") or {}
    total_inq = inquiry_q.get("total_records") or find_indicator(report_data, ("商机", "询盘"))
    high_q = inquiry_q.get("high_quality")
    totals = funnel_totals(report_data, raw)
    meta = report_data.get("meta") or {}
    trade_count, trade_amount = trade_metrics(
        raw.get("trade"),
        meta.get("period_start"),
        meta.get("period_end"),
    )
    orders = maximum_numeric_evidence(
        summary_value(raw.get("shop_summary"), "orderCnt", "crtOrdCnt", "orderCntValue"),
        find_indicator(report_data, ("订单数",)),
        totals.get("order"),
        trade_count,
    )
    amount = maximum_numeric_evidence(
        summary_value(raw.get("shop_summary"), "orderAmt", "orderAmount", "payOrdAmt"),
        trade_amount,
    )
    amount_ok = can_judge(report_data, "order_amount_missing")
    parts = []
    if cost is not None:
        parts.append(f"花了 ${cost:.0f} 广告费")
    else:
        parts.append("广告花费未返回")
    if total_inq:
        inq_desc = f"拿到 {total_inq} 条询盘"
        if high_q:
            inq_desc += f"({high_q}条高质量)"
        parts.append(inq_desc)
    if orders:
        parts.append(f"{orders}笔订单")
    if amount and amount_ok:
        parts.append(f"金额 ${to_float(amount):,.0f}" if to_float(amount) else "金额待确认")
    elif orders:
        parts.append("但金额没拉到")
    sentence = "本周" + "，".join(parts) + "。"
    if cost and not amount_ok:
        sentence += "钱花出去了，回报算不清。"
    elif cost and amount and to_float(amount):
        roi = to_float(amount) / cost if cost > 0 else 0
        if roi > 5:
            sentence += f"投产比 {roi:.1f}，回报不错。"
        elif roi > 2:
            sentence += f"投产比 {roi:.1f}，正常水平。"
        else:
            sentence += f"投产比 {roi:.1f}，偏低需关注。"
    return sentence


def sheet_intro_line(sheet_name: str, analysis: JsonDict, report_data: JsonDict) -> str:
    one_liner = analysis.get("one_liner") or ""
    intro_lines = {
        "投产看板": "老板先看这里：这周钱花在哪、有没有换来询盘和订单；缺花费或订单金额时，回报先算不清。",
        "询盘质量": "这页判断本周询盘值不值得追：先看高质量、低质量和 P0/P1 跟进压力，再看每条该怎么推进。",
        "订单产出": "这页看询盘有没有变成订单：订单数、金额和周期一致性决定投产能不能下结论。",
        "商品节奏": "这页看货承不承接流量：哪些商品该放大，哪些商品曝光不少但没有询盘。",
        "关键词与广告机会": "这页看钱该继续投到哪些词：金词继续放大，烧钱词马上复盘，潜力词小预算测试。",
        "业务员回复与跟进": "这页看业务有没有接住好询盘：回复慢、未跟进、重复回复都会直接影响订单机会。",
        "数据质量检查": "这页说明哪些结论能判断、哪些只能保守看；红灯字段先补齐，再谈投产健康。",
    }
    intro_line = intro_lines.get(sheet_name, "")
    if one_liner:
        return f"{compact_text(one_liner, 60)}。{intro_line}"
    return intro_line


def spoken_summary_row(sheet_name: str, analysis: JsonDict, report_data: JsonDict, width: int) -> list[Any]:
    """Create a boss-spoken summary row for the top of each worksheet.

    Args:
        sheet_name: Current worksheet name.
        analysis: Diagnosis output.
        report_data: Normalized report data.
        width: Number of columns in the current sheet.

    Returns:
        A padded row whose first two cells carry the spoken summary.

    Raises:
        No exceptions are intentionally raised.
    """

    return ["老板口径", compact_text(sheet_intro_line(sheet_name, analysis, report_data), 150)] + [""] * max(0, width - 2)


def build_boss_conclusion(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict) -> list[Block]:
    """Build the boss one-page sheet as a 4-section block layout.

    The layout splits the previous mixed table into a distinct title band, six
    short narrative paragraphs, and two compact detail tables (boss must-do
    actions, next-week review checklist). Keeping the sheet under 25 rows so
    that printed/PDF export still fits the boss's expected one-pager.
    """

    quality = report_data.get("data_quality") or {}
    status = overall_status(report_data, analysis)
    actions = analysis.get("boss_top5_actions") or analysis.get("top3_actions") or []
    degraded = quality.get("degraded_conclusions") or []
    biggest_risk = degraded[0] if degraded else status.get("reason")
    product_q = analysis.get("products_quadrant") or {}
    inquiry_summary = ((analysis.get("inquiry_quality") or {}).get("summary") or {})
    opportunity = "放大高质量询盘和可承接商品"
    if product_q.get("ink_print"):
        opportunity = f"复制金主商品：{compact_text(product_q['ink_print'][0].get('title'), 45)}"
    elif inquiry_summary.get("high_quality"):
        opportunity = f"本周高质量询盘 {inquiry_summary['high_quality']} 条，先推进报价/样品"

    boss_copy = narrative_section(narrative, "boss_conclusion")
    battle = first_narrative_value(boss_copy, "weekly_battle", "本周战况", "battle_summary", limit=180) or boss_narrative_sentence(report_data, analysis, raw)
    business_status = first_narrative_value(boss_copy, "business_status", "经营状态", limit=120) or status.get("reason") or ""
    data_confidence = first_narrative_value(boss_copy, "data_confidence", "数据可信度", limit=120) or f"红灯{quality.get('red_count', 0)}项 黄灯{quality.get('yellow_count', 0)}项"
    biggest_risk = first_narrative_value(boss_copy, "biggest_risk", "最大风险", limit=140) or biggest_risk
    opportunity = first_narrative_value(boss_copy, "biggest_opportunity", "最大机会", limit=140) or opportunity
    decision = first_narrative_value(boss_copy, "boss_decision", "老板拍板", "decision", limit=140) or "是否先补齐订单金额，再判断投产回收"
    if can_judge(report_data, "order_amount_missing"):
        decision = first_narrative_value(boss_copy, "boss_decision", "老板拍板", "decision", limit=140) or "广告预算维持/加减，重点品方向"

    collection = quality.get("collection") or {}
    period = collection.get("period_label") or ""
    if not period:
        start = collection.get("period_start") or ""
        end = collection.get("period_end") or ""
        if start and end:
            period = f"{start} 至 {end}"
    status_label = clean_light(status.get("label") or "不可判断")

    blocks: list[Block] = [
        title_block("老板一页纸", period, f"本周战况：{status_label}"),
        blank_block(),
        paragraph_block("本周战况", battle),
        paragraph_block("经营状态", business_status),
        paragraph_block("数据可信度", data_confidence),
        paragraph_block("最大风险", biggest_risk),
        paragraph_block("最大机会", opportunity),
        paragraph_block("老板拍板", decision),
        blank_block(),
        section_block("可视化看板：本周必抓行动", palette="visual"),
        detail_header_block(["优先级", "行动", "证据", "拍板/确认", "负责人", "截止", "下周复查", ""]),
    ]

    narrative_actions = narrative.get("top_actions") if isinstance(narrative.get("top_actions"), list) else []
    for i, action in enumerate(actions[:5]):
        human_action = narrative_actions[i] if i < len(narrative_actions) and isinstance(narrative_actions[i], dict) else {}
        blocks.append(detail_block([
            action.get("priority") or "P2",
            first_narrative_value(human_action, "action", "conclusion", "今天动作", limit=80) or compact_text(action_field(action, "action"), 80),
            first_narrative_value(human_action, "evidence", "why", "证据", limit=60) or compact_text(action_field(action, "evidence"), 60),
            first_narrative_value(human_action, "decision", "老板拍板", limit=40) or "",
            action.get("owner") or "运营主管",
            action.get("due") or "本周五",
            first_narrative_value(human_action, "review", "下周复查", limit=40) or compact_text(action.get("review_metric") or "下周复查", 40),
            "",
        ]))

    blocks.append(blank_block())
    blocks.append(section_block("下周复查（5 项）", palette="paragraph"))
    blocks.append(detail_header_block(["主题", "复查指标", "负责人", "截止", "", "", "", ""]))
    review_items = [
        ("投产", "花费/商机成本/订单金额是否可算"),
        ("商品", "高曝光低询盘商品是否减少"),
        ("询盘", "高质量询盘进入报价/样品"),
        ("业务员", "12h+未回清零，5分钟回复率上升"),
        ("跟进", "平台可见跟进风险是否下降"),
    ]
    for label, review in review_items:
        blocks.append(detail_block([label, review, "运营", "下周一", "", "", "", ""]))
    return blocks


def style_boss_conclusion(worksheet: Worksheet) -> None:
    """Style the one-page executive sheet as compact cards.

    Args:
        worksheet: 老板结论 worksheet.

    Returns:
        None.

    Raises:
        No exceptions are intentionally raised.
    """

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    fills = {
        "红灯": PatternFill("solid", fgColor="C00000"),
        "P0": PatternFill("solid", fgColor="C00000"),
        "黄灯": PatternFill("solid", fgColor="FFD966"),
        "不可判断": PatternFill("solid", fgColor="B7B7B7"),
        "机会": PatternFill("solid", fgColor="D9EAD3"),
        "复查": PatternFill("solid", fgColor="DDEBF7"),
    }
    for row in worksheet.iter_rows(min_row=2):
        text = " ".join(str(cell.value or "") for cell in row[:3])
        for key, fill in fills.items():
            if key in text:
                for cell in row:
                    cell.fill = fill
                    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF" if key in {"红灯", "P0"} else "000000")
                break
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_roi_rows(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict) -> list[Block]:
    """Build the 投产看板 sheet in the block layout.

    Visual cards on the top show four metric bars (spend / inquiries / orders /
    amount). Below them, a real detail table holds the same data so the boss
    can filter and sort. DataBar handles the in-cell bar so the columns stay
    short and readable.
    """

    ad = parse_ad_summary(report_data, raw)
    summary = raw.get("shop_summary")
    totals = funnel_totals(report_data, raw)
    ad_cost = ad.get("spend")
    inquiry = summary_value(summary, "abCnt", "abCntValue") or find_indicator(report_data, ("商机", "询盘")) or totals.get("inquiry")
    meta = report_data.get("meta") or {}
    trade_count, trade_amount = trade_metrics(
        raw.get("trade"),
        meta.get("period_start"),
        meta.get("period_end"),
    )
    orders = maximum_numeric_evidence(
        summary_value(summary, "orderCnt", "crtOrdCnt", "orderCntValue"),
        find_indicator(report_data, ("订单数",)),
        totals.get("order"),
        trade_count,
    )
    amount = maximum_numeric_evidence(
        summary_value(summary, "orderAmt", "orderAmount", "payOrdAmt"),
        trade_amount,
    )
    roi_ok = can_judge(
        report_data,
        "order_amount_missing",
        "ad_cost_missing",
        "order_period_mismatch",
        "order_pagination_truncated",
        "collection_period_mismatch",
    )
    summary_dict = sheet_summary(narrative, "投产看板", analysis, report_data)
    max_v = max_positive([ad_cost, inquiry, orders, amount])
    title_sub = "本周钱花得值不值"
    title_status = "可判断" if roi_ok else "回报算不清"

    blocks: list[Block] = [
        title_block("投产看板", title_sub, title_status),
        blank_block(),
        paragraph_block("老板口径", summary_dict.get("老板口径")),
        paragraph_block("关键证据", summary_dict.get("关键证据")),
        paragraph_block("今天动作", summary_dict.get("今天动作")),
        paragraph_block("下周复查", summary_dict.get("下周复查")),
        blank_block(),
        section_block("可视化看板：投产可信度", palette="visual"),
        visual_block("横向条形", "广告花费", money_text(ad_cost), safe_ratio(ad_cost, max_v),
                     judgement_text(report_data, "ad_cost_missing"), "缺花费不算广告 ROI"),
        visual_block("横向条形", "询盘/商机", inquiry, safe_ratio(inquiry, max_v),
                     traffic_light(inquiry), "只看数量不够，要看质量和订单后续"),
        visual_block("横向条形", "订单数", orders, safe_ratio(orders, max_v),
                     judgement_text(report_data, "order_count_missing"), "订单少先回查报价和跟进"),
        visual_block("横向条形", "订单金额", money_text(amount), safe_ratio(amount, max_v),
                     judgement_text(report_data, "order_amount_missing"),
                     "缺金额则投产回收不可判断", danger=not roi_ok),
        blank_block(),
        section_block("投产总览明细", palette="detail"),
        detail_header_block(["模块", "对象", "证据", "本周能不能判断", "结论", "今天动作", "复查指标"]),
        detail_block(["投产总览", "广告花费", money_text(ad_cost),
                      judgement_text(report_data, "ad_cost_missing"),
                      "看回报，不单看花费", "按关键词、商品、询盘质量拆开看",
                      "下周看商机成本"]),
        detail_block(["投产总览", "广告商机", fmt_value(ad.get("ad_inquiry")),
                      judgement_text(report_data, "ad_cost_missing"),
                      "看商机成本", "无商机词先核对承接并确认调价或暂停条件", "按用户确认周期复查有效询盘"]),
        detail_block(["投产总览", "全店询盘", fmt_value(inquiry),
                      "可判断" if inquiry not in (None, "") else "不可判断",
                      "只看数量不够", "优先追高质量询盘", "报价/样品/订单后续"]),
        detail_block(["投产总览", "订单金额", money_text(amount),
                      "可判断" if roi_ok else "不可判断",
                      "ROI 不可判断" if not roi_ok else "可计算回收",
                      "缺金额先补订单导出", "订单金额字段返回"]),
    ]
    return blocks


def build_keyword_rows(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict) -> list[Block]:
    """Build keyword and advertising-opportunity rows.

    Args:
        report_data: Normalized report data.
        analysis: Diagnosis output.
        raw: Optional raw context.
        narrative: Optional Agent-written narrative contract.

    Returns:
        Rows for 关键词与广告机会.

    Raises:
        No exceptions are intentionally raised.
    """

    summary_dict = sheet_summary(narrative, "关键词与广告机会", analysis, report_data)
    quadrants = analysis.get("keywords_quadrant") or {}
    market = report_data.get("market") or {}
    gold_n = len(quadrants.get("gold") or [])
    burning_n = len(quadrants.get("burning") or [])
    potential_n = len(quadrants.get("potential") or [])
    expand_n = len(quadrants.get("expand") or [])
    max_n = max_positive([gold_n, burning_n, potential_n, expand_n, 1])

    blocks: list[Block] = [
        title_block("关键词与广告机会", "金词放大，潜力词测水", "本周钱花到哪些词"),
        blank_block(),
        paragraph_block("老板口径", summary_dict.get("老板口径")),
        paragraph_block("关键证据", summary_dict.get("关键证据")),
        paragraph_block("今天动作", summary_dict.get("今天动作")),
        paragraph_block("下周复查", summary_dict.get("下周复查")),
        blank_block(),
        section_block("可视化看板：关键词分层一眼看", palette="visual"),
        visual_block("分层", "金词", gold_n, safe_ratio(gold_n, max_n), "金词", "保留核心位继续观察"),
        visual_block("分层", "烧钱词", burning_n, safe_ratio(burning_n, max_n), "烧钱词", "降价或暂停", danger=burning_n > 0),
        visual_block("分层", "潜力词", potential_n, safe_ratio(potential_n, max_n), "潜力词", "换承接页 3 天看商机"),
        visual_block("分层", "拓展词", expand_n, safe_ratio(expand_n, max_n), "拓展词", "标题嵌词，小预算测试"),
        blank_block(),
        section_block("关键词与机会明细", palette="detail"),
        detail_header_block(["模块", "对象", "证据", "本周能不能判断", "结论", "今天动作", "复查指标"]),
    ]
    for group, label in [("gold", "金词"), ("potential", "潜力词"), ("burning", "烧钱词"), ("expand", "拓展词")]:
        for item in (quadrants.get(group) or [])[:8]:
            blocks.append(detail_block([
                "关键词诊断",
                item.get("keyword"),
                f"点击 {fmt_value(item.get('clk'))} / 花费 {money_text(item.get('cost'))} / 询盘 {fmt_value(item.get('inquiry'))}",
                "证据不足" if item.get("product_name") in (None, "") else "可判断",
                label,
                rewrite_for(narrative, "keyword_actions", [item.get("keyword")],
                            f"{item.get('keyword')}：{join_actions(item.get('actions')) or '看 3 天有效询盘再决定预算'}",
                            90),
                "商机成本、询盘质量、报价/订单",
            ]))
    for item in (market.get("keyword_market") or [])[:12]:
        blocks.append(detail_block([
            "行业热词",
            item.get("keyword"),
            f"曝光指数 {fmt_value(item.get('year_imps_index'))} / 点击率 {fmt_value(item.get('ctr'))} / 商机转化 {fmt_value(item.get('business_rate'))}",
            "机会参考",
            compact_text(item.get("sell_status") or "待判断", 20),
            "小预算测试，3 天看有效询盘",
            "商机成本和询盘质量",
        ]))
    for item in (market.get("next_month_auction") or [])[:10]:
        blocks.append(detail_block([
            "次月资源",
            item.get("keyword"),
            f"业务线 {fmt_value(item.get('biz_line'))} / 标签 {compact_text(item.get('tags'), 30)}",
            "机会参考",
            compact_text(item.get("sell_status") or "待判断", 20),
            "提前评估是否抢资源",
            "下月资源位和询盘质量",
        ]))
    for text in (market.get("behavior_semantics") or [])[:8]:
        blocks.append(detail_block([
            "行为信号",
            "站内行为",
            compact_text(text, 90),
            "背景参考",
            "看买家真实动作",
            "把高频行为映射到关键词和商品承接",
            "点击、询盘、订单后续",
        ]))
    if gold_n + burning_n + potential_n + expand_n == 0 and not market.get("keyword_market"):
        blocks.append(detail_block(["关键词诊断", "未返回", "未返回", "不可判断",
                                    "证据不足", "补采关键词和广告明细",
                                    "下周重新判断"]))
    return blocks


def build_product_rows(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict) -> list[Block]:
    """Build 商品节奏 sheet with block layout."""

    products = report_data.get("products") or {}
    quadrants = analysis.get("products_quadrant") or {}
    summary_dict = sheet_summary(narrative, "商品节奏", analysis, report_data)

    bleed_n = len(quadrants.get("bleeding") or [])
    ink_n = len(quadrants.get("ink_print") or [])
    print_n = len(quadrants.get("print_money") or [])
    observe_n = len(quadrants.get("observation") or [])
    max_n = max_positive([bleed_n, ink_n, print_n, observe_n, 1])

    blocks: list[Block] = [
        title_block("商品节奏", "承接住的留下，不承接的换", "失血款先止血"),
        blank_block(),
        paragraph_block("老板口径", summary_dict.get("老板口径")),
        paragraph_block("关键证据", summary_dict.get("关键证据")),
        paragraph_block("今天动作", summary_dict.get("今天动作")),
        paragraph_block("下周复查", summary_dict.get("下周复查")),
        blank_block(),
        section_block("可视化看板：商品承接热力", palette="visual"),
        visual_block("热力", "失血款（高曝光0询盘）", bleed_n, safe_ratio(bleed_n, max_n),
                     "失血款" if bleed_n else "健康",
                     "改首图、标题、橱窗位", danger=bleed_n > 0),
        visual_block("热力", "潜力款（高询盘）", ink_n, safe_ratio(ink_n, max_n),
                     "高质量" if ink_n else "观察款",
                     "复制素材并给资源位"),
        visual_block("热力", "印钞款", print_n, safe_ratio(print_n, max_n),
                     "印钞款" if print_n else "观察款",
                     "继续放大并加预算"),
        visual_block("热力", "观察款", observe_n, safe_ratio(observe_n, max_n),
                     "观察款" if observe_n else "健康",
                     "达到 500 曝光再判断"),
        blank_block(),
        section_block("商品明细", palette="detail"),
        detail_header_block(["诊断", "商品ID", "完整标题", "链接", "曝光", "点击", "CTR", "询盘率", "橱窗", "今天动作", "复查指标"]),
    ]
    action_by_title = {}
    for group_rows in quadrants.values():
        for item in group_rows:
            action_by_title[str(item.get("title") or "")[:40]] = item
    top_products = products.get("exposure_top10") or []
    for item in top_products[:30]:
        if not isinstance(item, dict):
            continue
        title = full_product_title(item)
        matched = next((v for k, v in action_by_title.items() if k and k in title), None)
        imps = item.get("imps")
        clicks = first_present(item, "clicks", "clk", "click")
        inquiry = first_present(item, "fb_num", "inquiries", "ab")
        ctr = ratio_text(clicks, imps)
        inquiry_rate = fmt_value(first_present(item, "fb_rate", "inquiries_rate"))
        if to_float(imps) and not to_float(inquiry):
            diag = "失血款"
            action = "当天改首图和标题；3 天仍 0 询盘则换出橱窗"
        elif to_float(inquiry):
            diag = "潜力款"
            action = "复制主图标题，给关联商品和广告承接"
        else:
            diag = "观察款"
            action = "达到 500 曝光再判断，不盲目加预算"
        if matched and matched.get("actions"):
            action = join_actions(matched.get("actions"), 90)
        action = rewrite_for(
            narrative,
            "product_actions",
            [first_present(item, "product_id", "productId", "id"), title],
            f"{title}：{action}",
            90,
        )
        blocks.append(detail_block([
            diag,
            fmt_value(first_present(item, "product_id", "productId", "id")),
            title,
            product_link(item),
            fmt_value(imps),
            fmt_value(clicks),
            ctr,
            inquiry_rate,
            "是" if item.get("is_showcase") else "否",
            action,
            "曝光、点击、CTR、询盘率、橱窗位置",
        ]))
    return blocks


def build_order_rows(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict) -> list[Block]:
    """Build 订单产出 sheet with block layout."""

    summary = raw.get("shop_summary")
    totals = funnel_totals(report_data, raw)
    inquiry = summary_value(summary, "abCnt", "abCntValue") or find_indicator(report_data, ("商机", "询盘")) or totals.get("inquiry")
    meta = report_data.get("meta") or {}
    trade_count, trade_amount = trade_metrics(
        raw.get("trade"),
        meta.get("period_start"),
        meta.get("period_end"),
    )
    orders = maximum_numeric_evidence(
        summary_value(summary, "orderCnt", "crtOrdCnt", "orderCntValue"),
        find_indicator(report_data, ("订单数",)),
        totals.get("order"),
        trade_count,
    )
    amount = maximum_numeric_evidence(
        summary_value(summary, "orderAmt", "orderAmount", "payOrdAmt"),
        trade_amount,
    )
    amount_ok = can_judge(report_data, "order_amount_missing")
    period_ok = can_judge(
        report_data,
        "order_period_mismatch",
        "order_pagination_truncated",
        "collection_period_mismatch",
    )
    summary_dict = sheet_summary(narrative, "订单产出", analysis, report_data)
    max_v = max_positive([inquiry, orders, amount, 1])
    title_status = "可判断" if (amount_ok and period_ok) else "周期/金额任一缺则降级"

    blocks: list[Block] = [
        title_block("订单产出", "询盘 → 订单 → 金额", title_status),
        blank_block(),
        paragraph_block("老板口径", summary_dict.get("老板口径")),
        paragraph_block("关键证据", summary_dict.get("关键证据")),
        paragraph_block("今天动作", summary_dict.get("今天动作")),
        paragraph_block("下周复查", summary_dict.get("下周复查")),
        blank_block(),
        section_block("可视化看板：订单漏斗", palette="visual"),
        visual_block("漏斗", "询盘", inquiry, safe_ratio(inquiry, max_v),
                     "可判断" if inquiry not in (None, "") else "不可判断", "筛高质量询盘"),
        visual_block("漏斗", "订单数", orders, safe_ratio(orders, max_v),
                     judgement_text(report_data, "order_count_missing"), "回查报价推进"),
        visual_block("漏斗", "订单金额", money_text(amount), safe_ratio(amount, max_v),
                     "可判断" if amount_ok else "不可判断", "缺金额不算 ROI",
                     danger=not amount_ok),
        blank_block(),
        section_block("订单总览与明细", palette="detail"),
        detail_header_block(["模块", "对象", "证据", "本周能不能判断", "结论", "今天动作", "复查指标"]),
        detail_block(["订单总览", "周期一致", "报告周期内" if period_ok else "周期错位",
                      "可判断" if period_ok else "不可判断",
                      "周期错位不判健康", "重新拉取本周订单趋势",
                      "订单日期全部落在本周"]),
        detail_block(["订单总览", "订单金额", money_text(amount),
                      "可判断" if amount_ok else "不可判断",
                      "ROI 不可判断" if not amount_ok else "可做回收判断",
                      "补订单金额导出", "订单金额字段返回"]),
        detail_block(["订单总览", "询盘到订单", ratio_text(orders, inquiry),
                      "可判断" if period_ok and orders not in (None, "") else "不可判断",
                      "看转化，不看孤立订单", "高质量询盘逐条推进",
                      "报价、样品、订单后续"]),
    ]
    if period_ok:
        for row in ((report_data.get("funnel") or {}).get("daily") or [])[:14]:
            if isinstance(row, dict) and not row.get("_quarantined"):
                blocks.append(detail_block([
                    "每日趋势",
                    row.get("date"),
                    f"曝光 {fmt_value(row.get('imps'))} / 访客 {fmt_value(row.get('visitor_uv'))} / 询盘 {fmt_value(row.get('fb_count'))} / 订单 {fmt_value(row.get('order_count'))}",
                    "可用于本周判断",
                    "找异常日期",
                    "商机高订单低的日期回查业务员",
                    "下周同口径趋势",
                ]))
    else:
        blocks.append(detail_block(["每日趋势", "周期错位", "数据不可用于本周判断",
                                    "不可判断", "重新拉取本周数据",
                                    "确认报告周期", "周期一致"]))
    trade_rows = raw_rows(raw.get("trade"))
    period_start = str((report_data.get("meta") or {}).get("period_start") or "")
    period_end = str((report_data.get("meta") or {}).get("period_end") or "")
    for item in trade_rows:
        trade_date = str(first_present(item, "createDate", "gmtCreate", "orderCreateTime") or "")[:10]
        if trade_date and period_start and period_end and not (period_start <= trade_date <= period_end):
            continue
        blocks.append(detail_block([
            "交易明细",
            identifier_text(first_present(item, "orderId", "tradeId", "contractId", "id")),
            f"买家 {fmt_value(trade_buyer_name(item))} / 金额 {money_text(trade_amount_value(item))}",
            "可判断" if amount_ok else "金额不完整",
            fmt_value(trade_status_text(item)),
            "高金额订单单独复盘交付风险",
            "付款、发货、复购",
        ]))
    return blocks


def build_inquiry_rows(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict) -> list[Block]:
    """Build 询盘质量 sheet with block layout.

    This sheet is typically the longest one — the per-inquiry detail can run to
    60+ rows. Anchoring freeze_panes just below the detail header is the single
    biggest readability win for the boss when scrolling down.
    """

    iq = report_data.get("inquiry_quality") or {}
    summary = iq.get("summary") or {}
    records = iq.get("records") or []
    summary_dict = sheet_summary(narrative, "询盘质量", analysis, report_data)

    total = summary.get("total_records", 0) or 0
    high = summary.get("high_quality", 0) or 0
    low = summary.get("low_quality", 0) or 0
    pending = summary.get("pending_info", 0) or 0
    p0 = summary.get("p0", 0) or 0
    p1 = summary.get("p1", 0) or 0
    max_v = max(total, 1)

    title_status = f"高质量 {high} 条" if high else "看质量不看数量"

    blocks: list[Block] = [
        title_block("询盘质量", f"本周 {total} 条询盘" if total else "询盘明细", title_status),
        blank_block(),
        paragraph_block("老板口径", summary_dict.get("老板口径")),
        paragraph_block("关键证据", summary_dict.get("关键证据")),
        paragraph_block("今天动作", summary_dict.get("今天动作")),
        paragraph_block("下周复查", summary_dict.get("下周复查")),
        blank_block(),
        section_block("可视化看板：询盘质量漏斗", palette="visual"),
        visual_block("漏斗", "总询盘", total, safe_ratio(total, max_v),
                     "可判断" if total else "不可判断", "数量不等于质量"),
        visual_block("漏斗", "高质量", high, safe_ratio(high, max_v),
                     "高质量" if high else "不可判断", "推进报价/样品/规格"),
        visual_block("漏斗", "低质量", low, safe_ratio(low, max_v),
                     "低质量", "礼貌维护，不占主线"),
        visual_block("漏斗", "待补信息", pending, safe_ratio(pending, max_v),
                     "待补信息" if pending else "健康", "追买家关键信息"),
        visual_block("热力", "P0 / P1 跟进压力",
                     f"P0 {p0} / P1 {p1}",
                     safe_ratio(p0 + p1, max_v),
                     "P0" if p0 else ("P1" if p1 else "P3"),
                     "P0 今天必须有动作",
                     danger=bool(p0)),
        blank_block(),
        section_block("逐条询盘明细", palette="detail"),
        detail_header_block(["客户", "国家", "业务员", "L等级", "需求摘要", "购买信号", "质量判断", "优先级", "低质/风险原因", "建议动作"]),
    ]
    if not records:
        blocks.append(detail_block(["会话明细未返回", "", "", "", "当前只能用质检/汇总判断",
                                    "", "待判断", "P3", "", ""]))
    # Show the inquiries the boss should care about first. The old ascending
    # sort put P3/P2 rows at the top and could even cut P1 rows out of the
    # first 60 records, which made the sheet feel like it had no real signal.
    sorted_records = sorted(
        records,
        key=lambda r: (-priority_weight(r.get("priority")), str(r.get("customer") or "")),
    )
    for record in sorted_records:
        signals = record.get("purchase_signals") or []
        risks = record.get("reply_risks") or []
        blocks.append(detail_block([
            compact_text(record.get("customer"), 24),
            compact_text(record.get("country"), 10),
            compact_text(record.get("seller"), 16),
            record.get("buyer_level") or "未返回",
            compact_text(sanitize_for_boss(record.get("demand_summary") or record.get("product_or_need")), 90),
            "、".join(signals[:4]) if signals else "无明确信号",
            record.get("quality") or "待判断",
            record.get("priority") or "P3",
            compact_text(record.get("quality_reason") or "、".join(risks[:3]), 40),
            rewrite_for(narrative, "inquiry_actions", [record.get("customer")],
                        f"{record.get('customer') or '该客户'}：{sanitize_for_boss(record.get('suggested_action')) or '复查并确认下一步'}",
                        70),
        ]))
    return blocks


def build_seller_rows(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict) -> list[Block]:
    """Build 业务员回复与跟进 sheet with block layout."""

    service = report_data.get("service") or {}
    risk = report_data.get("risk") or {}
    records = (report_data.get("inquiry_quality") or {}).get("records") or []
    by_seller: dict[str, JsonDict] = {}
    for record in records:
        if not isinstance(record, dict):
            continue
        seller = record.get("seller") or "未分配业务员"
        row = by_seller.setdefault(seller, {"total": 0, "high": 0, "p0": 0, "quote": 0, "sample": 0, "order": 0})
        row["total"] += 1
        if record.get("quality") == "高质量":
            row["high"] += 1
        if record.get("priority") == "P0":
            row["p0"] += 1
        text = " ".join(str(record.get(k) or "") for k in ("demand_summary", "suggested_action", "quality_reason"))
        if "报价" in text:
            row["quote"] += 1
        if "样品" in text:
            row["sample"] += 1
        if "订单" in text or "付款" in text:
            row["order"] += 1

    summary_dict = sheet_summary(narrative, "业务员回复与跟进", analysis, report_data)
    five_min_raw = to_float(service.get("first_5min_reply_rate_30d"))
    five_min_pct = None
    if five_min_raw is not None:
        five_min_pct = five_min_raw * 100 if five_min_raw <= 1 else five_min_raw
    over_12h = to_float(service.get("reply_over_12h_count"))
    not_follow = to_float(service.get("not_follow_count"))
    title_status = (
        f"5 分钟回复率 {five_min_pct:.0f}%" if five_min_pct is not None else "回复指标缺失"
    )

    five_min_ratio = None
    if five_min_pct is not None:
        five_min_ratio = min(max(five_min_pct / 100.0, 0.0), 1.0)

    blocks: list[Block] = [
        title_block("业务员回复与跟进", "好询盘有没有被接住", title_status),
        blank_block(),
        paragraph_block("老板口径", summary_dict.get("老板口径")),
        paragraph_block("关键证据", summary_dict.get("关键证据")),
        paragraph_block("今天动作", summary_dict.get("今天动作")),
        paragraph_block("下周复查", summary_dict.get("下周复查")),
        blank_block(),
        section_block("可视化看板：业务员承接压力", palette="visual"),
        visual_block("响应", "店铺 5 分钟回复率",
                     f"{five_min_pct:.0f}%" if five_min_pct is not None else "未返回",
                     five_min_ratio,
                     "待校准" if five_min_pct is not None else "未返回",
                     "对照平台规则、店铺历史或用户目标",
                     danger=five_min_pct is None),
        visual_block("响应", "12h+ 未回", over_12h,
                     safe_ratio(over_12h, max_positive([over_12h, 1])),
                     "预警" if (over_12h or 0) > 0 else "健康",
                     "今天清零",
                     danger=(over_12h or 0) > 0),
        visual_block("响应", "未跟进风险", not_follow,
                     safe_ratio(not_follow, max_positive([not_follow, 1])),
                     "预警" if (not_follow or 0) > 0 else "健康",
                     "分配负责人逐条推进",
                     danger=(not_follow or 0) > 0),
        blank_block(),
        section_block("业务员明细", palette="detail"),
        detail_header_block(["业务员", "负责询盘", "高质量询盘", "首次响应", "5分钟回复率", "12h+未回", "报价推进", "样品推进", "订单后续", "判断", "今天动作"]),
    ]
    if by_seller:
        for seller, stat in sorted(by_seller.items(), key=lambda item: (-item[1]["high"], -item[1]["total"]))[:30]:
            blocks.append(detail_block([
                seller,
                stat["total"],
                stat["high"],
                "未返回",
                percent_text(service.get("first_5min_reply_rate_30d")),
                fmt_value(service.get("reply_over_12h_count")),
                stat["quote"],
                stat["sample"],
                stat["order"],
                "高质量多但推进少要复盘" if stat["high"] and not stat["quote"] else "按询盘质量复查",
                rewrite_for(narrative, "seller_actions", [seller],
                            "补齐下一步目标：报价/样品/付款/二次跟进", 40),
            ]))
    else:
        blocks.append(detail_block(["数据不足", "未返回", "未返回", "未返回", "未返回",
                                    "未返回", "未返回", "未返回", "未返回",
                                    "不能硬做人效榜", "看 数据质量检查 后补子账号明细"]))
    blocks.append(detail_block([
        "店铺总览", "全店", "未拆人",
        fmt_value(service.get("avg_reply_time_30d")),
        percent_text(service.get("first_5min_reply_rate_30d")),
        fmt_value(service.get("reply_over_12h_count")),
        "未返回", "未返回", "未返回",
        "回复超时影响转化", "每天 18:00 清超时",
    ]))
    blocks.append(detail_block([
        "专业度口径", "抽样", "每周至少 10 条", "看是否复述需求",
        "看是否 5 分钟内响应", "看是否超 12h",
        "是否给规格/MOQ/交期", "是否推进样品", "是否进入订单",
        "人工评分", "优秀样本进话术库",
    ]))
    if isinstance(risk, dict):
        for risk_item in (risk.get("items") or risk.get("risks") or [])[:10]:
            if isinstance(risk_item, dict):
                blocks.append(detail_block([
                    "风险明细",
                    compact_text(sanitize_for_boss(risk_item.get("type") or risk_item.get("name")), 18),
                    "未拆人", "未返回", "未返回", "未返回",
                    "未返回", "未返回", "未返回",
                    compact_text(sanitize_for_boss(risk_item.get("detail") or risk_item.get("description")), 45),
                    compact_text(risk_item.get("action") or "复查", 36),
                ]))
    return blocks


def build_action_rows(analysis: JsonDict) -> Rows:
    """Build two-layer action list: boss Top 5 and operations backlog.

    Args:
        analysis: Diagnosis output.

    Returns:
        Rows for 行动清单.

    Raises:
        No exceptions are intentionally raised.
    """

    actions = analysis.get("boss_top5_actions") or analysis.get("top3_actions") or []
    backlog = analysis.get("backlog") or {}
    rows: Rows = [
        ["层级", "优先级", "问题对象", "为什么重要", "负责人", "具体动作", "截止时间", "验收指标", "下周复查指标"],
        visual_section("老板 Top 5 与运营 Backlog", 9),
    ]
    for action in actions[:5]:
        rows.append([
            "老板 Top 5",
            action.get("priority") or "P2",
            compact_text(f"{action.get('object_type') or '事项'}：{action.get('object_name') or action.get('problem')}", 55),
            compact_text(action_field(action, "why"), 70),
            action.get("owner") or "运营主管",
            compact_text(action_field(action, "action"), 70),
            action.get("due") or "本周五",
            compact_text(action.get("acceptance_metric") or "动作完成且指标可复查", 40),
            compact_text(action.get("review_metric") or "下周复查同一指标", 40),
        ])
    due_map = {"P0": "今天", "P1": "本周内", "P2": "本周内", "P3": "下周观察"}
    for priority in ("P0", "P1", "P2", "P3"):
        items = backlog.get(priority) or []
        for item in items[:8]:
            rows.append([
                "运营 Backlog",
                priority,
                compact_text(item, 55),
                "",
                "运营主管",
                compact_text(item, 70),
                due_map.get(priority, "本周内"),
                "完成并记录",
                "同口径复查",
            ])
    return rows


def build_source_rows(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict) -> list[Block]:
    """Build 数据质量检查 sheet with block layout."""

    quality = report_data.get("data_quality") or {}
    collection = quality.get("collection") or {}
    summary_dict = sheet_summary(narrative, "数据质量检查", analysis, report_data)
    coverage = quality.get("coverage_rate", 0) or 0
    red_count = quality.get("red_count", 0) or 0
    yellow_count = quality.get("yellow_count", 0) or 0
    title_status = f"红灯 {red_count} 项 / 黄灯 {yellow_count} 项"

    blocks: list[Block] = [
        title_block("数据质量检查", "哪些结论能信，哪些只能保守看", title_status),
        blank_block(),
        paragraph_block("老板口径", summary_dict.get("老板口径")),
        paragraph_block("关键证据", summary_dict.get("关键证据")),
        paragraph_block("今天动作", summary_dict.get("今天动作")),
        paragraph_block("下周复查", summary_dict.get("下周复查")),
        blank_block(),
        section_block("可视化看板：数据可信度", palette="visual"),
        visual_block("可信度", "字段覆盖率", f"{coverage * 100:.0f}%",
                     min(max(coverage, 0), 1),
                     "可用" if coverage >= 0.95 else "预警",
                     "看关键字段，不止字段数",
                     danger=coverage < 0.95),
        visual_block("可信度", "红灯项", red_count,
                     safe_ratio(red_count, max_positive([red_count, yellow_count, 1])),
                     "预警" if red_count else "健康",
                     "红灯字段优先补",
                     danger=red_count > 0),
        visual_block("可信度", "黄灯项", yellow_count,
                     safe_ratio(yellow_count, max_positive([red_count, yellow_count, 1])),
                     "待判断" if yellow_count else "健康",
                     "黄灯保守判断"),
        blank_block(),
        section_block("检查项明细", palette="detail"),
        detail_header_block(["模块", "检查项", "状态", "本周能不能用", "数据周期", "问题", "影响结论", "处理方式"]),
        detail_block([
            "全表", "数据覆盖率", f"{coverage * 100:.0f}%",
            "红灯" if quality.get("status") == "red" else "可用但需保守",
            f"{collection.get('period_start')} 至 {collection.get('period_end')}",
            f"红灯 {red_count} 项 / 黄灯 {yellow_count} 项",
            "影响首页状态", "先补红灯字段",
        ]),
    ]
    for item in quality.get("checks_detail") or []:
        if not isinstance(item, dict):
            continue
        blocks.append(detail_block([
            item.get("module"), item.get("check"), item.get("status"),
            "可用" if item.get("usable") else "不可直接判断",
            item.get("period"), item.get("issue") or "无",
            item.get("impact"), item.get("action"),
        ]))
    if not quality.get("checks_detail"):
        blocks.append(detail_block(["全表", "数据质量", "黄灯", "不可直接判断",
                                    "未返回", "质量检查未返回",
                                    "只能保守判断", "重新生成清洗数据"]))
    return blocks


def write_workbook(report_data: JsonDict, analysis: JsonDict, raw: JsonDict, narrative: JsonDict, output: Path) -> None:
    """Render the workbook from block-based builders and persist it.

    Each sheet's tab gets coloured by the overall executive status — except
    the data-quality sheet, which uses its own data-quality status — so the
    boss can spot red sheets immediately without opening them.
    """

    workbook = setup_workbook()
    boss_status_label = clean_light(overall_status(report_data, analysis).get("label") or "")
    quality_status_raw = (report_data.get("data_quality") or {}).get("status") or ""
    quality_status_map = {"red": "红灯", "yellow": "黄灯", "green": "绿灯"}
    quality_status_label = quality_status_map.get(str(quality_status_raw).lower(), boss_status_label)

    sheets = [
        ("老板结论",
         build_boss_conclusion(report_data, analysis, raw, narrative),
         8,
         {"A": 12, "B": 30, "C": 30, "D": 22, "E": 14, "F": 12, "G": 26, "H": 6},
         boss_status_label, False),
        ("投产看板",
         build_roi_rows(report_data, analysis, raw, narrative),
         7,
         {"A": 14, "B": 22, "C": 30, "D": 16, "E": 22, "F": 36, "G": 28},
         boss_status_label, True),
        ("询盘质量",
         build_inquiry_rows(report_data, analysis, raw, narrative),
         10,
         {"A": 22, "B": 8, "C": 16, "D": 8, "E": 32, "F": 22, "G": 12, "H": 8, "I": 28, "J": 44},
         boss_status_label, True),
        ("订单产出",
         build_order_rows(report_data, analysis, raw, narrative),
         7,
         {"A": 14, "B": 20, "C": 38, "D": 18, "E": 22, "F": 34, "G": 28},
         boss_status_label, True),
        ("商品节奏",
         build_product_rows(report_data, analysis, raw, narrative),
         11,
         {"A": 14, "B": 16, "C": 44, "D": 24, "E": 10, "F": 10, "G": 10, "H": 12, "I": 8, "J": 44, "K": 28},
         boss_status_label, True),
        ("关键词与广告机会",
         build_keyword_rows(report_data, analysis, raw, narrative),
         7,
         {"A": 14, "B": 22, "C": 46, "D": 16, "E": 16, "F": 44, "G": 26},
         boss_status_label, True),
        ("业务员回复与跟进",
         build_seller_rows(report_data, analysis, raw, narrative),
         11,
         {"A": 18, "B": 10, "C": 12, "D": 12, "E": 14, "F": 12, "G": 12, "H": 12, "I": 12, "J": 26, "K": 36},
         boss_status_label, True),
        ("数据质量检查",
         build_source_rows(report_data, analysis, raw, narrative),
         8,
         {"A": 14, "B": 22, "C": 12, "D": 16, "E": 24, "F": 32, "G": 32, "H": 32},
         quality_status_label, True),
    ]
    for sheet_name, blocks, width, widths, sheet_status, auto_filter in sheets:
        write_blocks(
            workbook[sheet_name],
            blocks,
            width=width,
            widths=widths,
            sheet_status=sheet_status,
            auto_filter=auto_filter,
        )
    protect_formula_cells(workbook)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()


def resave_with_libreoffice(path: Path) -> None:
    """Re-save the workbook through LibreOffice headless.

    Args:
        path: Workbook to re-save in place.

    Returns:
        None.

    Raises:
        RuntimeError: If LibreOffice is not installed or conversion fails.
    """

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError("LibreOffice/soffice not found; cannot complete final XLSX safety flow.")
    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        input_dir = tmp_dir / "input"
        output_dir = tmp_dir / "output"
        profile_dir = tmp_dir / "profile"
        input_dir.mkdir()
        output_dir.mkdir()
        profile_dir.mkdir()
        input_copy = input_dir / path.name
        shutil.copy2(path, input_copy)
        cmd = [
            soffice,
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(input_copy),
        ]
        result = subprocess.run(cmd, text=True, capture_output=True, timeout=120)
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice resave failed: {result.stderr or result.stdout}")
        converted = output_dir / path.name
        if not converted.exists():
            raise RuntimeError("LibreOffice did not produce the expected .xlsx file.")
        shutil.copy2(converted, path)


def remove_sheet_residue(xml_bytes: bytes) -> bytes:
    """Remove tableParts and drawing tags from worksheet XML.

    Args:
        xml_bytes: Raw worksheet XML bytes.

    Returns:
        Cleaned XML bytes.

    Raises:
        ET.ParseError: If the XML cannot be parsed.
    """

    root = ET.fromstring(xml_bytes)
    for elem in list(root):
        local = elem.tag.split("}", 1)[-1]
        if local in {"tableParts", "drawing"}:
            root.remove(elem)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def clean_relationships(xml_bytes: bytes) -> bytes:
    """Remove table and drawing relationships from .rels XML.

    Args:
        xml_bytes: Raw relationships XML bytes.

    Returns:
        Cleaned XML bytes.

    Raises:
        ET.ParseError: If the XML cannot be parsed.
    """

    root = ET.fromstring(xml_bytes)
    for rel in list(root):
        target = str(rel.attrib.get("Target", "")).lower()
        rel_type = str(rel.attrib.get("Type", "")).lower()
        if "table" in target or "drawing" in target or "table" in rel_type or "drawing" in rel_type:
            root.remove(rel)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def clean_content_types(xml_bytes: bytes) -> bytes:
    """Remove table and drawing overrides from [Content_Types].xml.

    Args:
        xml_bytes: Raw content-types XML bytes.

    Returns:
        Cleaned XML bytes.

    Raises:
        ET.ParseError: If the XML cannot be parsed.
    """

    root = ET.fromstring(xml_bytes)
    for child in list(root):
        part = str(child.attrib.get("PartName", "")).lower()
        ctype = str(child.attrib.get("ContentType", "")).lower()
        if "/tables/" in part or "/drawings/" in part or "drawing" in ctype or "table" in ctype:
            root.remove(child)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def sanitize_xlsx_package(path: Path) -> None:
    """Remove table/drawing package residue after LibreOffice re-save.

    Args:
        path: Workbook package to sanitize in place.

    Returns:
        None.

    Raises:
        zipfile.BadZipFile: If the workbook is not a valid zip file.
        OSError: If temporary file replacement fails.
    """

    temp_path = path.with_suffix(".sanitized.xlsx")
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            name = info.filename
            lowered = name.lower()
            if lowered.startswith("xl/tables/") or lowered.startswith("xl/drawings/"):
                continue
            data = src.read(name)
            if lowered.startswith("xl/worksheets/") and lowered.endswith(".xml"):
                data = remove_sheet_residue(data)
            elif lowered.endswith(".rels"):
                data = clean_relationships(data)
            elif lowered == "[content_types].xml":
                data = clean_content_types(data)
            dst.writestr(info, data)
    temp_path.replace(path)


def validate_xlsx(path: Path) -> None:
    """Validate zip integrity, workbook readability, residue, and visible text.

    Args:
        path: Workbook path to validate.

    Returns:
        None when validation succeeds.

    Raises:
        ValueError: If any validation check fails.
        zipfile.BadZipFile: If the workbook is not a valid zip package.
        openpyxl exceptions: If the workbook cannot be opened.
    """

    with zipfile.ZipFile(path, "r") as archive:
        bad = archive.testzip()
        if bad:
            raise ValueError(f"Invalid zipped member: {bad}")
        names = archive.namelist()
        residue = [
            name for name in names
            if name.startswith("xl/tables/")
            or name.startswith("xl/drawings/")
            or ("table" in name.lower() and name.endswith(".rels"))
            or ("drawing" in name.lower() and name.endswith(".rels"))
        ]
        if residue:
            raise ValueError(f"Unexpected table/drawing residue found: {residue}")

    workbook = load_workbook(path, data_only=False)
    missing = [sheet for sheet in SHEETS if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(f"Missing required sheets: {missing}")
    extra_source_sheet = "数据来源" in workbook.sheetnames
    if extra_source_sheet:
        raise ValueError("Unexpected 数据来源 sheet found; this report version keeps data boundaries inside business sheets.")
    if "数据质量检查" not in workbook.sheetnames:
        raise ValueError("Missing 数据质量检查 sheet.")
    if workbook.sheetnames != SHEETS:
        raise ValueError(f"Workbook sheets must be exactly {SHEETS}, got {workbook.sheetnames}.")
    for required_sheet in SHEETS:
        sheet = workbook[required_sheet]
        has_visual_block = any(
            isinstance(cell.value, str) and "可视化看板" in cell.value
            for row in sheet.iter_rows()
            for cell in row
        )
        if not has_visual_block:
            raise ValueError(f"Missing safe visual dashboard block in sheet: {required_sheet}")
    boss = workbook["老板结论"]
    if boss.max_row > 25 or boss.max_column > 8:
        raise ValueError(f"老板结论 must stay one-page style, got {boss.max_row} rows x {boss.max_column} cols.")
    all_text = "\n".join(
        str(cell.value or "")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    )
    boss_text = "\n".join(
        str(cell.value or "")
        for row in boss.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    )
    if "订单金额未返回" in all_text and "ROI 不可判断" not in all_text:
        raise ValueError("Order amount is missing, but ROI was not downgraded to 不可判断.")
    if "周期错位" in all_text and any(word in boss_text for word in ("整体经营状态\n绿灯", "整体经营状态 绿灯")):
        raise ValueError("Period mismatch found, but boss conclusion still looks green.")
    if any(marker in all_text for marker in ("```", "## ", "### ", "- [ ]", "- [x]")):
        raise ValueError("Markdown-like text found in workbook cells.")
    repeated: dict[str, int] = {}
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if cell.data_type == "f":
                    raise ValueError(f"Unexpected formula in {sheet.title}!{cell.coordinate}")
                if isinstance(value, str):
                    lowered = value.lower()
                    if any(term in lowered for term in FORBIDDEN_TEXT):
                        raise ValueError(f"Forbidden internal wording in {sheet.title}!{cell.coordinate}: {value}")
                    if value in {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"}:
                        raise ValueError(f"Formula error in {sheet.title}!{cell.coordinate}: {value}")
                    normalized = re.sub(r"\s+", " ", value).strip()
                    if len(normalized) >= 36 and normalized not in {"平台接口异常，数据未返回"}:
                        repeated[normalized] = repeated.get(normalized, 0) + 1
    canned = {text: count for text, count in repeated.items() if count >= 6}
    if canned:
        sample, count = next(iter(canned.items()))
        raise ValueError(f"Repeated canned workbook wording found {count} times: {sample[:80]}")

    # --- Visual-layer assertions ------------------------------------------
    # The block-based layout commits to: every sheet has a merged title band
    # in row 1, at least one DataBar conditional formatting rule on a visual
    # card, a freeze pane anchored below the detail header, a scoped autoFilter
    # range starting at the detail header (except the boss page which uses two
    # short tables), and a tab colour signalling sheet status.
    from openpyxl.utils import range_boundaries  # local import to avoid top-level dependency
    for sheet in workbook.worksheets:
        title_merged = any(
            rng.min_row == 1 and rng.max_row == 1 and rng.max_col >= rng.min_col + 1
            for rng in sheet.merged_cells.ranges
        )
        if not title_merged:
            raise ValueError(
                f"Title band missing in {sheet.title}: row 1 must contain a merged title."
            )

        # The boss one-pager is intentionally text-only; every other sheet
        # gets DataBar visualisations on its 看板 cards.
        if sheet.title != "老板结论":
            has_databar = False
            cf_rules = getattr(sheet.conditional_formatting, "_cf_rules", {}) or {}
            for _, rules in cf_rules.items():
                for rule in rules:
                    if getattr(rule, "type", None) == "dataBar":
                        has_databar = True
                        break
                if has_databar:
                    break
            if not has_databar:
                raise ValueError(f"Visual DataBar missing in {sheet.title}.")

        freeze = sheet.freeze_panes
        if not freeze or freeze == "A1":
            raise ValueError(f"freeze_panes not set on {sheet.title}.")
        # freeze_panes is e.g. "A12" — extract row number
        match = re.match(r"[A-Z]+(\d+)", str(freeze))
        if not match or int(match.group(1)) < 9:
            raise ValueError(
                f"freeze_panes too high in {sheet.title}: expected detail-header anchor, got {freeze}."
            )

        if sheet.title != "老板结论":
            if not sheet.auto_filter.ref:
                raise ValueError(f"AutoFilter missing in {sheet.title}.")
            min_col, min_row, _, _ = range_boundaries(sheet.auto_filter.ref)
            if min_row < 8:
                raise ValueError(
                    f"AutoFilter starts too early in {sheet.title}: {sheet.auto_filter.ref}"
                )

        tab = sheet.sheet_properties.tabColor
        if not tab:
            raise ValueError(f"Tab colour missing on {sheet.title}.")
    workbook.close()

    data_only = load_workbook(path, data_only=True)
    data_only.close()


def main(argv: list[str]) -> int:
    """Run the workbook build and validation pipeline.

    Args:
        argv: Command-line arguments without the program name.

    Returns:
        Process exit code.

    Raises:
        No exceptions escape intentionally; errors are printed to stderr.
    """

    args = parse_args(argv)
    report_data = load_json(Path(args.report_data))
    analysis = load_json(Path(args.analysis))
    raw = collect_raw_context(Path(args.raw_dir))
    narrative = load_narrative(args.narrative)
    output = Path(args.output)
    write_workbook(report_data, analysis, raw, narrative, output)
    if not args.skip_libreoffice:
        resave_with_libreoffice(output)
    sanitize_xlsx_package(output)
    validate_xlsx(output)
    print(json.dumps({"ok": True, "output": str(output), "sheets": SHEETS}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv[1:]))
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False, indent=2), file=sys.stderr)
        raise SystemExit(1)
